"""
input check pin interm 시트 생성 모듈
DE requirement 시트에서 address 데이터를 가져와서 input check pin interm 시트를 생성합니다.
NET 파일에서 PIECE 정보를 읽어 그룹 헤더를 생성합니다.
"""

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

from logic.file_reader import find_piece_lines


def auto_adjust_column_width(worksheet):
    """
    워크시트의 각 컬럼 너비를 내용에 맞게 자동 조절
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max(max_length + 2, 8), 100)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def apply_header_style(worksheet, row_number, fill_color="ED7D31", start_col=1, end_col=None):
    """
    지정된 행에 헤더 스타일 적용 (주황색 배경)
    """
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True)
    
    if end_col is None:
        end_col = worksheet.max_column
    
    for col in range(start_col, end_col + 1):
        cell = worksheet.cell(row=row_number, column=col)
        cell.fill = fill
        cell.font = font


def make_input_check_pin_sheet(outfile_path: str, net_file_path: str = None) -> str:
    """
    DE requirement 시트에서 address 데이터를 가져와서 input check pin interm 시트를 생성합니다.
    NET 파일에서 PIECE 정보를 읽어 그룹을 구성합니다.
    
    Args:
        outfile_path: 출력 파일 경로 (DE requirement 시트가 이미 있는 파일)
        net_file_path: NET 파일 경로 (PIECE 정보를 읽기 위함)
    
    Returns:
        결과 메시지
    """
    try:
        debug_info = []
        
        # 출력 파일 열기
        if not os.path.exists(outfile_path):
            return f"Error: Output file not found: {outfile_path}"
        
        wb_out = load_workbook(outfile_path)
        
        # DE requirement 시트 확인
        if "DE requirement" not in wb_out.sheetnames:
            wb_out.close()
            return "Error: 'DE requirement' sheet not found. Please run 'Make DE Requirement' first."
        
        ws_de = wb_out["DE requirement"]
        
        # DE requirement에서 part 이름 가져오기 (C, E 컬럼 헤더에서)
        # 실제 part 이름은 데이터 첫 행에서 가져옴
        part1_name = ws_de.cell(row=3, column=3).value or "J_TELE"  # C3
        part2_name = ws_de.cell(row=3, column=5).value or "U0200"   # E3
        
        debug_info.append(f"Part names: {part1_name}, {part2_name}")
        
        # DE requirement에서 address 데이터 읽기 (I-L 컬럼, 9-12)
        address_data = []
        
        # part1.1과 part2.1의 최소 address 찾기
        part1_min_addr = None
        part2_min_addr = None
        
        # 데이터 읽기 (3행부터)
        for row_idx in range(3, ws_de.max_row + 1):
            add1 = ws_de.cell(row=row_idx, column=9).value   # J_TELE_add1
            add2 = ws_de.cell(row=row_idx, column=10).value  # J_TELE_add2
            add3 = ws_de.cell(row=row_idx, column=11).value  # U0200_add1
            add4 = ws_de.cell(row=row_idx, column=12).value  # U0200_add2
            
            # 최소 address 찾기
            if add1 is not None:
                try:
                    val = int(add1)
                    if part1_min_addr is None or val < part1_min_addr:
                        part1_min_addr = val
                except (ValueError, TypeError):
                    pass
            if add2 is not None:
                try:
                    val = int(add2)
                    if part1_min_addr is None or val < part1_min_addr:
                        part1_min_addr = val
                except (ValueError, TypeError):
                    pass
            if add3 is not None:
                try:
                    val = int(add3)
                    if part2_min_addr is None or val < part2_min_addr:
                        part2_min_addr = val
                except (ValueError, TypeError):
                    pass
            if add4 is not None:
                try:
                    val = int(add4)
                    if part2_min_addr is None or val < part2_min_addr:
                        part2_min_addr = val
                except (ValueError, TypeError):
                    pass
            
            # 데이터가 있는 행만 추가
            if add1 is not None or add2 is not None or add3 is not None or add4 is not None:
                address_data.append([add1, add2, add3, add4])
        
        debug_info.append(f"Found {len(address_data)} rows of address data")
        debug_info.append(f"Min addr: {part1_name}={part1_min_addr}, {part2_name}={part2_min_addr}")
        
        # NET 파일에서 PIECE 정보 읽기
        piece_groups = []
        if net_file_path and os.path.exists(net_file_path):
            piece_lines = find_piece_lines(net_file_path)
            for line, nums in piece_lines:
                if len(nums) >= 4:
                    # nums = [J_TELE_start, J_TELE_end, U0200_start, U0200_end, ...]
                    piece_groups.append(nums[:4])  # 처음 4개만 사용
            debug_info.append(f"Found {len(piece_groups)} PIECE groups")
        else:
            debug_info.append("No NET file provided or file not found")
        
        # 기존 input check pin interm 시트가 있으면 삭제
        if "input check pin interm" in wb_out.sheetnames:
            del wb_out["input check pin interm"]
        
        # 새 input check pin interm 시트 생성
        ws_input = wb_out.create_sheet("input check pin interm")
        
        # === Row 1: 헤더 정보 ===
        # A1-D1: "Input: Start Pin number each side" (4칸 병합)
        ws_input.merge_cells('A1:D1')
        ws_input.cell(row=1, column=1, value="Input: Start Pin number each side")
        ws_input.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        
        # E1-F1: part1_name (2칸 병합)
        ws_input.merge_cells('E1:F1')
        ws_input.cell(row=1, column=5, value=part1_name)
        ws_input.cell(row=1, column=5).alignment = Alignment(horizontal='center')
        
        # G1-H1: part2_name (2칸 병합)
        ws_input.merge_cells('G1:H1')
        ws_input.cell(row=1, column=7, value=part2_name)
        ws_input.cell(row=1, column=7).alignment = Alignment(horizontal='center')
        
        # Row 1 스타일 적용 (주황색)
        apply_header_style(ws_input, 1, fill_color="ED7D31", end_col=8)
        
        # === Row 2: 시작 핀 번호 (DE requirement에서 가져온 최소 address) ===
        ws_input.cell(row=2, column=5, value=part1_min_addr if part1_min_addr else 1)
        ws_input.cell(row=2, column=7, value=part2_min_addr if part2_min_addr else 2049)
        
        # === Row 3: 비워두기 ===
        
        # === A4-A9: 라벨 추가 ===
        # A4: 빈칸, A5: 빈칸, A6: Address, A7: Space, A8: Judge, A9: Select
        ws_input.cell(row=4, column=1, value="")
        ws_input.cell(row=5, column=1, value="")
        ws_input.cell(row=6, column=1, value="Address")
        ws_input.cell(row=7, column=1, value="Space")
        ws_input.cell(row=8, column=1, value="Judge")
        ws_input.cell(row=9, column=1, value="Select")
        
        # === Row 4-9: PIECE 그룹 데이터 (B열부터 시작) ===
        num_groups = len(piece_groups) if piece_groups else 1
        
        # Row 4: Group 헤더 (Group 1, Group 2, ...) - B4부터 시작
        col = 2  # B열부터 시작
        for g in range(num_groups):
            # 4개 컬럼 병합
            start_col_letter = get_column_letter(col)
            end_col_letter = get_column_letter(col + 3)
            ws_input.merge_cells(f'{start_col_letter}4:{end_col_letter}4')
            ws_input.cell(row=4, column=col, value=f"Group {g + 1}")
            ws_input.cell(row=4, column=col).alignment = Alignment(horizontal='center')
            col += 4
        
        # Row 4 스타일 적용
        apply_header_style(ws_input, 4, fill_color="ED7D31", start_col=2, end_col=1 + num_groups * 4)
        
        # Row 5: 각 그룹의 서브헤더 (part1, part2) - B5부터 시작
        col = 2  # B열부터 시작
        for g in range(num_groups):
            # part1 (2칸 병합)
            start_col_letter = get_column_letter(col)
            end_col_letter = get_column_letter(col + 1)
            ws_input.merge_cells(f'{start_col_letter}5:{end_col_letter}5')
            ws_input.cell(row=5, column=col, value=part1_name)
            ws_input.cell(row=5, column=col).alignment = Alignment(horizontal='center')
            
            # part2 (2칸 병합)
            start_col_letter = get_column_letter(col + 2)
            end_col_letter = get_column_letter(col + 3)
            ws_input.merge_cells(f'{start_col_letter}5:{end_col_letter}5')
            ws_input.cell(row=5, column=col + 2, value=part2_name)
            ws_input.cell(row=5, column=col + 2).alignment = Alignment(horizontal='center')
            
            col += 4
        
        # Row 5 스타일 적용
        apply_header_style(ws_input, 5, fill_color="ED7D31", start_col=2, end_col=1 + num_groups * 4)
        
        # Row 6: PIECE 데이터 (각 그룹의 4개 숫자) - B6부터 시작
        # 배치: 각 그룹은 4개 컬럼
        if piece_groups:
            col = 2  # B열부터 시작
            for g, group_nums in enumerate(piece_groups):
                if len(group_nums) >= 4:
                    # 가로로 4개 숫자 배치 (Row 6에 한 줄로)
                    ws_input.cell(row=6, column=col, value=group_nums[0])      # J_TELE start
                    ws_input.cell(row=6, column=col + 1, value=group_nums[1])  # J_TELE end
                    ws_input.cell(row=6, column=col + 2, value=group_nums[2])  # U0200 start
                    ws_input.cell(row=6, column=col + 3, value=group_nums[3])  # U0200 end
                col += 4
        
        # === Row 7 (Space): Excel 수식 추가 ===
        # Group 2부터 시작 (각 열 - 4칸 앞의 열)
        # F7 = IF(F6="","",F6-B6), J7 = IF(J6="","",J6-F6), ...
        if num_groups > 1:
            for g in range(1, num_groups):  # Group 2부터
                group_start_col = 2 + g * 4  # Group 2는 col 6 (F)
                for offset in range(4):  # 각 그룹의 4개 컬럼
                    curr_col = group_start_col + offset
                    base_col = curr_col - 4  # 4칸 앞의 열 (바로 앞 그룹)
                    
                    curr_col_letter = get_column_letter(curr_col)
                    base_col_letter = get_column_letter(base_col)
                    
                    # 수식: IF(현재셀6="","",현재셀6-4칸앞셀6)
                    formula = f'=IF({curr_col_letter}6="","",{curr_col_letter}6-{base_col_letter}6)'
                    ws_input.cell(row=7, column=curr_col, value=formula)
        
        # === Row 8 (Judge): Excel 수식 추가 (2칸씩 병합) ===
        # F8:G8 = IF(OR(F7="",G7=""),"",IF(F7=G7,"OK","NG"))
        # H8:I8 = IF(OR(H7="",I7=""),"",IF(H7=I7,"OK","NG"))
        if num_groups > 1:
            for g in range(1, num_groups):  # Group 2부터
                group_start_col = 2 + g * 4  # Group 2는 col 6 (F)
                
                # J_TELE 부분 (2칸 병합: col, col+1)
                col1_letter = get_column_letter(group_start_col)
                col2_letter = get_column_letter(group_start_col + 1)
                ws_input.merge_cells(f'{col1_letter}8:{col2_letter}8')
                formula1 = f'=IF(OR({col1_letter}7="",{col2_letter}7=""),"",IF({col1_letter}7={col2_letter}7,"OK","NG"))'
                ws_input.cell(row=8, column=group_start_col, value=formula1)
                ws_input.cell(row=8, column=group_start_col).alignment = Alignment(horizontal='center')
                
                # U0200 부분 (2칸 병합: col+2, col+3)
                col3_letter = get_column_letter(group_start_col + 2)
                col4_letter = get_column_letter(group_start_col + 3)
                ws_input.merge_cells(f'{col3_letter}8:{col4_letter}8')
                formula2 = f'=IF(OR({col3_letter}7="",{col4_letter}7=""),"",IF({col3_letter}7={col4_letter}7,"OK","NG"))'
                ws_input.cell(row=8, column=group_start_col + 2, value=formula2)
                ws_input.cell(row=8, column=group_start_col + 2).alignment = Alignment(horizontal='center')
        
        # === Row 9 (Select): Excel 수식 추가 (2칸씩 병합) ===
        # F9:G9 = IF(F8="OK",F7,"")
        # H9:I9 = IF(H8="OK",H7,"")
        if num_groups > 1:
            for g in range(1, num_groups):  # Group 2부터
                group_start_col = 2 + g * 4  # Group 2는 col 6 (F)
                
                # J_TELE 부분 (2칸 병합: col, col+1)
                col1_letter = get_column_letter(group_start_col)
                col2_letter = get_column_letter(group_start_col + 1)
                ws_input.merge_cells(f'{col1_letter}9:{col2_letter}9')
                formula1 = f'=IF({col1_letter}8="OK",{col1_letter}7,"")'
                ws_input.cell(row=9, column=group_start_col, value=formula1)
                ws_input.cell(row=9, column=group_start_col).alignment = Alignment(horizontal='center')
                
                # U0200 부분 (2칸 병합: col+2, col+3)
                col3_letter = get_column_letter(group_start_col + 2)
                col4_letter = get_column_letter(group_start_col + 3)
                ws_input.merge_cells(f'{col3_letter}9:{col4_letter}9')
                formula2 = f'=IF({col3_letter}8="OK",{col3_letter}7,"")'
                ws_input.cell(row=9, column=group_start_col + 2, value=formula2)
                ws_input.cell(row=9, column=group_start_col + 2).alignment = Alignment(horizontal='center')
        
        # === Row 10: 비워두기 ===
        
        # === Row 11: 기존 데이터 헤더 ===
        ws_input.cell(row=11, column=1, value="NO")
        
        # B11-E11에 "Group 1" (4개 셀 병합)
        ws_input.merge_cells('B11:E11')
        ws_input.cell(row=11, column=2, value="Group 1")
        ws_input.cell(row=11, column=2).alignment = Alignment(horizontal='center')
        
        # Row 11 스타일 적용
        apply_header_style(ws_input, 11, fill_color="ED7D31", end_col=5)
        
        # === Row 12부터: 기존 address 데이터 ===
        for row_offset, row_data in enumerate(address_data):
            row_num = 12 + row_offset
            
            # NO 컬럼 (1부터 시작)
            ws_input.cell(row=row_num, column=1, value=row_offset + 1)
            
            # address 데이터 (B-E 컬럼)
            for col_offset, value in enumerate(row_data):
                ws_input.cell(row=row_num, column=2 + col_offset, value=value)
        
        # 모든 컬럼 너비를 10으로 고정
        max_col = 1 + num_groups * 4  # A열 + 그룹 수 * 4
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            ws_input.column_dimensions[col_letter].width = 10
        
        # 파일 저장
        wb_out.save(outfile_path)
        wb_out.close()
        
        result_msg = f"Success: Created 'input check pin interm' sheet with {len(address_data)} rows, {num_groups} groups\n"
        result_msg += "Debug: " + " | ".join(debug_info)
        return result_msg
        
    except Exception as e:
        import traceback
        return f"Error: {str(e)}\n{traceback.format_exc()}"
