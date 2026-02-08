"""
Form Measurement Result 파일 생성 모듈
Form measurement result files_form.xlsx 템플릿을 기반으로 결과 파일 생성
DK 파일들에서 TDR 데이터를 읽어서 Impedance NET resistance 행에 채움
"""

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from copy import copy
import os
import sys
import re
import pandas as pd

from logic.visualizer import save_form_plots_from_workbook


def get_template_path() -> str:
    """
    템플릿 파일 경로 반환 (PyInstaller 지원)
    """
    if getattr(sys, 'frozen', False):
        # exe 실행 시: PyInstaller가 파일을 추출하는 임시 디렉토리(_MEIPASS) 확인
        # 만약 번들링하지 않고 exe와 같은 위치에 두는 경우를 위해 executable 경로도 확인
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(sys.executable))
    else:
        # 개발 중
        base_path = os.path.dirname(os.path.dirname(__file__))
    
    template_name = "Form measurement result files_form.xlsx"
    template_path = os.path.join(base_path, template_name)
    
    # 만약 _MEIPASS에서 못 찾으면 exe 실행 위치에서 다시 확인
    if getattr(sys, 'frozen', False) and not os.path.exists(template_path):
        template_path = os.path.join(os.path.dirname(sys.executable), template_name)
        
    return template_path


def create_form_measurement_file(output_path: str) -> str:
    """
    Form measurement result 파일을 생성합니다.
    템플릿의 구조를 하드카피하여 새 파일 생성
    
    Args:
        output_path: 출력 파일 경로
        
    Returns:
        결과 메시지
    """
    try:
        template_path = get_template_path()
        
        if not os.path.exists(template_path):
            return f"Error: Template file not found: {template_path}"
        
        # 템플릿 파일 열기
        wb_template = load_workbook(template_path)
        ws_template = wb_template.active
        
        debug_info = []
        debug_info.append(f"Template: {ws_template.title}")
        debug_info.append(f"Size: {ws_template.max_row} rows x {ws_template.max_column} cols")
        
        # 새 워크북 생성
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Physical Analysis"
        
        # === Row 1-2 (헤더) 하드카피 ===
        # 병합 셀 정보 먼저 복사
        for merged_range in ws_template.merged_cells.ranges:
            ws_out.merge_cells(str(merged_range))
        
        debug_info.append(f"Merged cells: {len(list(ws_template.merged_cells.ranges))}")
        
        # 모든 데이터 및 스타일 복사
        for row in range(1, ws_template.max_row + 1):
            for col in range(1, ws_template.max_column + 1):
                source_cell = ws_template.cell(row=row, column=col)
                target_cell = ws_out.cell(row=row, column=col)
                
                try:
                    # 값 복사
                    target_cell.value = source_cell.value
                    
                    # 스타일 복사
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.border = copy(source_cell.border)
                        target_cell.alignment = copy(source_cell.alignment)
                        target_cell.number_format = source_cell.number_format
                except AttributeError:
                    # MergedCell인 경우 스킵
                    pass
        
        # 컬럼 너비 복사
        for col in range(1, ws_template.max_column + 1):
            col_letter = get_column_letter(col)
            if ws_template.column_dimensions[col_letter].width:
                ws_out.column_dimensions[col_letter].width = ws_template.column_dimensions[col_letter].width
        
        # 행 높이 복사
        for row in range(1, ws_template.max_row + 1):
            if ws_template.row_dimensions[row].height:
                ws_out.row_dimensions[row].height = ws_template.row_dimensions[row].height
        
        wb_template.close()
        
        # 파일 저장
        wb_out.save(output_path)
        wb_out.close()
        
        result_msg = f"Success: Created Form Measurement Result file\n"
        result_msg += f"Output: {output_path}\n"
        result_msg += "Debug: " + " | ".join(debug_info)
        return result_msg
        
    except Exception as e:
        import traceback
        return f"Error: {str(e)}\n{traceback.format_exc()}"


def get_inner_value_from_filename(filename: str) -> str:
    """
    파일명에서 inner 값 추출
    예: DK1.5.xls -> 1.5, DK CENTER.xls -> CENTER
    """
    basename = os.path.splitext(os.path.basename(filename))[0]
    # DK 제거
    value = basename.replace("DK", "").replace("dk", "").strip()
    return value


def read_tdr_data_from_dk_file(file_path: str) -> list:
    """
    DK 파일의 'Form kq' 시트에서 TDR 데이터를 읽음
    
    Args:
        file_path: DK 파일 경로 (.xls)
        
    Returns:
        TDR 값 리스트 (1-32)
    """
    try:
        df = pd.read_excel(file_path, sheet_name='Form kq', header=None)
        
        # STT와 TDR 컬럼 찾기 (Row 10에서)
        tdr_col = None
        for col in range(df.shape[1]):
            val = df.iloc[10, col] if 10 < df.shape[0] else None
            if pd.notna(val) and str(val).upper() == 'TDR':
                tdr_col = col
                break
        
        if tdr_col is None:
            return []
        
        # TDR 데이터 추출 (Row 12부터 32개)
        tdr_data = []
        for row in range(12, min(44, df.shape[0])):  # Row 12-43 (32개)
            val = df.iloc[row, tdr_col]
            if pd.notna(val) and isinstance(val, (int, float)):
                tdr_data.append(val)
        
        return tdr_data
        
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return []


def get_dk_files_in_directory(etching_dir: str) -> list:
    """
    etching 디렉토리에서 DK 파일들의 목록 반환
    
    Args:
        etching_dir: etching 디렉토리 경로
        
    Returns:
        [(inner_value, file_path), ...] 리스트
    """
    dk_files = []
    
    if not os.path.exists(etching_dir):
        return dk_files
    
    for filename in os.listdir(etching_dir):
        if filename.upper().startswith('DK') and filename.endswith('.xls'):
            inner_val = get_inner_value_from_filename(filename)
            file_path = os.path.join(etching_dir, filename)
            dk_files.append((inner_val, file_path))
    
    return dk_files


def fill_impedance_data(output_path: str, etching_dir: str) -> str:
    """
    DK 파일들에서 TDR 데이터를 읽어서 
    Form Measurement Result 파일의 Impedance NET resistance 행에 채움
    
    Args:
        output_path: 출력 파일 경로
        etching_dir: etching 디렉토리 경로
        
    Returns:
        결과 메시지
    """
    try:
        if not os.path.exists(output_path):
            return f"Error: Output file not found: {output_path}"
        
        if not os.path.exists(etching_dir):
            return f"Error: Etching directory not found: {etching_dir}"
        
        debug_info = []
        
        # DK 파일 목록 가져오기
        dk_files = get_dk_files_in_directory(etching_dir)
        if not dk_files:
            return f"Error: No DK files found in {etching_dir}"
        
        debug_info.append(f"Found {len(dk_files)} DK files")
        
        # 출력 파일 열기
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Inner 값과 행 매핑 찾기
        # B열에서 Inner 값을 찾아서 해당 그룹의 시작 행 결정
        inner_to_row = {}  # inner_value -> row_number (Impedance NET resistance 행)
        
        for row in range(3, ws.max_row + 1):
            inner_val = ws.cell(row=row, column=2).value  # B열 (Inner)
            content_val = ws.cell(row=row, column=4).value  # D열 (Contents)
            
            if inner_val and content_val:
                inner_str = str(inner_val).strip()
                content_str = str(content_val).lower()
                
                # Impedance NET resistance 행 찾기
                if 'impedance' in content_str and 'resistance' in content_str:
                    inner_to_row[inner_str] = row
        
        debug_info.append(f"Inner mappings: {list(inner_to_row.keys())}")
        
        # 각 DK 파일 처리
        processed = 0
        # 시각화를 위한 TDR 모음
        tdr_map = {}
        for inner_val, file_path in dk_files:
            tdr_data = read_tdr_data_from_dk_file(file_path)
            
            if not tdr_data:
                debug_info.append(f"No TDR data in {os.path.basename(file_path)}")
                continue
            
            # Inner 값으로 해당 행 찾기
            target_row = inner_to_row.get(inner_val)
            
            if target_row:
                # E열(5)부터 데이터 채우기 (Impedance NET resistance 행)
                for idx, val in enumerate(tdr_data):
                    col = 5 + idx  # E=5, F=6, ...
                    ws.cell(row=target_row, column=col, value=val)
                # 시각화용 저장
                tdr_map[inner_val] = tdr_data
                
                # 각 Inner 그룹의 4개 행 모두에 수식 추가
                # Row: Impedance NET resistance (target_row)
                # Row+1: Impedance NET Circuit width
                # Row+2: Impedance NET thickness
                # Row+3: Minimum NET Circuit width
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 초록색
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # 빨간색
                
                for row_offset in range(4):  # 0, 1, 2, 3
                    row = target_row + row_offset
                    
                    # Min = MIN(E{row}:AJ{row})
                    ws.cell(row=row, column=40, value=f"=MIN(E{row}:AJ{row})")
                    
                    # Max = MAX(E{row}:AJ{row})
                    ws.cell(row=row, column=41, value=f"=MAX(E{row}:AJ{row})")
                    
                    # Ave = AVERAGE(E{row}:AJ{row})
                    ws.cell(row=row, column=42, value=f"=AVERAGE(E{row}:AJ{row})")
                    
                    # Judge1: Min >= LSL -> =IF(AN{row}>=AK{row},"OK","NG")
                    ws.cell(row=row, column=43, value=f'=IF(AN{row}>=AK{row},"OK","NG")')
                    
                    # Judge2: Max <= USL -> =IF(AO{row}<=AM{row},"OK","NG")
                    ws.cell(row=row, column=44, value=f'=IF(AO{row}<=AM{row},"OK","NG")')
                    
                    # Judge3: 둘 다 만족 -> =IF(AND(AN{row}>=AK{row},AO{row}<=AM{row}),"OK","NG")
                    ws.cell(row=row, column=45, value=f'=IF(AND(AN{row}>=AK{row},AO{row}<=AM{row}),"OK","NG")')
                    
                    # 조건부 서식 추가 - AS열(Judge3) 값에 따라 전체 행(A~AS) 색상 적용
                    # A열(1)부터 AS열(45)까지 전체 행에 조건부 서식 적용
                    row_range = f"A{row}:AS{row}"
                    
                    # AS열(Judge3)이 OK이면 전체 행 초록색
                    ws.conditional_formatting.add(
                        row_range,
                        FormulaRule(formula=[f'$AS{row}="OK"'], fill=green_fill)
                    )
                    
                    # AS열(Judge3)이 NG이면 전체 행 빨간색
                    ws.conditional_formatting.add(
                        row_range,
                        FormulaRule(formula=[f'$AS{row}="NG"'], fill=red_fill)
                    )
                
                processed += 1
                debug_info.append(f"{os.path.basename(file_path)} -> Rows {target_row}-{target_row+3} ({len(tdr_data)} values + formulas)")
            else:
                debug_info.append(f"No row found for inner={inner_val}")
        
        # 파일 저장
        wb.save(output_path)
        wb.close()
        
        result_msg = f"Success: Filled Impedance data from {processed} DK files\n"
        result_msg += "Debug:\n  " + "\n  ".join(debug_info)
        # 결과 메시지와 시각화 데이터 반환을 위해 dict 형태로 래핑
        return {"message": result_msg, "tdr_map": tdr_map}
        
    except Exception as e:
        import traceback
        return f"Error: {str(e)}\n{traceback.format_exc()}"


def fill_dimension_data(output_path: str, dimension_file: str, sheet_name: str = "") -> str:
    """
    Dimension 파일(7E3493-00003.xlsx 형태)에서 dimension 데이터를 읽어서
    Form Measurement Result 파일에 채움
    
    파일 구조:
    - Row 5에 DK 섹션 헤더 (DK 1.5, DK 1.6, ..., DK CENTER)
    - 각 DK 섹션은 3개 열 (TOP, BOTTOM, CIRCUIT HIGHT)
    - Row 34-43에 1-10번 데이터
    - BOTTOM 열: Impedance NET Circuit width (um)
    - CIRCUIT HIGHT 열: Impedance NET thickness (um)
    - Minimum NET Circuit width = Circuit width와 같은 값
    
    Args:
        output_path: 출력 파일 경로
        dimension_file: dimension 파일 경로
        sheet_name: 사용할 시트 이름 (빈 문자열이면 자동 선택)
        
    Returns:
        결과 메시지
    """
    try:
        if not os.path.exists(output_path):
            return f"Error: Output file not found: {output_path}"
        
        if not os.path.exists(dimension_file):
            return f"Error: Dimension file not found: {dimension_file}"
        
        debug_info = []
        
        # Dimension 파일 읽기
        if sheet_name:
            # 지정된 시트 사용
            try:
                df = pd.read_excel(dimension_file, sheet_name=sheet_name, header=None)
                debug_info.append(f"Using sheet: {sheet_name}")
            except:
                return f"Error: Sheet '{sheet_name}' not found in dimension file"
        else:
            # 시트 이름이 지정되지 않으면 B2 또는 첫 번째 시트 사용
            try:
                df = pd.read_excel(dimension_file, sheet_name='B2', header=None)
                debug_info.append("Using sheet: B2")
            except:
                df = pd.read_excel(dimension_file, sheet_name=0, header=None)
                debug_info.append("Using first sheet")
        
        debug_info.append(f"Dimension file shape: {df.shape}")
        
        # DK 섹션 위치 찾기 (Row 5에서 DK로 시작하는 셀 찾기)
        dk_sections = {}  # inner_value -> column_index (BOTTOM 열)
        for col in range(df.shape[1]):
            val = df.iloc[5, col] if 5 < df.shape[0] else None
            if pd.notna(val) and 'DK' in str(val).upper():
                # DK 값에서 inner 추출 (DK 1.5 -> 1.5, DK CENTER -> CENTER)
                inner_val = str(val).replace('DK', '').strip().upper()
                # BOTTOM 열은 DK 헤더 열 + 1, CIRCUIT HIGHT 열은 + 2
                dk_sections[inner_val] = col + 1  # BOTTOM 열
                debug_info.append(f"Found DK section: {val} at col {col}")
        
        if not dk_sections:
            return f"Error: No DK sections found in dimension file"
        
        # 출력 파일 열기
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Inner 값과 시작 행 매핑 찾기
        inner_to_row = {}
        for row in range(3, ws.max_row + 1):
            inner_val = ws.cell(row=row, column=2).value  # B열 (Inner)
            if inner_val:
                inner_str = str(inner_val).strip().upper()
                inner_to_row[inner_str] = row
        
        debug_info.append(f"Output file inner mappings: {list(inner_to_row.keys())}")
        
        # 데이터 행 범위 (Row 34-43 = 1-10번 데이터)
        data_start_row = 34
        data_end_row = 43
        
        processed = 0
        dim_map = {}
        
        for inner_val, bottom_col in dk_sections.items():
            # 해당 Inner의 시작 행 찾기
            target_base_row = inner_to_row.get(inner_val)
            if target_base_row is None:
                debug_info.append(f"No output row for inner={inner_val}")
                continue
            
            circuit_hight_col = bottom_col + 1  # CIRCUIT HIGHT 열
            
            # 1-10번 데이터 읽기
            circuit_width_data = []  # BOTTOM 열 데이터
            thickness_data = []  # CIRCUIT HIGHT 열 데이터
            
            for data_row in range(data_start_row, data_end_row + 1):
                # BOTTOM 열 (Circuit width)
                width_val = df.iloc[data_row, bottom_col] if data_row < df.shape[0] and bottom_col < df.shape[1] else None
                circuit_width_data.append(width_val if pd.notna(width_val) else None)
                
                # CIRCUIT HIGHT 열 (thickness)
                height_val = df.iloc[data_row, circuit_hight_col] if data_row < df.shape[0] and circuit_hight_col < df.shape[1] else None
                thickness_data.append(height_val if pd.notna(height_val) else None)
            
            # Circuit width 데이터 채우기 (Row: target_base_row + 1)
            target_row_width = target_base_row + 1
            for idx, val in enumerate(circuit_width_data):
                if val is not None:
                    ws.cell(row=target_row_width, column=5 + idx, value=val)  # E=5
            
            # Thickness 데이터 채우기 (Row: target_base_row + 2)
            target_row_thickness = target_base_row + 2
            for idx, val in enumerate(thickness_data):
                if val is not None:
                    ws.cell(row=target_row_thickness, column=5 + idx, value=val)  # E=5
            
            # Minimum Circuit width = Circuit width와 같은 값 (Row: target_base_row + 3)
            target_row_minimum = target_base_row + 3
            for idx, val in enumerate(circuit_width_data):
                if val is not None:
                    ws.cell(row=target_row_minimum, column=5 + idx, value=val)  # E=5
            
            processed += 1
            debug_info.append(f"Inner {inner_val}: filled width/thickness/minimum data")

            # 시각화용 평균값 저장
            valid_widths = [v for v in circuit_width_data if v is not None]
            valid_thicks = [v for v in thickness_data if v is not None]
            width_avg = float(pd.Series(valid_widths).mean()) if valid_widths else None
            thick_avg = float(pd.Series(valid_thicks).mean()) if valid_thicks else None
            if width_avg is not None and thick_avg is not None:
                dim_map[inner_val] = (width_avg, thick_avg)
        
        # 파일 저장
        wb.save(output_path)
        wb.close()
        
        result_msg = f"Success: Filled dimension data from {processed} DK sections\n"
        result_msg += "Debug:\n  " + "\n  ".join(debug_info)
        return {"message": result_msg, "dim_map": dim_map}
        
    except Exception as e:
        import traceback
        return f"Error: {str(e)}\n{traceback.format_exc()}"


def fill_lslusl_data(output_path: str, lslusl_file: str) -> str:
    """
    LSLUSL 파일에서 LSL/CENTER/USL 데이터를 읽어서
    Form Measurement Result 파일에 채움
    
    Args:
        output_path: 출력 파일 경로
        lslusl_file: LSLUSL 파일 경로
        
    Returns:
        결과 메시지
    """
    try:
        if not os.path.exists(output_path):
            return f"Error: Output file not found: {output_path}"
        
        if not os.path.exists(lslusl_file):
            return f"Error: LSLUSL file not found: {lslusl_file}"
        
        debug_info = []
        
        # LSLUSL 파일 읽기
        try:
            df_lslusl = pd.read_excel(lslusl_file, sheet_name='LSLUSL', header=None)
        except:
            df_lslusl = pd.read_excel(lslusl_file, sheet_name=0, header=None)
        
        debug_info.append(f"LSLUSL sheet: {df_lslusl.shape}")
        
        # 출력 파일 열기
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Inner 값과 시작 행 매핑 찾기
        inner_to_row = {}
        for row in range(3, ws.max_row + 1):
            inner_val = ws.cell(row=row, column=2).value  # B열 (Inner)
            if inner_val:
                inner_str = str(inner_val).strip().upper()
                inner_to_row[inner_str] = row
        
        debug_info.append(f"Output file inner mappings: {list(inner_to_row.keys())}")
        
        processed_lslusl = 0
        
        # LSLUSL 데이터 처리
        # Row 0: 헤더 (LSL, Center, USL, inner)
        # Column 0: LSL, Column 1: Center, Column 2: USL, Column 3: inner
        current_inner = None
        for lsl_row in range(2, len(df_lslusl)):
            # inner 열(column 3) 확인
            inner_cell = df_lslusl.iloc[lsl_row, 3] if df_lslusl.shape[1] > 3 else None
            if pd.notna(inner_cell):
                current_inner = str(inner_cell).strip().upper()
            
            if current_inner is None:
                continue
            
            # 해당 Inner의 시작 행 찾기
            target_base_row = inner_to_row.get(current_inner)
            if target_base_row is None:
                continue
            
            # LSLUSL의 행 순서: resistance, width, thickness, minimum (4행 단위)
            row_offset_in_group = (lsl_row - 2) % 4
            target_row = target_base_row + row_offset_in_group
            
            # LSL, CENTER, USL 값 가져오기
            lsl_val = df_lslusl.iloc[lsl_row, 0]
            center_val = df_lslusl.iloc[lsl_row, 1]
            usl_val = df_lslusl.iloc[lsl_row, 2]
            
            # AK(37): LSL, AL(38): Center, AM(39): USL
            if pd.notna(lsl_val):
                ws.cell(row=target_row, column=37, value=lsl_val)  # AK
            if pd.notna(center_val):
                ws.cell(row=target_row, column=38, value=center_val)  # AL
            if pd.notna(usl_val):
                ws.cell(row=target_row, column=39, value=usl_val)  # AM
            
            processed_lslusl += 1
        
        debug_info.append(f"LSLUSL data: {processed_lslusl} rows filled")
        
        # E열부터 마지막 열까지 AutoFilter 추가
        last_col_letter = get_column_letter(ws.max_column)
        filter_range = f"E1:{last_col_letter}{ws.max_row}"
        ws.auto_filter.ref = filter_range
        debug_info.append(f"AutoFilter added: {filter_range}")
        
        # 파일 저장
        wb.save(output_path)
        wb.close()
        
        result_msg = f"Success: Filled LSL/USL data\n"
        result_msg += "Debug:\n  " + "\n  ".join(debug_info)
        return result_msg
        
    except Exception as e:
        import traceback
        return f"Error: {str(e)}\n{traceback.format_exc()}"


def get_template_structure() -> dict:
    """
    템플릿 파일의 구조 정보를 반환합니다 (디버깅/참조용)
    
    Returns:
        템플릿 구조 정보 딕셔너리
    """
    try:
        template_path = get_template_path()
        
        if not os.path.exists(template_path):
            return {"error": f"Template not found: {template_path}"}
        
        wb = load_workbook(template_path)
        ws = wb.active
        
        structure = {
            "sheet_name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged_cells_count": len(list(ws.merged_cells.ranges)),
            "headers_row1": [],
            "headers_row2": [],
            "columns_A_D": []
        }
        
        # Row 1 헤더
        for col in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(row=1, column=col).value
            structure["headers_row1"].append(str(val) if val else "")
        
        # Row 2 서브헤더
        for col in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(row=2, column=col).value
            structure["headers_row2"].append(str(val) if val else "")
        
        # A-D 열 데이터 (첫 10행)
        for row in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col in range(1, 5):
                val = ws.cell(row=row, column=col).value
                row_data.append(str(val) if val else "")
            structure["columns_A_D"].append(row_data)
        
        wb.close()
        return structure
        
    except Exception as e:
        return {"error": str(e)}

