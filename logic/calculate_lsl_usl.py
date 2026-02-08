"""
LSL/USL 계산 모듈
merged_file에서 데이터를 읽어 통계 계산 후 출력 파일 생성
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

from logic.visualizer import save_lslusl_plots_from_data


def convert_to_number_if_possible(val):
    """
    값을 숫자로 변환 가능하면 숫자로, 아니면 원래 값 반환
    
    Args:
        val: 변환할 값
        
    Returns:
        숫자(float/int) 또는 원래 문자열
    """
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    
    # 이미 숫자인 경우
    if isinstance(val, (int, float)):
        return val
    
    # 문자열인 경우 숫자로 변환 시도
    if isinstance(val, str):
        val_stripped = val.strip()
        if val_stripped == '':
            return None
        try:
            # 정수로 변환 시도
            if '.' not in val_stripped:
                return int(val_stripped)
            # 실수로 변환 시도
            return float(val_stripped)
        except (ValueError, TypeError):
            # 변환 실패 시 원래 문자열 반환
            return val
    
    return val


def get_x_from_dcr(dcr_file: str) -> int:
    """
    DCR 파일의 DCR sheet에서 C열의 마지막 숫자(x)를 가져옴
    
    Args:
        dcr_file: DCR 파일 경로
        
    Returns:
        x 값 (데이터 세트 크기)
    """
    try:
        wb = openpyxl.load_workbook(dcr_file, data_only=True)
        if 'DCR' not in wb.sheetnames:
            wb.close()
            return 0
        
        ws = wb['DCR']
        last_val = 0
        
        # C열에서 숫자인 마지막 값 찾기
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=3).value
            if val is not None and isinstance(val, (int, float)):
                last_val = int(val)
        
        wb.close()
        return last_val
    except Exception as e:
        print(f"Error reading DCR file: {e}")
        return 0


def calculate_lsl_usl(merged_file: str, dcr_file: str, output_file: str) -> str:
    """
    merged_file에서 데이터를 읽어 LSL/USL 통계 계산 후 출력 파일 생성
    
    Args:
        merged_file: merged_file.xlsx 경로
        dcr_file: DCR_format_yamaha.xlsx 경로 (x 값을 얻기 위함)
        output_file: 출력 파일 경로
        
    Returns:
        결과 메시지
    """
    try:
        debug_info = []
        
        # 1. DCR 파일에서 x 값 가져오기
        x = get_x_from_dcr(dcr_file)
        if x == 0:
            return "Error: Could not get x value from DCR file"
        
        debug_info.append(f"x value from DCR: {x}")
        
        # 2. merged_file 읽기
        if not os.path.exists(merged_file):
            return f"Error: Merged file not found: {merged_file}"
        
        df = pd.read_excel(merged_file, sheet_name=0, header=None)
        debug_info.append(f"Merged file shape: {df.shape}")
        
        # 3. G열(column 6)부터, row 4(index 3)부터 데이터 추출
        # G열부터 끝까지
        data_start_col = 6  # G열 (0-indexed)
        data_start_row = 3  # row 4 (0-indexed)
        
        # 데이터 영역 추출
        data_df = df.iloc[data_start_row:, data_start_col:].copy()
        data_df = data_df.apply(pd.to_numeric, errors='coerce')
        
        debug_info.append(f"Data shape (from G4): {data_df.shape}")
        
        num_rows = data_df.shape[0]
        num_cols = data_df.shape[1]
        
        # 4. x개씩 잘라서 재구성
        # 각 열에 대해 x개씩 잘라서 옆으로 붙임
        num_sets = num_rows // x
        debug_info.append(f"Number of sets: {num_sets}")
        
        # 결과를 저장할 리스트
        all_data = []
        
        # 각 열(측정 포인트)에 대해
        for col_idx in range(num_cols):
            col_data = data_df.iloc[:, col_idx].values
            
            # x개씩 잘라서 리스트로 만듦
            sets_data = []
            for set_idx in range(num_sets):
                start = set_idx * x
                end = start + x
                if end <= len(col_data):
                    sets_data.append(col_data[start:end])
            
            # 세로로 쌓기 (각 세트가 한 열이 됨)
            if sets_data:
                stacked = np.column_stack(sets_data)  # shape: (x, num_sets)
                all_data.append(stacked)
        
        # 모든 데이터 합치기
        # all_data는 리스트, 각 요소는 (x, num_sets) shape
        # 최종: (x, num_cols * num_sets) 또는 적절히 배열
        
        # 실제로는 각 열을 하나의 데이터 포인트로 보고
        # 여러 측정(set)을 행으로 가짐
        
        # 더 간단한 접근: 첫 번째 열(G)만 처리
        # 실제로는 각 열이 하나의 NET에 대한 측정값
        
        # 각 열을 독립적으로 처리
        # 출력: 행은 NET (1~x), 열은 통계 + 측정값들
        
        debug_info.append(f"Processing {num_cols} measurement points")
        
        # 출력 워크북 생성
        wb_out = openpyxl.Workbook()
        ws = wb_out.active
        ws.title = "tinh LCLUCL"
        
        # 헤더 스타일
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True)
        
        # 헤더 행 작성
        headers = ["Min", "Max", "Average", "Median", "Stdev", "IQR", 
                   "1stQuat-4IQR", "3rdQuat+4IQR", "(A)AverageIfs", "(B)Stdev Ifs",
                   "LSL(A-3B)", "USL(A+3B)"]
        
        # 첫 행에 헤더
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # 측정 세트 헤더 (통계 컬럼 다음)
        stat_cols = len(headers)
        for set_idx in range(num_sets):
            cell = ws.cell(row=1, column=stat_cols + set_idx + 1, value=f"Set{set_idx + 1}")
        
        # 각 NET(행)에 대해 통계 계산
        current_row = 2
        
        for col_idx in range(num_cols):
            col_data = data_df.iloc[:, col_idx].values
            
            # x개씩 잘라서 각 세트의 데이터 수집
            measurement_values = []
            for set_idx in range(num_sets):
                start = set_idx * x
                end = start + x
                if end <= len(col_data):
                    measurement_values.append(col_data[start:end])
            
            if not measurement_values:
                continue
            
            # measurement_values: list of arrays, 각 array는 x개 값
            # 우리가 원하는 것: 각 NET 위치(0~x-1)에 대해 모든 세트의 값으로 통계 계산
            
            for net_idx in range(x):
                # 해당 NET 위치의 모든 측정값 수집
                net_values = []
                for set_data in measurement_values:
                    if net_idx < len(set_data):
                        val = set_data[net_idx]
                        if pd.notna(val):
                            net_values.append(val)
                
                if not net_values:
                    continue
                
                net_values = np.array(net_values)
                
                # 통계 계산
                min_val = np.min(net_values)
                max_val = np.max(net_values)
                avg_val = np.mean(net_values)
                median_val = np.median(net_values)
                std_val = np.std(net_values, ddof=1) if len(net_values) > 1 else 0
                
                q1 = np.percentile(net_values, 25)
                q3 = np.percentile(net_values, 75)
                iqr_val = q3 - q1
                
                q1_4iqr = q1 - 4 * iqr_val
                q3_4iqr = q3 + 4 * iqr_val
                
                # AverageIfs, StdevIfs - IQR 범위 내의 값들만 사용
                lower_bound = q1 - 1.5 * iqr_val
                upper_bound = q3 + 1.5 * iqr_val
                filtered_values = net_values[(net_values >= lower_bound) & (net_values <= upper_bound)]
                
                if len(filtered_values) > 0:
                    avg_ifs = np.mean(filtered_values)
                    std_ifs = np.std(filtered_values, ddof=1) if len(filtered_values) > 1 else 0
                else:
                    avg_ifs = avg_val
                    std_ifs = std_val
                
                # LSL, USL 계산
                lsl = avg_ifs - 3 * std_ifs
                usl = avg_ifs + 3 * std_ifs
                
                # 행 작성
                row_data = [min_val, max_val, avg_val, median_val, std_val, iqr_val,
                           q1_4iqr, q3_4iqr, avg_ifs, std_ifs, lsl, usl]
                
                for col, val in enumerate(row_data, start=1):
                    ws.cell(row=current_row, column=col, value=val)
                
                # 측정값 추가
                for set_idx, set_data in enumerate(measurement_values):
                    if net_idx < len(set_data):
                        ws.cell(row=current_row, column=stat_cols + set_idx + 1, 
                               value=set_data[net_idx])
                
                current_row += 1
            
            # 첫 번째 열만 처리 (테스트용)
            # 실제로는 모든 열 처리
            break  # 일단 첫 번째 열만
        
        debug_info.append(f"Total rows written: {current_row - 2}")
        
        # 열 너비 조정
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        # 파일 저장
        wb_out.save(output_file)
        wb_out.close()
        
        result = f"Success: Created {output_file}\n"
        result += "Debug:\n  " + "\n  ".join(debug_info)
        return result
        
    except Exception as e:
        import traceback
        return f"Error: {str(e)}\n{traceback.format_exc()}"


def calculate_lsl_usl_full(merged_file: str, dcr_file: str, output_file: str, operator: str = "") -> str:
    """
    merged_file의 모든 데이터를 처리하여 통계 계산
    
    처리 과정:
    1. merged file 시트: 원본 데이터 추출 (G열부터, Row 4부터)
    2. Sap xep 시트: 데이터 재배열 (N개씩 잘라서 옆으로)
    3. tinh LCLUCL 시트: 통계 계산
    
    Args:
        merged_file: merged_file.xlsx 경로
        dcr_file: DCR_format_yamaha.xlsx 경로
        output_file: 출력 파일 경로
        
    Returns:
        결과 메시지
    """
    try:
        debug_info = []
        
        # 1. DCR 파일에서 x (N) 값 가져오기
        x = get_x_from_dcr(dcr_file)
        if x == 0:
            return "Error: Could not get x value from DCR file"
        
        debug_info.append(f"N value from DCR (NET count): {x}")
        
        # 2. merged_file 읽기
        if not os.path.exists(merged_file):
            return f"Error: Merged file not found: {merged_file}"
        
        df = pd.read_excel(merged_file, sheet_name=0, header=None)
        debug_info.append(f"Merged file original shape: {df.shape}")
        
        # Method=3인 행만 필터링 (Column D = index 3)
        # Row 0은 헤더이므로 Row 1부터 필터링
        method_col = 3  # Column D (0-indexed)
        data_col_start = 6  # Column G (0-indexed)
        
        # Method=3인 행의 인덱스 찾기 (헤더 제외)
        method_3_mask = df.iloc[1:, method_col] == 3
        method_3_indices = method_3_mask[method_3_mask].index.tolist()
        
        debug_info.append(f"Method=3 rows: {len(method_3_indices)}")
        
        # Method=3인 행만 추출
        df_filtered = df.iloc[method_3_indices]
        
        # 메타데이터 추출 (Method=3인 행들의 PinA, PinB)
        meta_pina = df_filtered.iloc[:, 0].values  # A열 (PinA)
        meta_pinb = df_filtered.iloc[:, 1].values  # B열 (PinB)
        
        # G열 이후 데이터만 추출
        data_df = df_filtered.iloc[:, data_col_start:].copy()
        
        num_rows = data_df.shape[0]
        num_cols = data_df.shape[1]
        num_sets_per_col = num_rows // x  # 각 열에서 세트 수
        
        debug_info.append(f"Data rows: {num_rows}, cols: {num_cols}")
        debug_info.append(f"Sets per column: {num_sets_per_col}")
        
        # 출력 워크북 생성
        wb_out = openpyxl.Workbook()
        
        # ============================================
        # Sheet 1: merged file (원본 데이터 추출)
        # G열부터, Row 4부터의 데이터만 추출
        # ============================================
        ws_merged = wb_out.active
        ws_merged.title = "merged file"
        
        # 데이터를 복사 (숫자로 변환 가능한 것은 숫자로, 아니면 문자로)
        for row_idx in range(num_rows):
            for col_idx in range(num_cols):
                val = data_df.iloc[row_idx, col_idx]
                if pd.notna(val):
                    converted_val = convert_to_number_if_possible(val)
                    if converted_val is not None:
                        ws_merged.cell(row=row_idx + 1, column=col_idx + 1, value=converted_val)
        
        debug_info.append(f"merged file sheet: {num_rows} rows × {num_cols} cols")
        
        # ============================================
        # Sheet 2: Cal_merged (데이터 재배열 - transpose)
        # Excel 열 한계(16,384)를 피하기 위해 transpose해서 저장
        # - 행: 각 측정값 (piece × set)
        # - 열: 각 NET (1~x)
        # 문자열 데이터도 그대로 포함
        # ============================================
        ws_cal = wb_out.create_sheet("Cal_merged")
        
        # 원본 데이터 사용 (숫자 변환 없음, 문자열도 포함)
        total_measurements = num_cols * num_sets_per_col  # 총 측정 횟수
        
        # 각 측정(행)에 대해 모든 NET의 값을 열로 나열
        current_row = 1
        
        for set_idx in range(num_sets_per_col):  # 각 세트
            for piece_idx in range(num_cols):  # 각 piece
                # 이 측정의 각 NET 값을 열로 쓰기
                for net_idx in range(x):
                    row_in_merged = set_idx * x + net_idx  # merged_file에서의 행 위치
                    
                    if row_in_merged < data_df.shape[0]:
                        val = data_df.iloc[row_in_merged, piece_idx]
                        # 숫자로 변환 가능한 것은 숫자로, 아니면 문자로
                        if pd.notna(val):
                            converted_val = convert_to_number_if_possible(val)
                            if converted_val is not None:
                                ws_cal.cell(row=current_row, column=net_idx + 1, value=converted_val)
                
                current_row += 1
        
        total_cal_rows = current_row - 1
        debug_info.append(f"Cal_merged: {total_cal_rows} measurement rows × {x} NET cols")
        
        # ============================================
        # Sheet 3: Sap xep (메타데이터 + 데이터)
        # Cal_merged와 같은 구조 (행=측정값, 열=NET) + 메타데이터
        # ============================================
        ws_sap = wb_out.create_sheet("Sap xep")
        
        # 헤더 정보 (Row 1-5)
        ws_sap.cell(row=1, column=1, value="Input data")
        ws_sap.cell(row=1, column=2, value="NET count")
        ws_sap.cell(row=1, column=3, value=x)
        ws_sap.cell(row=2, column=2, value="Piece count")
        ws_sap.cell(row=2, column=3, value=num_cols)
        ws_sap.cell(row=3, column=2, value="Sets per piece")
        ws_sap.cell(row=3, column=3, value=num_sets_per_col)
        ws_sap.cell(row=4, column=2, value="Total measurements")
        ws_sap.cell(row=4, column=3, value=total_cal_rows)
        
        # Row 6: NET 번호 헤더
        ws_sap.cell(row=6, column=1, value="NET No")
        for net_idx in range(x):
            ws_sap.cell(row=6, column=net_idx + 2, value=net_idx + 1)
        
        # Row 7-12: 메타데이터 (PinA, PinB 등)
        ws_sap.cell(row=7, column=1, value="PinA")
        ws_sap.cell(row=8, column=1, value="PinB")
        ws_sap.cell(row=9, column=1, value="StatementID")
        ws_sap.cell(row=10, column=1, value="Method")
        ws_sap.cell(row=11, column=1, value="Threshold L")
        ws_sap.cell(row=12, column=1, value="Threshold U")
        
        # PinA, PinB 데이터 채우기 (숫자로 변환 가능한 것은 숫자로)
        for net_idx in range(x):
            if net_idx < len(meta_pina):
                val = meta_pina[net_idx]
                if pd.notna(val):
                    converted_val = convert_to_number_if_possible(val)
                    if converted_val is not None:
                        ws_sap.cell(row=7, column=net_idx + 2, value=converted_val)
            if net_idx < len(meta_pinb):
                val = meta_pinb[net_idx]
                if pd.notna(val):
                    converted_val = convert_to_number_if_possible(val)
                    if converted_val is not None:
                        ws_sap.cell(row=8, column=net_idx + 2, value=converted_val)
        
        # Row 13~: Cal_merged 데이터를 복사 (문자열 포함)
        # Cal_merged: 행=측정값, 열=NET (이미 같은 구조)
        data_start_excel_row = 13
        total_data_rows = total_cal_rows
        
        for measure_idx in range(total_cal_rows):
            excel_row = data_start_excel_row + measure_idx
            for net_idx in range(x):
                val = ws_cal.cell(row=measure_idx + 1, column=net_idx + 1).value
                # 모든 값 복사 (None이 아닌 경우, 문자열 포함)
                if val is not None:
                    ws_sap.cell(row=excel_row, column=net_idx + 2, value=val)
        
        debug_info.append(f"Sap xep: {total_data_rows} data rows × {x} NET columns")
        
        # ============================================
        # Sheet 4: tinh LCLUCL (통계 계산 - Excel 수식 사용)
        # ============================================
        ws_tinh = wb_out.create_sheet("tinh LCLUCL")
        
        # 스타일 정의
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        header_font = Font(bold=True)
        
        # A열 헤더 (행 레이블)
        row_labels = [
            "NET no",           # Row 1
            "PinA",             # Row 2
            "PinB",             # Row 3
            "Method",           # Row 4
            "Threshold L",      # Row 5
            "Threshold U",      # Row 6
            "UnderNG",          # Row 7
            "OverNG",           # Row 8
            "Min",              # Row 9
            "Max",              # Row 10
            "Average",          # Row 11
            "Median",           # Row 12
            "Stdev",            # Row 13
            "IQR",              # Row 14
            "1stQuat-4IQR",     # Row 15
            "3rdQuat+4IQR",     # Row 16
            "(A)AverageIfs",    # Row 17
            "(B)Stdev ifs",     # Row 18
            "LSL(A-3B)",        # Row 19
            "USL(A+3B)",        # Row 20
        ]
        
        for row_idx, label in enumerate(row_labels, start=1):
            cell = ws_tinh.cell(row=row_idx, column=1, value=label)
            cell.font = header_font
            # A열에도 색상 적용 (Row 9-16: 회색, Row 17-20: 노란색)
            if 9 <= row_idx <= 16:
                cell.fill = gray_fill
            elif 17 <= row_idx <= 20:
                cell.fill = yellow_fill
        
        # 데이터 끝 행 계산 (Row 21부터 데이터 시작)
        data_start_row = 21
        data_end_row = data_start_row + total_cal_rows - 1
        
        # 각 NET 열에 대해 처리
        for net_idx in range(x):
            col = net_idx + 2  # B열부터 시작 (tinh LCLUCL 시트)
            col_letter = get_column_letter(col)
            
            # Row 1: NET 번호
            ws_tinh.cell(row=1, column=col, value=net_idx + 1)
            
            # Row 2-3: 메타데이터 (숫자로 변환 가능한 것은 숫자로)
            if net_idx < len(meta_pina):
                val = meta_pina[net_idx]
                if pd.notna(val):
                    converted_val = convert_to_number_if_possible(val)
                    ws_tinh.cell(row=2, column=col, value=converted_val if converted_val is not None else "")
            if net_idx < len(meta_pinb):
                val = meta_pinb[net_idx]
                if pd.notna(val):
                    converted_val = convert_to_number_if_possible(val)
                    ws_tinh.cell(row=3, column=col, value=converted_val if converted_val is not None else "")
            
            # Row 4-6: 빈 값
            ws_tinh.cell(row=7, column=col, value=0)  # UnderNG
            ws_tinh.cell(row=8, column=col, value=0)  # OverNG
            
            # Row 21~: Cal_merged에서 모든 값을 그대로 복사 (문자열 포함)
            for measure_row in range(1, total_cal_rows + 1):
                val = ws_cal.cell(row=measure_row, column=net_idx + 1).value
                # 모든 값 저장 (None이 아닌 경우, 문자열 포함)
                if val is not None:
                    ws_tinh.cell(row=data_start_row + measure_row - 1, column=col, value=val)
            
            # 데이터 범위 문자열 생성
            data_range = f"{col_letter}${data_start_row}:{col_letter}${data_end_row}"
            
            # Row 9-16: Excel 수식 (회색 배경)
            # Row 9: Min
            cell = ws_tinh.cell(row=9, column=col, value=f"=MIN({data_range})")
            cell.fill = gray_fill
            
            # Row 10: Max
            cell = ws_tinh.cell(row=10, column=col, value=f"=MAX({data_range})")
            cell.fill = gray_fill
            
            # Row 11: Average
            cell = ws_tinh.cell(row=11, column=col, value=f"=AVERAGE({data_range})")
            cell.fill = gray_fill
            
            # Row 12: Median
            cell = ws_tinh.cell(row=12, column=col, value=f"=MEDIAN({data_range})")
            cell.fill = gray_fill
            
            # Row 13: Stdev
            cell = ws_tinh.cell(row=13, column=col, value=f"=STDEV({data_range})")
            cell.fill = gray_fill
            
            # Row 14: IQR = Q3 - Q1
            cell = ws_tinh.cell(row=14, column=col, 
                value=f"=QUARTILE({data_range},3)-QUARTILE({data_range},1)")
            cell.fill = gray_fill
            
            # Row 15: 1stQuat-4IQR
            cell = ws_tinh.cell(row=15, column=col, 
                value=f"=IF(QUARTILE({data_range},1)-(4*{col_letter}14)<0,0,QUARTILE({data_range},1)-(4*{col_letter}14))")
            cell.fill = gray_fill
            
            # Row 16: 3rdQuat+4IQR
            cell = ws_tinh.cell(row=16, column=col, 
                value=f"=QUARTILE({data_range},3)+(4*{col_letter}14)")
            cell.fill = gray_fill
            
            # Row 17-18: Python에서 직접 계산한 값 (노란색 배경)
            # Row 19-20: Excel 수식 (Row 17, 18 참조하므로 자동 계산)
            
            # Cal_merged에서 해당 NET의 숫자 값만 수집
            net_values = []
            for measure_row in range(1, total_cal_rows + 1):
                val = ws_cal.cell(row=measure_row, column=net_idx + 1).value
                if val is not None and isinstance(val, (int, float)):
                    net_values.append(float(val))
            
            if len(net_values) > 0:
                net_arr = np.array(net_values)
                
                # IQR 범위 계산 (Row 15, 16의 값)
                q1 = np.percentile(net_arr, 25)
                q3 = np.percentile(net_arr, 75)
                iqr = q3 - q1
                lower_bound = max(0, q1 - 4 * iqr)  # 1stQuat-4IQR
                upper_bound = q3 + 4 * iqr          # 3rdQuat+4IQR
                
                # 범위 내의 값만 필터링
                filtered = net_arr[(net_arr > lower_bound) & (net_arr < upper_bound)]
                
                if len(filtered) > 0:
                    avg_ifs = np.mean(filtered)
                    std_ifs = np.std(filtered, ddof=1) if len(filtered) > 1 else 0
                else:
                    avg_ifs = np.mean(net_arr)
                    std_ifs = np.std(net_arr, ddof=1) if len(net_arr) > 1 else 0
                
                # Row 17: (A)AverageIfs - 값으로 저장
                cell = ws_tinh.cell(row=17, column=col, value=avg_ifs)
                cell.fill = yellow_fill
                
                # Row 18: (B)Stdev ifs - 값으로 저장
                cell = ws_tinh.cell(row=18, column=col, value=std_ifs)
                cell.fill = yellow_fill
            else:
                # 데이터가 없는 경우 빈 셀
                cell = ws_tinh.cell(row=17, column=col, value="")
                cell.fill = yellow_fill
                cell = ws_tinh.cell(row=18, column=col, value="")
                cell.fill = yellow_fill
            
            # Row 19: LSL(A-3B) - 수식으로 저장 (Row 17, 18 참조)
            cell = ws_tinh.cell(row=19, column=col, 
                value=f"=ROUNDDOWN(IF({col_letter}17-(3*{col_letter}18)<0,0,{col_letter}17-(3*{col_letter}18)),3)")
            cell.fill = yellow_fill
            
            # Row 20: USL(A+3B) - 수식으로 저장 (Row 17, 18 참조)
            cell = ws_tinh.cell(row=20, column=col, 
                value=f"=ROUNDUP({col_letter}17+(3*{col_letter}18),3)")
            cell.fill = yellow_fill
        
        debug_info.append(f"tinh LCLUCL: {x} NETs, {total_cal_rows} measurements each (formulas applied)")
        
        # 열 너비 조정
        ws_tinh.column_dimensions['A'].width = 15
        for col in range(2, x + 2):
            ws_tinh.column_dimensions[get_column_letter(col)].width = 12
        
        # ============================================
        # Sheet 5: Calculate USL LSL
        # 참조: Calculator LSL,USL 1.xlsm의 "Calculate USL, LSL " 시트
        # ============================================
        ws_calc = wb_out.create_sheet("Calculate USL LSL")
        
        # 색상 정의
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        header_font = Font(bold=True)
        
        # DCR 파일에서 DCR 시트 및 vendor 시트 읽기
        try:
            wb_dcr = openpyxl.load_workbook(dcr_file, data_only=True)
            
            # vendor 시트에서 ERS 값 매핑 (make_dcr.py와 동일한 방식)
            # Key: Pin1 + Pin2 (예: "J_TELE.18U0200.8")
            # Value: (Nominal, USL, LSL)
            vendor_map = {}
            if 'vendor' in wb_dcr.sheetnames:
                ws_vendor = wb_dcr['vendor']
                for row_idx in range(10, ws_vendor.max_row + 1):
                    pin1 = ws_vendor.cell(row=row_idx, column=7).value  # G열 (Pin 1)
                    pin2 = ws_vendor.cell(row=row_idx, column=8).value  # H열 (Pin 2)
                    nominal = ws_vendor.cell(row=row_idx, column=15).value  # O열 (Nominal)
                    usl = ws_vendor.cell(row=row_idx, column=16).value  # P열 (USL)
                    lsl = ws_vendor.cell(row=row_idx, column=17).value  # Q열 (LSL)
                    
                    if pin1 and pin2:
                        key = f"{pin1}{pin2}"
                        vendor_map[key] = (nominal, usl, lsl)
                
                debug_info.append(f"vendor_map entries: {len(vendor_map)}")
            else:
                debug_info.append("Warning: vendor sheet not found in DCR file")
            
            if 'DCR' in wb_dcr.sheetnames:
                ws_dcr = wb_dcr['DCR']
                
                # DCR 시트에서 데이터 행 수 (No 열 기준)
                data_count = 0
                for row_idx in range(2, ws_dcr.max_row + 1):
                    if ws_dcr.cell(row=row_idx, column=3).value:  # C열 (No)
                        data_count += 1
                if data_count == 0:
                    data_count = x
                
                # 데이터 시작 행 (Row 5부터)
                data_start_row = 5
                
                # ===== Row 1-2: 상단 텍스트 =====
                ws_calc.cell(row=1, column=2, value="Yamaha: ± ( 5mohm)")
                ws_calc.cell(row=2, column=2, value="Taiyo: ± ( 10mohm)")
                
                # ===== Row 3: 헤더 1 (병합 셀 포함) =====
                # A: No
                ws_calc.cell(row=3, column=1, value="No")
                ws_calc.cell(row=3, column=1).font = header_font
                ws_calc.merge_cells('A3:A4')
                
                # B: Net name
                ws_calc.cell(row=3, column=2, value="Net name")
                ws_calc.cell(row=3, column=2).font = header_font
                ws_calc.cell(row=3, column=2).fill = yellow_fill
                ws_calc.merge_cells('B3:B4')
                
                # C-D: BtoB (Name, Pin)
                ws_calc.cell(row=3, column=3, value="BtoB")
                ws_calc.cell(row=3, column=3).font = header_font
                ws_calc.cell(row=3, column=3).fill = yellow_fill
                ws_calc.merge_cells('C3:D3')
                ws_calc.cell(row=4, column=3, value="Name")
                ws_calc.cell(row=4, column=4, value="Pin")
                
                # E-F: ACF (Name, Pin)
                ws_calc.cell(row=3, column=5, value="ACF")
                ws_calc.cell(row=3, column=5).font = header_font
                ws_calc.cell(row=3, column=5).fill = yellow_fill
                ws_calc.merge_cells('E3:F3')
                ws_calc.cell(row=4, column=5, value="Name")
                ws_calc.cell(row=4, column=6, value="Pin")
                
                # G-I: ERS (Nominal, LSL, USL)
                ws_calc.cell(row=3, column=7, value="ERS")
                ws_calc.cell(row=3, column=7).font = header_font
                ws_calc.cell(row=3, column=7).fill = light_blue_fill
                ws_calc.merge_cells('G3:I3')
                ws_calc.cell(row=4, column=7, value="Nominal")
                ws_calc.cell(row=4, column=8, value="LSL")
                ws_calc.cell(row=4, column=9, value="USL")
                
                # J-K: Internal (LSL, USL) - 반올림 값
                ws_calc.cell(row=3, column=10, value="Internal")
                ws_calc.cell(row=3, column=10).font = header_font
                ws_calc.cell(row=3, column=10).fill = light_green_fill
                ws_calc.merge_cells('J3:K3')
                ws_calc.cell(row=4, column=10, value="LSL")
                ws_calc.cell(row=4, column=11, value="USL")
                
                # L-M: Judgement (LSL, USL)
                ws_calc.cell(row=3, column=12, value="Judgement")
                ws_calc.cell(row=3, column=12).font = header_font
                ws_calc.cell(row=3, column=12).fill = yellow_fill
                ws_calc.merge_cells('L3:M3')
                ws_calc.cell(row=4, column=12, value="LSL")
                ws_calc.cell(row=4, column=13, value="USL")
                
                # N-O: Internal (반올림 전, LSL, USL)
                ws_calc.cell(row=3, column=14, value="Internal (raw)")
                ws_calc.cell(row=3, column=14).font = header_font
                ws_calc.merge_cells('N3:O3')
                ws_calc.cell(row=4, column=14, value="LSL")
                ws_calc.cell(row=4, column=15, value="USL")
                
                # P: 빈칸
                
                # Q-R: LSL, USL (계산값)
                ws_calc.cell(row=3, column=17, value="Calculated")
                ws_calc.cell(row=3, column=17).font = header_font
                ws_calc.merge_cells('Q3:R3')
                ws_calc.cell(row=4, column=17, value="LSL")
                ws_calc.cell(row=4, column=18, value="USL")
                
                # S-T: *1000
                ws_calc.cell(row=3, column=19, value="*1000")
                ws_calc.cell(row=3, column=19).font = header_font
                ws_calc.merge_cells('S3:T3')
                ws_calc.cell(row=4, column=19, value="LSL")
                ws_calc.cell(row=4, column=20, value="USL")
                
                # ===== DCR에서 데이터 가져오기 =====
                # DCR 시트 구조: Row 2-3 헤더, Row 4부터 데이터
                # DCR 열: C=No, D=Net name, E-F=pin1(BtoB), G-H=pin2(ACF), 
                #         I=ERS Nominal, J=ERS LSL, K=ERS USL, L-M=3sigma, N-O=OnMachine
                dcr_data_start = 4  # DCR 데이터 시작 행
                
                for net_idx in range(min(x, data_count)):
                    row_idx = data_start_row + net_idx  # 출력 행 (Row 5부터)
                    tinh_col = net_idx + 2  # tinh LCLUCL의 B열부터
                    dcr_row = dcr_data_start + net_idx  # DCR 시트의 Row 4부터
                    
                    # A: No (DCR C열)
                    no_val = ws_dcr.cell(row=dcr_row, column=3).value  # DCR C열
                    ws_calc.cell(row=row_idx, column=1, value=no_val if no_val else net_idx + 1)
                    
                    # B: Net name (DCR D열)
                    ws_calc.cell(row=row_idx, column=2, value=ws_dcr.cell(row=dcr_row, column=4).value)
                    
                    # C-D: BtoB Name, Pin (DCR E, F열)
                    btob_name = ws_dcr.cell(row=dcr_row, column=5).value
                    btob_pin = ws_dcr.cell(row=dcr_row, column=6).value
                    ws_calc.cell(row=row_idx, column=3, value=btob_name)
                    ws_calc.cell(row=row_idx, column=4, value=btob_pin)
                    
                    # E-F: ACF Name, Pin (DCR G, H열)
                    acf_name = ws_dcr.cell(row=dcr_row, column=7).value
                    acf_pin = ws_dcr.cell(row=dcr_row, column=8).value
                    ws_calc.cell(row=row_idx, column=5, value=acf_name)
                    ws_calc.cell(row=row_idx, column=6, value=acf_pin)
                    
                    # vendor 시트에서 ERS 값 조회 (make_dcr.py와 동일한 방식)
                    # Key: "Part1.Pin1Part2.Pin2" (예: "J_TELE.1U0200.25")
                    ers_nominal = None
                    ers_usl = None
                    ers_lsl = None
                    
                    if btob_name and btob_pin and acf_name and acf_pin:
                        lookup_key = f"{btob_name}.{btob_pin}{acf_name}.{acf_pin}"
                        if lookup_key in vendor_map:
                            ers_nominal, ers_usl, ers_lsl = vendor_map[lookup_key]
                    
                    # G-I: ERS Nominal, LSL, USL (vendor 시트에서 가져온 값)
                    ws_calc.cell(row=row_idx, column=7, value=ers_nominal)  # ERS Nominal
                    ws_calc.cell(row=row_idx, column=8, value=ers_lsl)      # ERS LSL
                    ws_calc.cell(row=row_idx, column=9, value=ers_usl)      # ERS USL
                    
                    # 계산된 LSL/USL 값 (tinh LCLUCL에서)
                    avg_ifs = ws_tinh.cell(row=17, column=tinh_col).value
                    std_ifs = ws_tinh.cell(row=18, column=tinh_col).value
                    
                    lsl_val = None
                    usl_val = None
                    
                    if avg_ifs is not None and std_ifs is not None:
                        try:
                            avg_ifs = float(avg_ifs)
                            std_ifs = float(std_ifs)
                            
                            # Calculated LSL/USL (소수점)
                            lsl_val = max(0, avg_ifs - 3 * std_ifs)
                            lsl_val = np.floor(lsl_val * 1000) / 1000
                            usl_val = avg_ifs + 3 * std_ifs
                            usl_val = np.ceil(usl_val * 1000) / 1000
                        except:
                            pass
                    
                    # J-K: Internal LSL, USL (반올림)
                    ws_calc.cell(row=row_idx, column=10, value=f"=ROUNDUP(N{row_idx},0)")   # LSL
                    ws_calc.cell(row=row_idx, column=11, value=f"=ROUNDDOWN(O{row_idx},0)") # USL
                    
                    # L-M: Judgement (Judge 수식)
                    ws_calc.cell(row=row_idx, column=12, 
                        value=f'=IF(H{row_idx}="","",IF(J{row_idx}<H{row_idx},"NG","OK"))')
                    ws_calc.cell(row=row_idx, column=13, 
                        value=f'=IF(I{row_idx}="","",IF(K{row_idx}>I{row_idx},"NG","OK"))')
                    
                    # N-O: Internal raw (LSL*1000-5, USL*1000+5)
                    ws_calc.cell(row=row_idx, column=14, value=f"=S{row_idx}-5")  # LSL
                    ws_calc.cell(row=row_idx, column=15, value=f"=T{row_idx}+5")  # USL
                    
                    # Q-R: Calculated LSL, USL (값)
                    ws_calc.cell(row=row_idx, column=17, value=lsl_val)
                    ws_calc.cell(row=row_idx, column=18, value=usl_val)
                    
                    # S-T: *1000 (수식)
                    ws_calc.cell(row=row_idx, column=19, value=f"=Q{row_idx}*1000")
                    ws_calc.cell(row=row_idx, column=20, value=f"=R{row_idx}*1000")
                
                # 마지막 행 (GND-SUS): 모든 LSL/USL 값을 0, 50으로 고정
                last_data_row = data_start_row + min(x, data_count) - 1
                # J-K: Internal (반올림 값)
                ws_calc.cell(row=last_data_row, column=10, value=0)   # Internal LSL
                ws_calc.cell(row=last_data_row, column=11, value=50)  # Internal USL
                # N-O: Internal raw
                ws_calc.cell(row=last_data_row, column=14, value=0)   # Internal raw LSL
                ws_calc.cell(row=last_data_row, column=15, value=50)  # Internal raw USL
                # Q-R: Calculated LSL/USL (0으로 고정하면 계산 무시)
                ws_calc.cell(row=last_data_row, column=17, value=0)   # Calculated LSL
                ws_calc.cell(row=last_data_row, column=18, value=0.05)  # Calculated USL (0.05 * 1000 = 50)
                # S-T: *1000
                ws_calc.cell(row=last_data_row, column=19, value=0)   # LSL * 1000
                ws_calc.cell(row=last_data_row, column=20, value=50)  # USL * 1000
                
                # ===== 열 너비 조정 =====
                col_widths = {
                    1: 5,    # A: No
                    2: 25,   # B: Net name
                    3: 10,   # C: BtoB Name
                    4: 6,    # D: BtoB Pin
                    5: 10,   # E: ACF Name
                    6: 6,    # F: ACF Pin
                    7: 15,   # G: ERS Nominal
                    8: 15,   # H: ERS LSL
                    9: 15,   # I: ERS USL
                    10: 10,  # J: Internal LSL
                    11: 10,  # K: Internal USL
                    12: 8,   # L: Judge LSL
                    13: 8,   # M: Judge USL
                    14: 10,  # N: Internal raw LSL
                    15: 10,  # O: Internal raw USL
                    16: 3,   # P: 빈칸
                    17: 12,  # Q: Calculated LSL
                    18: 12,  # R: Calculated USL
                    19: 12,  # S: *1000 LSL
                    20: 12,  # T: *1000 USL
                }
                for col_idx, width in col_widths.items():
                    ws_calc.column_dimensions[get_column_letter(col_idx)].width = width
                
                debug_info.append(f"Calculate USL LSL: Created with {min(x, data_count)} rows (Reference format)")
                
                wb_dcr.close()
            else:
                debug_info.append("Warning: DCR sheet not found in DCR file")
        except Exception as e:
            import traceback
            debug_info.append(f"Warning: Could not create Calculate USL LSL: {str(e)}")
        
        # 중간 과정 시트 삭제 (merged file, Cal_merged)
        sheets_to_delete = ["merged file", "Cal_merged"]
        for sheet_name in sheets_to_delete:
            if sheet_name in wb_out.sheetnames:
                del wb_out[sheet_name]
                debug_info.append(f"Deleted intermediate sheet: {sheet_name}")
        
        # 파일 저장
        wb_out.save(output_file)
        wb_out.close()
        
        # ============================================
        # DCR_format_yamaha.xlsx 파일 업데이트
        # 계산된 LSL/USL 값을 3 sigma spec (L-M) 및 On machine (N-O) 열에 복사
        # ============================================
        try:
            wb_dcr_update = openpyxl.load_workbook(dcr_file)
            if 'DCR' in wb_dcr_update.sheetnames:
                ws_dcr_update = wb_dcr_update['DCR']
                dcr_data_start = 4  # DCR 데이터 시작 행
                
                updated_count = 0
                for net_idx in range(x):
                    row_idx = dcr_data_start + net_idx
                    tinh_col = net_idx + 2  # tinh LCLUCL의 B열부터
                    
                    # 마지막 행 (GND-SUS): 고정값 0, 50 사용
                    if net_idx == x - 1:
                        # L-M열: 3 sigma spec (LSL=0, USL=50)
                        ws_dcr_update.cell(row=row_idx, column=12, value=0)
                        ws_dcr_update.cell(row=row_idx, column=13, value=50)
                        # N-O열: On machine (LSL=0, USL=50)
                        ws_dcr_update.cell(row=row_idx, column=14, value=0)
                        ws_dcr_update.cell(row=row_idx, column=15, value=50)
                        updated_count += 1
                        continue
                    
                    # tinh LCLUCL에서 AverageIfs, StdevIfs 값 가져오기
                    avg_ifs = ws_tinh.cell(row=17, column=tinh_col).value
                    std_ifs = ws_tinh.cell(row=18, column=tinh_col).value
                    
                    if avg_ifs is not None and std_ifs is not None:
                        try:
                            avg_ifs = float(avg_ifs)
                            std_ifs = float(std_ifs)
                            
                            # LSL/USL 계산
                            lsl_calculated = max(0, avg_ifs - 3 * std_ifs)
                            usl_calculated = avg_ifs + 3 * std_ifs
                            
                            # Internal raw 계산: (LSL * 1000) - 5, (USL * 1000) + 5
                            internal_lsl = int(lsl_calculated * 1000) - 5
                            internal_usl = int(usl_calculated * 1000) + 5
                            
                            # L-M열: 3 sigma spec (LSL, USL)
                            ws_dcr_update.cell(row=row_idx, column=12, value=internal_lsl)
                            ws_dcr_update.cell(row=row_idx, column=13, value=internal_usl)
                            
                            # N-O열: On machine (LSL, USL) - 동일한 값
                            ws_dcr_update.cell(row=row_idx, column=14, value=internal_lsl)
                            ws_dcr_update.cell(row=row_idx, column=15, value=internal_usl)
                            
                            updated_count += 1
                        except:
                            pass
                
                wb_dcr_update.save(dcr_file)
                wb_dcr_update.close()
                debug_info.append(f"Updated DCR file: {updated_count} rows with 3 sigma spec & On machine values")
            else:
                debug_info.append("Warning: DCR sheet not found for update")
        except Exception as e:
            debug_info.append(f"Warning: Could not update DCR file: {str(e)}")
        
        result = f"Success: Created {output_file}\n"
        result += f"Sheet 'Sap xep': {total_data_rows} rows × {x} NET columns\n"
        result += f"Sheet 'tinh LCLUCL': {x} NETs × {total_data_rows} measurements\n"
        result += f"Sheet 'Calculate USL LSL': DCR data with calculated ERS values\n"
        result += f"Updated DCR_format_yamaha.xlsx: 3 sigma spec & On machine columns\n"
        result += "Debug:\n  " + "\n  ".join(debug_info)
        # ============================================
        # Visualization (Top NETs + Control 스타일)
        # ============================================
        try:
            # NET별 데이터프레임 구성 (행=NET, 열=측정값)
            data_by_net = []
            for net_idx in range(x):
                net_rows = data_df.iloc[net_idx::x, :]
                flat_vals = []
                for col in range(net_rows.shape[1]):
                    col_vals = net_rows.iloc[:, col].tolist()
                    for v in col_vals:
                        cv = convert_to_number_if_possible(v)
                        if cv is not None and not (isinstance(cv, float) and np.isnan(cv)):
                            flat_vals.append(cv)
                data_by_net.append(flat_vals)

            data_by_net_df = pd.DataFrame(data_by_net)

            # Python 기준 LSL/USL 계산
            lsl_list = []
            usl_list = []
            for vals in data_by_net:
                series = pd.Series([v for v in vals if pd.notna(v)])
                avg = series.mean() if not series.empty else None
                std = series.std() if not series.empty else None
                if avg is None or std is None:
                    lsl_list.append(None)
                    usl_list.append(None)
                else:
                    lsl = max(0, avg - 3 * std)
                    usl = avg + 3 * std
                    lsl_list.append(lsl)
                    usl_list.append(usl)

            plots = save_lslusl_plots_from_data(data_by_net_df, lsl_list, usl_list, operator)
            for p in plots:
                debug_info.append(f"Plot saved: {p}")
        except Exception as e:
            debug_info.append(f"Warning: Plot generation failed - {str(e)}")

        return result
        
    except Exception as e:
        import traceback
        return f"Error: {str(e)}\n{traceback.format_exc()}"

