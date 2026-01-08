"""
파일 읽기 로직 모듈
.NET 파일 및 .xlsx 파일을 읽어서 문자열로 반환
"""

from openpyxl import load_workbook


def read_net_file(file_path: str) -> str:
    """
    .NET 파일을 읽어서 내용을 문자열로 반환
    
    Args:
        file_path: .NET 파일 경로
        
    Returns:
        파일 내용 문자열
    """
    try:
        # .NET 파일은 텍스트 파일이므로 일반 읽기로 처리
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content
    except UnicodeDecodeError:
        # UTF-8이 실패하면 다른 인코딩 시도
        with open(file_path, 'r', encoding='cp949') as f:
            content = f.read()
        return content
    except Exception as e:
        return f"Error reading .NET file: {str(e)}"


def find_piece_lines(file_path: str) -> list:
    """
    .NET 파일에서 PIECE로 시작하는 줄을 찾아서 반환
    
    Args:
        file_path: .NET 파일 경로
        
    Returns:
        PIECE 줄 리스트 [(line_content, [num1, num2, num3, num4, ...]), ...]
        각 PIECE는 4개 이상의 숫자를 가질 수 있음
    """
    try:
        # 파일 읽기
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
        except UnicodeDecodeError:
            with open(file_path, 'r', encoding='cp949') as f:
                lines = f.readlines()
        
        piece_lines = []
        for line in lines:
            line = line.strip()
            if line.upper().startswith('PIECE'):
                # PIECE 뒤의 숫자들 파싱
                # 예: "PIECE:1,48,2049,2178" 또는 "PIECE 1,48,2049,2178"
                # "PIECE:" 또는 "PIECE" 제거
                parts = line.upper().replace('PIECE:', '').replace('PIECE', '').strip()
                # 콤마로 분리
                numbers = [p.strip() for p in parts.split(',') if p.strip()]
                # 숫자만 추출
                nums = []
                for n in numbers:
                    try:
                        nums.append(int(n))
                    except ValueError:
                        pass
                if nums:
                    piece_lines.append((line, nums))
        
        return piece_lines
    except Exception as e:
        return [(f"Error: {str(e)}", [])]


def parse_4w_section(file_path: str) -> tuple:
    """
    .NET 파일에서 #4W 섹션을 파싱하여 그룹별 데이터 추출
    
    Args:
        file_path: .NET 파일 경로
        
    Returns:
        (groups_dict, debug_info_list)
        groups_dict: {group_name: [[num1, num2, num3, num4], ...], ...}
    """
    debug_info = []
    
    try:
        # 파일 읽기
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
        except UnicodeDecodeError:
            with open(file_path, 'r', encoding='cp949') as f:
                lines = f.readlines()
        
        debug_info.append(f"Total lines: {len(lines)}")
        
        groups = {}
        current_group = None
        in_4w_section = False
        found_4w = False
        
        for i, line in enumerate(lines):
            original_line = line
            line = line.strip()
            
            # #4W 섹션 시작
            if '#4W' in line.upper():
                in_4w_section = True
                found_4w = True
                debug_info.append(f"Found #4W at line {i+1}: {line[:50]}")
                continue
            
            # %END로 섹션 종료
            if '%END' in line.upper() and in_4w_section:
                debug_info.append(f"Found %END at line {i+1}")
                break
            
            if not in_4w_section:
                continue
            
            # 그룹 헤더 (#Gr01, #Gr02, #Gr1, #Gr2, ...)
            if '#GR' in line.upper():
                # 그룹 번호 추출
                group_num = ''.join(filter(str.isdigit, line))
                if group_num:
                    current_group = f"Group {int(group_num)}"
                    if current_group not in groups:
                        groups[current_group] = []
                    debug_info.append(f"Found group: {current_group} at line {i+1}")
                continue
            
            # EXR4W 라인에서 숫자 추출
            if 'EXR4W' in line.upper() and current_group:
                # EXR4W 뒤의 부분 추출
                idx = line.upper().find('EXR4W')
                parts = line[idx+5:].strip()  # EXR4W 이후 부분
                
                # 콤마, 공백, 탭 등으로 분리
                numbers = []
                for part in parts.replace(',', ' ').replace('\t', ' ').split():
                    # 숫자만 추출 (소수점, 음수 등 처리)
                    clean_part = ''.join(c for c in part if c.isdigit() or c == '-')
                    if clean_part:
                        try:
                            numbers.append(int(clean_part))
                        except ValueError:
                            pass
                
                if len(numbers) >= 4:
                    groups[current_group].append(numbers[:4])
                elif numbers:
                    debug_info.append(f"Line {i+1}: Found only {len(numbers)} numbers: {numbers}")
        
        if not found_4w:
            debug_info.append("WARNING: #4W section not found!")
        
        debug_info.append(f"Total groups: {len(groups)}")
        for g, data in groups.items():
            debug_info.append(f"  {g}: {len(data)} rows")
        
        return groups, debug_info
    except Exception as e:
        import traceback
        return {"Error": [[str(e), "", "", ""]]}, [f"Exception: {str(e)}", traceback.format_exc()]


def read_xlsx_file(file_path: str) -> str:
    """
    .xlsx 파일을 읽어서 내용을 문자열로 반환
    모든 시트의 내용을 포함
    
    Args:
        file_path: .xlsx 파일 경로
        
    Returns:
        파일 내용 문자열 (시트별로 구분)
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        result_lines = []
        
        for sheet_name in wb.sheetnames:
            result_lines.append(f"\n{'='*50}")
            result_lines.append(f"Sheet: {sheet_name}")
            result_lines.append('='*50)
            
            sheet = wb[sheet_name]
            
            # 각 행을 읽어서 출력
            for row in sheet.iter_rows():
                row_values = []
                for cell in row:
                    # None인 경우 빈 문자열로 처리
                    value = cell.value if cell.value is not None else ""
                    row_values.append(str(value))
                
                # 빈 행은 건너뛰기
                if any(v.strip() for v in row_values):
                    result_lines.append("\t".join(row_values))
        
        wb.close()
        return "\n".join(result_lines)
    except Exception as e:
        return f"Error reading .xlsx file: {str(e)}"

