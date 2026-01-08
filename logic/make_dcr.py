"""
DCR 시트 생성 모듈
DE Requirement에서 Net name과 pin을 복사하고,
Input Check Pin에서 Group pin을 복사합니다.
"""

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy
import os


def get_de_requirement_data(wb) -> tuple:
    """
    DE Requirement 시트에서 데이터를 가져옵니다.
    
    Returns:
        (data_list, part1_name, part2_name)
        data_list: [(net, part1, pin1, part2, pin2, part_pin1, part_pin2), ...]
    """
    if "DE requirement" not in wb.sheetnames:
        return [], "", ""
    
    ws = wb["DE requirement"]
    
    # Row 2에서 part 이름 확인 (C2, E2)
    # Row 3부터 데이터
    data_list = []
    part1_name = ""
    part2_name = ""
    
    # part 이름 찾기 (C열과 E열의 헤더 또는 첫 데이터에서)
    for row in range(3, ws.max_row + 1):
        net = ws.cell(row=row, column=2).value  # B열: NET
        part1 = ws.cell(row=row, column=3).value  # C열: part1
        pin1 = ws.cell(row=row, column=4).value  # D열: pin1
        part2 = ws.cell(row=row, column=5).value  # E열: part2
        pin2 = ws.cell(row=row, column=6).value  # F열: pin2
        part_pin1 = ws.cell(row=row, column=7).value  # G열: Part&Pin1
        part_pin2 = ws.cell(row=row, column=8).value  # H열: Part&Pin2
        
        if net is not None:
            data_list.append((net, part1, pin1, part2, pin2, part_pin1, part_pin2))
            
            # part 이름 저장 (첫 번째 데이터에서)
            if not part1_name and part1:
                part1_name = str(part1)
            if not part2_name and part2:
                part2_name = str(part2)
    
    return data_list, part1_name, part2_name


def get_input_check_pin_data(wb) -> tuple:
    """
    Input Check Pin 시트에서 그룹별 핀 데이터를 가져옵니다.
    
    Returns:
        (group_data, num_groups)
        group_data: [[(p1, p2, p3, p4), ...], ...] - 각 그룹별 행 데이터
    """
    if "input check pin" not in wb.sheetnames:
        return [], 0
    
    ws = wb["input check pin"]
    
    # Row 11에서 그룹 수 확인
    num_groups = 0
    col = 2  # B열부터
    while col <= ws.max_column:
        cell_val = ws.cell(row=11, column=col).value
        if cell_val and str(cell_val).startswith("Group"):
            num_groups += 1
            col += 4
        else:
            break
    
    if num_groups == 0:
        return [], 0
    
    # Row 12부터 데이터 읽기
    group_data = [[] for _ in range(num_groups)]
    
    for row in range(12, ws.max_row + 1):
        no_val = ws.cell(row=row, column=1).value  # A열: NO
        if no_val is None:
            break
        
        for g in range(num_groups):
            group_start_col = 2 + g * 4  # B, F, J, ...
            p1 = ws.cell(row=row, column=group_start_col).value
            p2 = ws.cell(row=row, column=group_start_col + 1).value
            p3 = ws.cell(row=row, column=group_start_col + 2).value
            p4 = ws.cell(row=row, column=group_start_col + 3).value
            group_data[g].append((p1, p2, p3, p4))
    
    return group_data, num_groups


def make_dcr_sheet(outfile_path: str) -> str:
    """
    DCR 시트를 outfile에 생성합니다.
    
    구조:
    - Row 1: 안내 텍스트 (선택)
    - Row 2: 섹션 헤더 (RGO NET name, ERS spec, etc.)
    - Row 3: 컬럼 헤더 (Net name, pin1, pin2, Gr1, Gr2, ...)
    - Row 4+: 데이터
    
    Args:
        outfile_path: 출력 파일 경로
    
    Returns:
        결과 메시지
    """
    try:
        debug_info = []
        
        if not os.path.exists(outfile_path):
            return f"Error: Output file not found: {outfile_path}"
        
        wb = load_workbook(outfile_path)
        
        # DE Requirement 데이터 가져오기
        de_data, part1_name, part2_name = get_de_requirement_data(wb)
        if not de_data:
            wb.close()
            return "Error: No data found in 'DE requirement' sheet"
        
        debug_info.append(f"DE data rows: {len(de_data)}")
        debug_info.append(f"Parts: {part1_name}, {part2_name}")
        
        # Input Check Pin 데이터 가져오기
        group_data, num_groups = get_input_check_pin_data(wb)
        if num_groups == 0:
            wb.close()
            return "Error: No groups found in 'input check pin' sheet"
        
        debug_info.append(f"Groups: {num_groups}")
        
        # 기존 시트 삭제
        if "DCR" in wb.sheetnames:
            del wb["DCR"]
        
        # 새 시트 생성
        ws = wb.create_sheet("DCR")
        
        # 스타일 정의
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # === Row 1: 안내 텍스트 (선택) ===
        ws['P1'] = "when submit, please delete these info"
        ws['P1'].font = Font(color="FF0000")
        
        # === Row 2: 섹션 헤더 ===
        # D2-H2: RGO NET name and pin assign
        ws.merge_cells('D2:H2')
        ws['D2'] = "RGO NET name and pin assign"
        ws['D2'].fill = header_fill
        ws['D2'].font = header_font
        ws['D2'].alignment = center_align
        
        # I2-K2: ERS spec
        ws.merge_cells('I2:K2')
        ws['I2'] = "ERS spec"
        ws['I2'].fill = header_fill
        ws['I2'].font = header_font
        ws['I2'].alignment = center_align
        
        # L2-M2: 3 sigma spec
        ws.merge_cells('L2:M2')
        ws['L2'] = "3 sigma spec"
        ws['L2'].fill = header_fill
        ws['L2'].font = header_font
        ws['L2'].alignment = center_align
        
        # N2-O2: On machine
        ws.merge_cells('N2:O2')
        ws['N2'] = "On machine"
        ws['N2'].fill = header_fill
        ws['N2'].font = header_font
        ws['N2'].alignment = center_align
        
        # P2 이후: our jig pin No (그룹별 4열씩)
        group_header_start = 16  # P열
        group_header_end = group_header_start + num_groups * 4 - 1
        ws.merge_cells(f'{get_column_letter(group_header_start)}2:{get_column_letter(group_header_end)}2')
        ws.cell(row=2, column=group_header_start).value = "our jig pin No correlated with RGO assignment"
        ws.cell(row=2, column=group_header_start).fill = header_fill
        ws.cell(row=2, column=group_header_start).font = header_font
        ws.cell(row=2, column=group_header_start).alignment = center_align
        
        # === Row 3: 컬럼 헤더 ===
        headers_row3 = [
            ('A3', ''),
            ('B3', ''),
            ('C3', 'No'),
            ('D3', 'Net name'),
            ('E3', 'pin1'),
            ('F3', ''),
            ('G3', 'pin2'),
            ('H3', ''),
            ('I3', 'Nominal'),
            ('J3', 'LSL'),
            ('K3', 'USL'),
            ('L3', 'LSL'),
            ('M3', 'USL'),
            ('N3', 'LSL'),
            ('O3', 'USL'),
        ]
        
        for cell_ref, value in headers_row3:
            ws[cell_ref] = value
            ws[cell_ref].font = Font(bold=True)
            ws[cell_ref].alignment = center_align
        
        # 그룹 헤더 (Gr1, Gr2, ...) - 4열씩 병합
        col = 16  # P열
        for g in range(1, num_groups + 1):
            start_col_letter = get_column_letter(col)
            end_col_letter = get_column_letter(col + 3)
            ws.merge_cells(f'{start_col_letter}3:{end_col_letter}3')
            cell = ws.cell(row=3, column=col)
            cell.value = f"Gr{g}"
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            col += 4
        
        # === Row 4+: 데이터 ===
        data_start_row = 4
        
        for idx, (net, part1, pin1, part2, pin2, part_pin1, part_pin2) in enumerate(de_data):
            row = data_start_row + idx
            
            # A열: =Q{row}&S{row} (Gr1의 2,4번째 값 결합)
            ws.cell(row=row, column=1).value = f'=Q{row}&S{row}'
            
            # B열: =E{row}&"."&F{row}&G{row}&"."&H{row}
            ws.cell(row=row, column=2).value = f'=E{row}&"."&F{row}&G{row}&"."&H{row}'
            
            # C열: NO
            ws.cell(row=row, column=3).value = idx + 1
            
            # D열: Net name
            ws.cell(row=row, column=4).value = net
            
            # E열: pin1 part
            ws.cell(row=row, column=5).value = part1
            
            # F열: pin1 number
            ws.cell(row=row, column=6).value = pin1
            
            # G열: pin2 part
            ws.cell(row=row, column=7).value = part2
            
            # H열: pin2 number
            ws.cell(row=row, column=8).value = pin2
            
            # I-K열: ERS spec (VLOOKUP from vendor)
            # I: Nominal
            ws.cell(row=row, column=9).value = f'=IFERROR(VLOOKUP(B{row},vendor!$A$10:$Q$100,15,0),"")'
            # J: LSL
            ws.cell(row=row, column=10).value = f'=IFERROR(VLOOKUP(B{row},vendor!$A$10:$Q$100,17,0),"")'
            # K: USL
            ws.cell(row=row, column=11).value = f'=IFERROR(VLOOKUP(B{row},vendor!$A$10:$Q$100,16,0),"")'
            
            # L-M열: 3 sigma spec (빈 칸 또는 수동 입력)
            # N-O열: On machine (빈 칸 또는 수동 입력)
            
            # P열 이후: 그룹별 핀 데이터 (Input Check Pin에서)
            col = 16  # P열
            for g in range(num_groups):
                if idx < len(group_data[g]):
                    p1, p2, p3, p4 = group_data[g][idx]
                    ws.cell(row=row, column=col).value = p1
                    ws.cell(row=row, column=col + 1).value = p2
                    ws.cell(row=row, column=col + 2).value = p3
                    ws.cell(row=row, column=col + 3).value = p4
                col += 4
        
        debug_info.append(f"Created {len(de_data)} data rows")
        
        # 컬럼 너비 설정
        col_widths = {
            'A': 15, 'B': 25, 'C': 5, 'D': 35, 'E': 10, 'F': 5,
            'G': 10, 'H': 5, 'I': 10, 'J': 8, 'K': 8, 'L': 8,
            'M': 8, 'N': 8, 'O': 8
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # 그룹 열 너비
        for col in range(16, 16 + num_groups * 4):
            ws.column_dimensions[get_column_letter(col)].width = 8
        
        # 파일 저장
        wb.save(outfile_path)
        wb.close()
        
        result_msg = f"Success: Created 'DCR' sheet\n"
        result_msg += f"Data rows: {len(de_data)}, Groups: {num_groups}\n"
        result_msg += "Debug: " + " | ".join(debug_info)
        return result_msg
        
    except Exception as e:
        import traceback
        return f"Error: {str(e)}\n{traceback.format_exc()}"

