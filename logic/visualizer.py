"""
시각화 모듈
matplotlib를 사용하여 주요 분석 결과를 PNG로 저장합니다.
"""

import os
from datetime import datetime
from typing import Dict, List, Tuple, Optional

import matplotlib

# GUI 백엔드 충돌을 피하기 위해 Agg 사용
matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
import numpy as np  # noqa: E402
import pandas as pd  # noqa: E402
from scipy import stats  # noqa: E402

from logic.config_manager import get_app_dir
from openpyxl import load_workbook


# 공통 경로 유틸
def _get_plots_dir() -> str:
    base = os.path.join(get_app_dir(), "output", "plots")
    os.makedirs(base, exist_ok=True)
    return base


def _timestamp_label(operator: str) -> str:
    op = operator if operator else "Unknown"
    date_str = datetime.now().strftime("%Y%m%d")
    return f"{op}_{date_str}"


# 공통 스타일 설정
def _setup_style():
    """플롯 스타일 설정"""
    plt.style.use('seaborn-v0_8-whitegrid')
    plt.rcParams['font.size'] = 10
    plt.rcParams['axes.titlesize'] = 12
    plt.rcParams['axes.labelsize'] = 10
    plt.rcParams['figure.facecolor'] = 'white'


# =========================
# Tab1: DCR Format
# =========================
def save_dcr_plots_from_file(
    output_file: str,
    operator: str = "",
) -> List[str]:
    """
    DCR Format 결과 파일에서 통계 플롯 생성

    Args:
        output_file: 생성된 DCR_format_yamaha 파일 경로
        operator: 작업자 이름
    Returns:
        저장된 파일 경로 리스트
    """
    _setup_style()
    saved_paths = []
    plots_dir = _get_plots_dir()
    label = _timestamp_label(operator)

    if not os.path.exists(output_file):
        return saved_paths

    try:
        wb = load_workbook(output_file, data_only=True)
    except Exception:
        return saved_paths

    # 1) Vendor Sheet 분석: NET별 저항 스펙 (Min/Typ/Max)
    if "vendorspec" in wb.sheetnames:
        ws = wb["vendorspec"]
        net_data = []
        
        # A열: NET, 스펙 데이터 찾기 (row 2부터)
        for row in range(2, min(ws.max_row + 1, 102)):  # 최대 100개 NET
            net_name = ws.cell(row=row, column=1).value
            if net_name is None:
                continue
            
            # 저항 스펙 컬럼 찾기 (보통 Min, Typ, Max)
            min_val = None
            typ_val = None
            max_val = None
            
            # 각 셀에서 숫자 데이터 찾기
            for col in range(2, min(ws.max_column + 1, 20)):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val is not None:
                    try:
                        val = float(cell_val)
                        if min_val is None:
                            min_val = val
                        elif typ_val is None:
                            typ_val = val
                        elif max_val is None:
                            max_val = val
                    except (ValueError, TypeError):
                        pass
            
            if min_val is not None or typ_val is not None or max_val is not None:
                net_data.append({
                    'NET': str(net_name),
                    'Min': min_val if min_val else 0,
                    'Typ': typ_val if typ_val else 0,
                    'Max': max_val if max_val else 0
                })
        
        if net_data:
            # 막대 차트: NET별 Min/Typ/Max 저항 스펙
            fig, ax = plt.subplots(figsize=(14, 6))
            
            x = np.arange(len(net_data))
            width = 0.25
            
            mins = [d['Min'] for d in net_data]
            typs = [d['Typ'] for d in net_data]
            maxs = [d['Max'] for d in net_data]
            
            ax.bar(x - width, mins, width, label='Min', color='#5B9BD5', edgecolor='black')
            ax.bar(x, typs, width, label='Typ', color='#70AD47', edgecolor='black')
            ax.bar(x + width, maxs, width, label='Max', color='#ED7D31', edgecolor='black')
            
            ax.set_xticks(x)
            ax.set_xticklabels([d['NET'][:15] for d in net_data], rotation=45, ha='right', fontsize=8)
            ax.set_title("DCR Vendor Spec by NET (Min/Typ/Max)", fontsize=14, fontweight='bold')
            ax.set_xlabel("NET Name")
            ax.set_ylabel("Resistance Value")
            ax.legend()
            ax.grid(True, axis='y', linestyle='--', alpha=0.4)
            
            path = os.path.join(plots_dir, f"DCR_VendorSpec_{label}.png")
            plt.tight_layout()
            plt.savefig(path, dpi=200)
            plt.close()
            saved_paths.append(path)

            # 스펙 범위 차트 (Max - Min)
            ranges = [d['Max'] - d['Min'] for d in net_data]
            fig, ax = plt.subplots(figsize=(12, 6))
            colors = plt.cm.viridis(np.linspace(0, 1, len(ranges)))
            ax.bar(range(len(ranges)), ranges, color=colors, edgecolor='black')
            ax.set_xticks(range(len(net_data)))
            ax.set_xticklabels([d['NET'][:10] for d in net_data], rotation=45, ha='right', fontsize=8)
            ax.set_title("DCR Spec Range by NET (Max - Min)", fontsize=14, fontweight='bold')
            ax.set_xlabel("NET Name")
            ax.set_ylabel("Spec Range")
            ax.grid(True, axis='y', linestyle='--', alpha=0.4)
            
            path = os.path.join(plots_dir, f"DCR_SpecRange_{label}.png")
            plt.tight_layout()
            plt.savefig(path, dpi=200)
            plt.close()
            saved_paths.append(path)

    # 2) DE Requirement 분석: Part별 Pin 분포
    if "DE requirement" in wb.sheetnames:
        ws = wb["DE requirement"]
        part_pins = {}  # {part_name: pin_count}
        nets = []
        
        for row in range(3, ws.max_row + 1):
            net = ws.cell(row=row, column=2).value
            part1 = ws.cell(row=row, column=3).value
            part2 = ws.cell(row=row, column=5).value
            
            if net:
                nets.append(str(net))
            if part1:
                part_pins[str(part1)] = part_pins.get(str(part1), 0) + 1
            if part2:
                part_pins[str(part2)] = part_pins.get(str(part2), 0) + 1
        
        if part_pins:
            fig, ax = plt.subplots(figsize=(10, 6))
            parts = list(part_pins.keys())
            counts = list(part_pins.values())
            colors = plt.cm.Set2(np.linspace(0, 1, len(parts)))
            
            wedges, texts, autotexts = ax.pie(counts, labels=parts, colors=colors, autopct='%1.1f%%',
                                              shadow=True, startangle=90)
            ax.set_title("Part Distribution in DE Requirement", fontsize=14, fontweight='bold')
            
            path = os.path.join(plots_dir, f"DCR_PartDist_{label}.png")
            plt.tight_layout()
            plt.savefig(path, dpi=200)
            plt.close()
            saved_paths.append(path)

    # 3) Judge(check pin) 분석: Pass/Fail 통계
    if "Judge(check pin)" in wb.sheetnames:
        ws = wb["Judge(check pin)"]
        pass_count = 0
        fail_count = 0
        
        # Judge 컬럼 찾기 (보통 마지막 컬럼에 Judge 결과)
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val is not None:
                    val_str = str(cell_val).upper()
                    if val_str in ['PASS', 'OK', 'O', 'GOOD']:
                        pass_count += 1
                    elif val_str in ['FAIL', 'NG', 'X', 'BAD']:
                        fail_count += 1
        
        total = pass_count + fail_count
        if total > 0:
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
            
            # 파이 차트
            sizes = [pass_count, fail_count]
            labels = ['Pass', 'Fail']
            colors = ['#70AD47', '#C00000']
            explode = (0, 0.1) if fail_count > 0 else (0, 0)
            
            ax1.pie(sizes, explode=explode, labels=labels, colors=colors, autopct='%1.1f%%',
                   shadow=True, startangle=90)
            ax1.set_title(f"DCR Judge Results\n(Total: {total})", fontsize=12, fontweight='bold')
            
            # 막대 차트
            ax2.bar(['Pass', 'Fail'], [pass_count, fail_count], color=colors, edgecolor='black')
            ax2.set_title("Pass/Fail Count", fontsize=12, fontweight='bold')
            ax2.set_ylabel("Count")
            
            for i, (count, pct) in enumerate(zip([pass_count, fail_count], 
                                                  [pass_count/total*100, fail_count/total*100])):
                ax2.annotate(f'{count:,}\n({pct:.1f}%)',
                            xy=(i, count), xytext=(0, 5),
                            textcoords="offset points", ha='center', va='bottom', fontsize=10)
            ax2.grid(True, axis='y', linestyle='--', alpha=0.4)
            
            path = os.path.join(plots_dir, f"DCR_JudgeResult_{label}.png")
            plt.tight_layout()
            plt.savefig(path, dpi=200)
            plt.close()
            saved_paths.append(path)

    # 4) DCR Sheet 분석: 데이터 요약
    if "DCR" in wb.sheetnames:
        ws = wb["DCR"]
        dcr_values = []
        
        # 숫자 데이터 수집
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val is not None:
                    try:
                        val = float(cell_val)
                        if 0 < val < 1000:  # 합리적인 범위의 값만
                            dcr_values.append(val)
                    except (ValueError, TypeError):
                        pass
        
        if dcr_values:
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
            
            # 히스토그램
            ax1.hist(dcr_values, bins=30, color='steelblue', edgecolor='black', alpha=0.7)
            ax1.axvline(np.mean(dcr_values), color='red', linestyle='--', linewidth=2, 
                       label=f'Mean: {np.mean(dcr_values):.4f}')
            ax1.set_title("DCR Values Distribution", fontsize=12, fontweight='bold')
            ax1.set_xlabel("Value")
            ax1.set_ylabel("Frequency")
            ax1.legend()
            ax1.grid(True, linestyle='--', alpha=0.4)
            
            # Box Plot
            bp = ax2.boxplot(dcr_values, vert=True, patch_artist=True)
            bp['boxes'][0].set_facecolor('#5B9BD5')
            bp['boxes'][0].set_alpha(0.7)
            ax2.set_title("DCR Values Box Plot", fontsize=12, fontweight='bold')
            ax2.set_ylabel("Value")
            ax2.grid(True, axis='y', linestyle='--', alpha=0.4)
            
            # 통계 정보 추가
            stats_text = f"Count: {len(dcr_values)}\nMean: {np.mean(dcr_values):.4f}\n"
            stats_text += f"Std: {np.std(dcr_values):.4f}\nMin: {np.min(dcr_values):.4f}\nMax: {np.max(dcr_values):.4f}"
            ax2.text(1.3, np.mean(dcr_values), stats_text, fontsize=9, verticalalignment='center',
                    bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
            
            plt.suptitle("DCR Sheet Analysis", fontsize=14, fontweight='bold')
            
            path = os.path.join(plots_dir, f"DCR_Analysis_{label}.png")
            plt.tight_layout()
            plt.savefig(path, dpi=200)
            plt.close()
            saved_paths.append(path)

    # 5) 전체 요약 테이블
    summary_data = []
    
    # 각 시트별 정보 수집
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        row_count = ws.max_row - 1 if ws.max_row > 1 else 0
        col_count = ws.max_column
        summary_data.append({
            'Sheet': sheet_name,
            'Rows': row_count,
            'Columns': col_count
        })
    
    if summary_data:
        fig, ax = plt.subplots(figsize=(10, len(summary_data) * 0.5 + 2))
        ax.axis('off')
        
        df = pd.DataFrame(summary_data)
        table = ax.table(
            cellText=df.values,
            colLabels=df.columns,
            cellLoc='center',
            loc='center',
            colColours=['#4472C4'] * len(df.columns)
        )
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        table.scale(1.2, 1.5)
        
        for i in range(len(df.columns)):
            table[(0, i)].set_text_props(color='white', fontweight='bold')
        
        ax.set_title("DCR File Sheet Summary", fontsize=14, fontweight='bold', pad=20)
        
        path = os.path.join(plots_dir, f"DCR_Summary_{label}.png")
        plt.tight_layout()
        plt.savefig(path, dpi=200, bbox_inches='tight')
        plt.close()
        saved_paths.append(path)

    wb.close()
    return saved_paths


# =========================
# Tab2: Form Measurement
# =========================
def save_form_plots_from_workbook(
    tdr_map: Dict[str, List[float]],
    dim_map: Dict[str, Tuple[float, float]],
    operator: str = "",
) -> List[str]:
    """
    Form Measurement 결과에서 추출한 TDR/치수 데이터를 시각화하여 PNG 저장

    Args:
        tdr_map: {inner: [tdr values...]}
        dim_map: {inner: (width_avg, thickness_avg)}
    Returns:
        저장된 파일 경로 리스트
    """
    _setup_style()
    saved_paths = []
    plots_dir = _get_plots_dir()
    label = _timestamp_label(operator)

    # 1) TDR Box Plot
    if tdr_map:
        inners = list(tdr_map.keys())
        data = [tdr_map[k] for k in inners if tdr_map[k]]
        if data:
            fig, ax = plt.subplots(figsize=(12, 6))
            bp = ax.boxplot(data, labels=inners, showfliers=True, patch_artist=True)
            
            # 색상 설정
            colors = plt.cm.Set3(np.linspace(0, 1, len(data)))
            for patch, color in zip(bp['boxes'], colors):
                patch.set_facecolor(color)
                patch.set_alpha(0.7)
            
            ax.set_title("TDR Distribution by Inner", fontsize=14, fontweight='bold')
            ax.set_xlabel("Inner")
            ax.set_ylabel("TDR Value")
            ax.grid(True, axis="y", linestyle="--", alpha=0.5)
            
            # 평균선 추가
            means = [np.mean(d) for d in data]
            ax.scatter(range(1, len(means)+1), means, color='red', marker='D', s=50, zorder=5, label='Mean')
            ax.legend()
            
            path = os.path.join(plots_dir, f"Form_TDR_BoxPlot_{label}.png")
            plt.tight_layout()
            plt.savefig(path, dpi=200)
            plt.close()
            saved_paths.append(path)

    # 2) TDR Violin Plot (분포 시각화)
    if tdr_map:
        inners = list(tdr_map.keys())
        data = [tdr_map[k] for k in inners if tdr_map[k]]
        if data and len(data) > 0:
            fig, ax = plt.subplots(figsize=(12, 6))
            parts = ax.violinplot(data, showmeans=True, showmedians=True)
            
            # 색상 설정
            colors = plt.cm.Pastel1(np.linspace(0, 1, len(data)))
            for i, pc in enumerate(parts['bodies']):
                pc.set_facecolor(colors[i])
                pc.set_alpha(0.7)
            
            ax.set_xticks(range(1, len(inners)+1))
            ax.set_xticklabels(inners)
            ax.set_title("TDR Distribution (Violin Plot)", fontsize=14, fontweight='bold')
            ax.set_xlabel("Inner")
            ax.set_ylabel("TDR Value")
            ax.grid(True, axis="y", linestyle="--", alpha=0.5)
            
            path = os.path.join(plots_dir, f"Form_TDR_Violin_{label}.png")
            plt.tight_layout()
            plt.savefig(path, dpi=200)
            plt.close()
            saved_paths.append(path)

    # 3) TDR Statistics Summary Table
    if tdr_map:
        inners = list(tdr_map.keys())
        data = [tdr_map[k] for k in inners if tdr_map[k]]
        if data:
            stats_data = []
            for inner, values in zip(inners, data):
                if values:
                    stats_data.append({
                        'Inner': inner,
                        'Count': len(values),
                        'Min': f"{np.min(values):.3f}",
                        'Max': f"{np.max(values):.3f}",
                        'Mean': f"{np.mean(values):.3f}",
                        'Median': f"{np.median(values):.3f}",
                        'Std': f"{np.std(values):.3f}",
                        'Range': f"{np.max(values) - np.min(values):.3f}"
                    })
            
            if stats_data:
                fig, ax = plt.subplots(figsize=(14, len(stats_data) * 0.6 + 2))
                ax.axis('off')
                
                df = pd.DataFrame(stats_data)
                table = ax.table(
                    cellText=df.values,
                    colLabels=df.columns,
                    cellLoc='center',
                    loc='center',
                    colColours=['#4472C4'] * len(df.columns)
                )
                table.auto_set_font_size(False)
                table.set_fontsize(10)
                table.scale(1.2, 1.5)
                
                # 헤더 텍스트 흰색
                for i in range(len(df.columns)):
                    table[(0, i)].set_text_props(color='white', fontweight='bold')
                
                ax.set_title("TDR Statistics Summary", fontsize=14, fontweight='bold', pad=20)
                
                path = os.path.join(plots_dir, f"Form_TDR_Stats_{label}.png")
                plt.tight_layout()
                plt.savefig(path, dpi=200, bbox_inches='tight')
                plt.close()
                saved_paths.append(path)

    # 4) Dimension Trend (Width & Thickness) - 개선된 버전
    if dim_map:
        inners = list(dim_map.keys())
        width_vals = [dim_map[k][0] for k in inners]
        thick_vals = [dim_map[k][1] for k in inners]

        x = np.arange(len(inners))
        bar_width = 0.35

        fig, ax = plt.subplots(figsize=(12, 6))
        bars1 = ax.bar(x - bar_width/2, width_vals, bar_width, label="Width (um)", color='#5B9BD5', edgecolor='black')
        bars2 = ax.bar(x + bar_width/2, thick_vals, bar_width, label="Thickness (um)", color='#70AD47', edgecolor='black')
        
        # 값 표시
        for bar in bars1:
            height = bar.get_height()
            ax.annotate(f'{height:.1f}',
                       xy=(bar.get_x() + bar.get_width()/2, height),
                       xytext=(0, 3), textcoords="offset points",
                       ha='center', va='bottom', fontsize=8)
        for bar in bars2:
            height = bar.get_height()
            ax.annotate(f'{height:.1f}',
                       xy=(bar.get_x() + bar.get_width()/2, height),
                       xytext=(0, 3), textcoords="offset points",
                       ha='center', va='bottom', fontsize=8)
        
        ax.set_xticks(x)
        ax.set_xticklabels(inners)
        ax.set_title("Dimension Trend by Inner", fontsize=14, fontweight='bold')
        ax.set_xlabel("Inner")
        ax.set_ylabel("Value (um)")
        ax.legend()
        ax.grid(True, axis="y", linestyle="--", alpha=0.4)
        
        path = os.path.join(plots_dir, f"Form_Dimension_{label}.png")
        plt.tight_layout()
        plt.savefig(path, dpi=200)
        plt.close()
        saved_paths.append(path)

    # 5) TDR All Data Histogram (전체 분포)
    if tdr_map:
        all_tdr = []
        for values in tdr_map.values():
            if values:
                all_tdr.extend(values)
        
        if all_tdr:
            fig, ax = plt.subplots(figsize=(10, 6))
            
            # 히스토그램
            n, bins, patches = ax.hist(all_tdr, bins=30, density=True, alpha=0.7, 
                                       color='steelblue', edgecolor='black')
            
            # 정규분포 피팅
            mu, std = stats.norm.fit(all_tdr)
            x_fit = np.linspace(min(all_tdr), max(all_tdr), 100)
            y_fit = stats.norm.pdf(x_fit, mu, std)
            ax.plot(x_fit, y_fit, 'r-', linewidth=2, label=f'Normal Fit (μ={mu:.3f}, σ={std:.3f})')
            
            ax.axvline(mu, color='red', linestyle='--', alpha=0.8, label=f'Mean: {mu:.3f}')
            ax.axvline(mu - 3*std, color='orange', linestyle=':', alpha=0.8, label=f'-3σ: {mu-3*std:.3f}')
            ax.axvline(mu + 3*std, color='orange', linestyle=':', alpha=0.8, label=f'+3σ: {mu+3*std:.3f}')
            
            ax.set_title("TDR All Data Distribution with Normal Fit", fontsize=14, fontweight='bold')
            ax.set_xlabel("TDR Value")
            ax.set_ylabel("Density")
            ax.legend(loc='upper right')
            ax.grid(True, linestyle="--", alpha=0.4)
            
            path = os.path.join(plots_dir, f"Form_TDR_Distribution_{label}.png")
            plt.tight_layout()
            plt.savefig(path, dpi=200)
            plt.close()
            saved_paths.append(path)

    return saved_paths


# =========================
# Tab3: LSL/USL 계산
# =========================
def save_lslusl_plots_from_data(
    data_df,
    lsl_values: List[float],
    usl_values: List[float],
    operator: str = "",
    top_k: int = 5,
) -> List[str]:
    """
    LSL/USL 계산 결과를 시각화하여 PNG 저장

    Args:
        data_df: pandas DataFrame (rows=NET, cols=측정값)
        lsl_values/usl_values: 각 NET별 LSL/USL 값 (길이 = NET 수)
        top_k: 표준편차 기준 상위 NET 개수
    Returns:
        저장된 파일 경로 리스트
    """
    _setup_style()
    saved_paths = []
    plots_dir = _get_plots_dir()
    label = _timestamp_label(operator)

    if not isinstance(data_df, pd.DataFrame) or data_df.empty:
        return saved_paths

    # NET별 통계 계산
    means = data_df.mean(axis=1, skipna=True)
    stds = data_df.std(axis=1, skipna=True)
    mins = data_df.min(axis=1, skipna=True)
    maxs = data_df.max(axis=1, skipna=True)

    # 1) Control Chart: NET별 평균 + LSL/USL (개선)
    x = np.arange(len(means)) + 1
    fig, ax = plt.subplots(figsize=(14, 6))
    
    # 평균 플롯
    ax.plot(x, means, marker="o", linewidth=2, markersize=4, label="Mean", color='#2E75B6')
    
    # LSL/USL 영역 채우기
    if lsl_values and usl_values:
        ax.fill_between(x, lsl_values, usl_values, alpha=0.2, color='green', label='Spec Range')
        ax.plot(x, lsl_values, linestyle="--", color="red", linewidth=1.5, label="LSL")
        ax.plot(x, usl_values, linestyle="--", color="green", linewidth=1.5, label="USL")
    
    # 벗어난 포인트 강조
    if lsl_values and usl_values:
        out_of_spec = []
        for i, (m, l, u) in enumerate(zip(means, lsl_values, usl_values)):
            if m < l or m > u:
                out_of_spec.append(i)
        if out_of_spec:
            ax.scatter([x[i] for i in out_of_spec], [means.iloc[i] for i in out_of_spec],
                      color='red', s=100, zorder=5, marker='x', label='Out of Spec')
    
    ax.set_title("NET-wise Mean with LSL/USL Control Chart", fontsize=14, fontweight='bold')
    ax.set_xlabel("NET No")
    ax.set_ylabel("Value")
    ax.grid(True, linestyle="--", alpha=0.4)
    ax.legend(loc='upper right')
    
    path = os.path.join(plots_dir, f"LSLUSL_Control_{label}.png")
    plt.tight_layout()
    plt.savefig(path, dpi=200)
    plt.close()
    saved_paths.append(path)

    # 2) 전체 데이터 히스토그램 (정규분포 오버레이)
    all_values = data_df.values.flatten()
    all_values = all_values[~np.isnan(all_values)]
    
    if len(all_values) > 0:
        fig, ax = plt.subplots(figsize=(12, 6))
        
        n, bins, patches = ax.hist(all_values, bins=50, density=True, alpha=0.7,
                                   color='steelblue', edgecolor='black')
        
        # 정규분포 피팅
        mu, std = stats.norm.fit(all_values)
        x_fit = np.linspace(min(all_values), max(all_values), 100)
        y_fit = stats.norm.pdf(x_fit, mu, std)
        ax.plot(x_fit, y_fit, 'r-', linewidth=2, label=f'Normal Fit\nμ={mu:.4f}\nσ={std:.4f}')
        
        # 전체 LSL/USL 평균
        if lsl_values and usl_values:
            avg_lsl = np.mean(lsl_values)
            avg_usl = np.mean(usl_values)
            ax.axvline(avg_lsl, color='red', linestyle='--', linewidth=2, label=f'Avg LSL: {avg_lsl:.4f}')
            ax.axvline(avg_usl, color='green', linestyle='--', linewidth=2, label=f'Avg USL: {avg_usl:.4f}')
        
        ax.set_title("All Measurements Distribution with Normal Fit", fontsize=14, fontweight='bold')
        ax.set_xlabel("Value")
        ax.set_ylabel("Density")
        ax.legend(loc='upper right')
        ax.grid(True, linestyle="--", alpha=0.4)
        
        path = os.path.join(plots_dir, f"LSLUSL_Distribution_{label}.png")
        plt.tight_layout()
        plt.savefig(path, dpi=200)
        plt.close()
        saved_paths.append(path)

    # 3) Scatter Plot (NET별 측정값 분포)
    fig, ax = plt.subplots(figsize=(14, 6))
    
    for net_idx in range(len(data_df)):
        values = data_df.iloc[net_idx].dropna().values
        x_scatter = np.full(len(values), net_idx + 1)
        ax.scatter(x_scatter, values, alpha=0.3, s=10)
    
    # 평균선
    ax.plot(range(1, len(means)+1), means, 'r-', linewidth=2, label='Mean', zorder=5)
    
    # LSL/USL
    if lsl_values and usl_values:
        ax.plot(range(1, len(lsl_values)+1), lsl_values, 'g--', linewidth=1.5, label='LSL')
        ax.plot(range(1, len(usl_values)+1), usl_values, 'b--', linewidth=1.5, label='USL')
    
    ax.set_title("Measurement Scatter Plot by NET", fontsize=14, fontweight='bold')
    ax.set_xlabel("NET No")
    ax.set_ylabel("Value")
    ax.legend(loc='upper right')
    ax.grid(True, linestyle="--", alpha=0.4)
    
    path = os.path.join(plots_dir, f"LSLUSL_Scatter_{label}.png")
    plt.tight_layout()
    plt.savefig(path, dpi=200)
    plt.close()
    saved_paths.append(path)

    # 4) Cpk/Ppk 지표 차트
    if lsl_values and usl_values:
        cpk_values = []
        ppk_values = []
        
        for net_idx in range(len(data_df)):
            values = data_df.iloc[net_idx].dropna().values
            if len(values) > 1 and net_idx < len(lsl_values):
                mean = np.mean(values)
                std = np.std(values)
                lsl = lsl_values[net_idx]
                usl = usl_values[net_idx]
                
                if std > 0:
                    cpu = (usl - mean) / (3 * std)
                    cpl = (mean - lsl) / (3 * std)
                    cpk = min(cpu, cpl)
                    cpk_values.append(cpk)
                else:
                    cpk_values.append(np.nan)
            else:
                cpk_values.append(np.nan)
        
        cpk_values = np.array(cpk_values)
        valid_cpk = cpk_values[~np.isnan(cpk_values)]
        
        if len(valid_cpk) > 0:
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
            
            # Cpk by NET
            colors = ['green' if c >= 1.33 else 'orange' if c >= 1.0 else 'red' for c in cpk_values if not np.isnan(c)]
            valid_idx = [i for i, c in enumerate(cpk_values) if not np.isnan(c)]
            ax1.bar(np.array(valid_idx) + 1, valid_cpk, color=colors, edgecolor='black')
            ax1.axhline(y=1.33, color='green', linestyle='--', label='Cpk ≥ 1.33 (Excellent)')
            ax1.axhline(y=1.0, color='orange', linestyle='--', label='Cpk ≥ 1.0 (Acceptable)')
            ax1.set_title("Cpk by NET", fontsize=12, fontweight='bold')
            ax1.set_xlabel("NET No")
            ax1.set_ylabel("Cpk")
            ax1.legend(loc='upper right', fontsize=8)
            ax1.grid(True, axis='y', linestyle='--', alpha=0.4)
            
            # Cpk 분포 히스토그램
            ax2.hist(valid_cpk, bins=20, color='steelblue', edgecolor='black', alpha=0.7)
            ax2.axvline(1.33, color='green', linestyle='--', linewidth=2, label='Cpk=1.33')
            ax2.axvline(1.0, color='orange', linestyle='--', linewidth=2, label='Cpk=1.0')
            ax2.axvline(np.mean(valid_cpk), color='red', linestyle='-', linewidth=2, 
                       label=f'Mean Cpk={np.mean(valid_cpk):.3f}')
            ax2.set_title("Cpk Distribution", fontsize=12, fontweight='bold')
            ax2.set_xlabel("Cpk")
            ax2.set_ylabel("Count")
            ax2.legend(loc='upper right', fontsize=8)
            ax2.grid(True, linestyle='--', alpha=0.4)
            
            plt.suptitle("Process Capability Analysis", fontsize=14, fontweight='bold')
            
            path = os.path.join(plots_dir, f"LSLUSL_Cpk_{label}.png")
            plt.tight_layout()
            plt.savefig(path, dpi=200)
            plt.close()
            saved_paths.append(path)

    # 5) Pass/Fail 파이 차트
    if lsl_values and usl_values:
        pass_count = 0
        fail_count = 0
        
        for net_idx in range(len(data_df)):
            values = data_df.iloc[net_idx].dropna().values
            if net_idx < len(lsl_values):
                lsl = lsl_values[net_idx]
                usl = usl_values[net_idx]
                for v in values:
                    if lsl <= v <= usl:
                        pass_count += 1
                    else:
                        fail_count += 1
        
        total = pass_count + fail_count
        if total > 0:
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
            
            # 파이 차트
            sizes = [pass_count, fail_count]
            labels = ['Pass', 'Fail']
            colors = ['#70AD47', '#C00000']
            explode = (0, 0.1)
            
            ax1.pie(sizes, explode=explode, labels=labels, colors=colors, autopct='%1.1f%%',
                   shadow=True, startangle=90)
            ax1.set_title(f"Pass/Fail Ratio\n(Total: {total:,} measurements)", fontsize=12, fontweight='bold')
            
            # 막대 차트
            ax2.bar(['Pass', 'Fail'], [pass_count, fail_count], color=colors, edgecolor='black')
            ax2.set_title("Pass/Fail Count", fontsize=12, fontweight='bold')
            ax2.set_ylabel("Count")
            
            for i, (count, pct) in enumerate(zip([pass_count, fail_count], [pass_count/total*100, fail_count/total*100])):
                ax2.annotate(f'{count:,}\n({pct:.1f}%)',
                            xy=(i, count), xytext=(0, 5),
                            textcoords="offset points", ha='center', va='bottom', fontsize=10)
            
            ax2.grid(True, axis='y', linestyle='--', alpha=0.4)
            
            path = os.path.join(plots_dir, f"LSLUSL_PassFail_{label}.png")
            plt.tight_layout()
            plt.savefig(path, dpi=200)
            plt.close()
            saved_paths.append(path)

    # 6) Statistics Summary Table
    stats_data = []
    for net_idx in range(min(20, len(data_df))):  # 최대 20개 NET만
        values = data_df.iloc[net_idx].dropna().values
        if len(values) > 0:
            lsl = lsl_values[net_idx] if lsl_values and net_idx < len(lsl_values) else None
            usl = usl_values[net_idx] if usl_values and net_idx < len(usl_values) else None
            
            stats_data.append({
                'NET': net_idx + 1,
                'Count': len(values),
                'Min': f"{np.min(values):.4f}",
                'Max': f"{np.max(values):.4f}",
                'Mean': f"{np.mean(values):.4f}",
                'Std': f"{np.std(values):.4f}",
                'LSL': f"{lsl:.4f}" if lsl else "N/A",
                'USL': f"{usl:.4f}" if usl else "N/A"
            })
    
    if stats_data:
        fig, ax = plt.subplots(figsize=(14, len(stats_data) * 0.4 + 2))
        ax.axis('off')
        
        df = pd.DataFrame(stats_data)
        table = ax.table(
            cellText=df.values,
            colLabels=df.columns,
            cellLoc='center',
            loc='center',
            colColours=['#4472C4'] * len(df.columns)
        )
        table.auto_set_font_size(False)
        table.set_fontsize(9)
        table.scale(1.2, 1.4)
        
        for i in range(len(df.columns)):
            table[(0, i)].set_text_props(color='white', fontweight='bold')
        
        ax.set_title("NET Statistics Summary (First 20 NETs)", fontsize=14, fontweight='bold', pad=20)
        
        path = os.path.join(plots_dir, f"LSLUSL_Stats_{label}.png")
        plt.tight_layout()
        plt.savefig(path, dpi=200, bbox_inches='tight')
        plt.close()
        saved_paths.append(path)

    # 7) 히스토그램: 표준편차 상위 top_k NET (개선)
    std_sorted = stds.sort_values(ascending=False).head(top_k)
    for rank, (net_idx, _) in enumerate(std_sorted.items(), 1):
        values = data_df.loc[net_idx].dropna().values
        if len(values) == 0:
            continue
        
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # 히스토그램
        n, bins, patches = ax.hist(values, bins=30, alpha=0.7, color="steelblue", 
                                   edgecolor="black", density=True)
        
        # 정규분포 피팅
        mu, std_val = stats.norm.fit(values)
        x_fit = np.linspace(min(values), max(values), 100)
        y_fit = stats.norm.pdf(x_fit, mu, std_val)
        ax.plot(x_fit, y_fit, 'r-', linewidth=2, label=f'Normal (μ={mu:.4f}, σ={std_val:.4f})')
        
        # LSL/USL 선
        lsl = lsl_values[net_idx] if lsl_values and net_idx < len(lsl_values) else None
        usl = usl_values[net_idx] if usl_values and net_idx < len(usl_values) else None
        if lsl is not None:
            ax.axvline(lsl, color="red", linestyle="--", linewidth=2, label=f"LSL={lsl:.4f}")
        if usl is not None:
            ax.axvline(usl, color="green", linestyle="--", linewidth=2, label=f"USL={usl:.4f}")
        
        # Cpk 계산
        if lsl is not None and usl is not None and std_val > 0:
            cpu = (usl - mu) / (3 * std_val)
            cpl = (mu - lsl) / (3 * std_val)
            cpk = min(cpu, cpl)
            ax.text(0.02, 0.98, f'Cpk: {cpk:.3f}', transform=ax.transAxes, 
                   fontsize=12, verticalalignment='top', fontweight='bold',
                   bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
        
        ax.set_title(f"NET {net_idx + 1} Distribution (Rank #{rank} by Std)", fontsize=14, fontweight='bold')
        ax.set_xlabel("Value")
        ax.set_ylabel("Density")
        ax.grid(True, linestyle="--", alpha=0.4)
        ax.legend(loc='upper right')
        
        path = os.path.join(plots_dir, f"LSLUSL_Hist_NET{net_idx + 1}_{label}.png")
        plt.tight_layout()
        plt.savefig(path, dpi=200)
        plt.close()
        saved_paths.append(path)

    # 8) Min/Max Range Chart
    fig, ax = plt.subplots(figsize=(14, 6))
    
    x = np.arange(len(means)) + 1
    ax.fill_between(x, mins, maxs, alpha=0.3, color='blue', label='Min-Max Range')
    ax.plot(x, means, 'r-', linewidth=2, marker='o', markersize=3, label='Mean')
    ax.plot(x, mins, 'b--', linewidth=1, alpha=0.7, label='Min')
    ax.plot(x, maxs, 'g--', linewidth=1, alpha=0.7, label='Max')
    
    ax.set_title("NET-wise Min/Max Range with Mean", fontsize=14, fontweight='bold')
    ax.set_xlabel("NET No")
    ax.set_ylabel("Value")
    ax.legend(loc='upper right')
    ax.grid(True, linestyle="--", alpha=0.4)
    
    path = os.path.join(plots_dir, f"LSLUSL_Range_{label}.png")
    plt.tight_layout()
    plt.savefig(path, dpi=200)
    plt.close()
    saved_paths.append(path)

    return saved_paths
