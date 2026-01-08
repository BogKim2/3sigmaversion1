"""
DE requirement 시트 생성 모듈
partpin 파일의 sheet2에서 continuity 데이터를 읽어서 가공
"""

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
import os


def auto_adjust_column_width(worksheet):
    """
    각 컬럼의 너비를 해당 컬럼에서 가장 긴 셀 내용에 맞게 자동 조절
    
    Args:
        worksheet: openpyxl worksheet 객체
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # 최소 너비 8, 최대 너비 100, 여유 공간 2 추가
        adjusted_width = min(max(max_length + 2, 8), 100)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def apply_header_style(worksheet, row_number, fill_color="4472C4"):
    """
    헤더 row에 스타일 적용 (파란색 배경)
    
    Args:
        worksheet: openpyxl worksheet 객체
        row_number: 스타일 적용할 row 번호
        fill_color: 배경색
    """
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    
    for cell in worksheet[row_number]:
        cell.fill = fill
        cell.font = white_font


def make_de_requirement_sheet(partpin_path: str, outfile_path: str) -> str:
    """
    partpin 파일의 sheet2에서 continuity 데이터를 읽어서
    outfile의 'DE requirement' sheet로 저장
    
    Args:
        partpin_path: partpin 파일 경로
        outfile_path: 출력 파일 경로
        
    Returns:
        결과 메시지
    """
    try:
        # partpin 파일 열기
        wb_partpin = load_workbook(partpin_path, data_only=True)
        
        # 디버그 정보
        debug_info = []
        debug_info.append(f"Sheets: {wb_partpin.sheetnames}")
        
        # sheet2 찾기 (두 번째 시트)
        if len(wb_partpin.sheetnames) < 2:
            wb_partpin.close()
            return f"Error: partpin file does not have sheet2. Found sheets: {wb_partpin.sheetnames}"
        
        # 두 번째 시트 선택
        sheet2_name = wb_partpin.sheetnames[1]
        sheet2 = wb_partpin[sheet2_name]
        debug_info.append(f"Sheet: {sheet2_name}")
        
        # "continuity" 헤더 셀을 정확히 찾기
        header_row = None
        continuity_col = None  # continuity 셀의 column 번호 (1-based)
        
        for row_idx in range(1, sheet2.max_row + 1):
            for col_idx in range(1, sheet2.max_column + 1):
                cell_val = sheet2.cell(row=row_idx, column=col_idx).value
                if cell_val and str(cell_val).lower() == "continuity":
                    header_row = row_idx
                    continuity_col = col_idx
                    break
            if header_row:
                break
        
        debug_info.append(f"Header row: {header_row}, Continuity col: {continuity_col}")
        
        if header_row is None:
            # 헤더를 못 찾으면 샘플 데이터 보여줌
            sample_data = []
            for row_idx in range(1, min(10, sheet2.max_row + 1)):
                row_vals = []
                for col_idx in range(1, min(10, sheet2.max_column + 1)):
                    val = sheet2.cell(row=row_idx, column=col_idx).value
                    row_vals.append(str(val)[:15] if val else "")
                sample_data.append(f"R{row_idx}: {row_vals}")
            wb_partpin.close()
            return f"Error: Could not find 'continuity' header.\nSample:\n" + "\n".join(sample_data)
        
        # 데이터 추출 (헤더 다음 행부터)
        # 컬럼 순서: continuity(col), NET(col+1), part(col+2), pin(col+3), part(col+4), pin(col+5)
        data_rows = []
        prev_net_name = ""  # 이전 행의 NET 이름 저장
        
        for row_idx in range(header_row + 1, sheet2.max_row + 1):
            # NET 컬럼 (continuity_col + 1)
            net_name = sheet2.cell(row=row_idx, column=continuity_col + 1).value
            
            # 각 컬럼 값 추출
            part1 = sheet2.cell(row=row_idx, column=continuity_col + 2).value or ""
            pin1 = sheet2.cell(row=row_idx, column=continuity_col + 3).value or ""
            part2 = sheet2.cell(row=row_idx, column=continuity_col + 4).value or ""
            pin2 = sheet2.cell(row=row_idx, column=continuity_col + 5).value or ""
            
            # part1 또는 part2가 없으면 완전히 빈 행으로 간주하고 건너뛰기
            if not part1 and not part2:
                continue
            
            # NET이 비어있는 경우: 이전 행의 NET + "-" + part2 형태로 자동 생성
            if not net_name:
                if prev_net_name and part2:
                    net_name = f"{prev_net_name}-{part2}"
                else:
                    net_name = part2 if part2 else "UNKNOWN"
            else:
                # 현재 NET을 이전 NET으로 저장
                prev_net_name = net_name
            
            # Part&Pin 형식으로 변환
            part_pin1 = f"{part1}.{pin1}" if part1 and pin1 else ""
            part_pin2 = f"{part2}.{pin2}" if part2 and pin2 else ""
            
            data_rows.append([
                len(data_rows) + 1,  # 행 번호
                net_name,
                part1,
                pin1,
                part2,
                pin2,
                part_pin1,
                part_pin2
            ])
        
        debug_info.append(f"Data rows: {len(data_rows)}")
        
        # === 첫 번째 sheet에서 Pin location과 Address image 테이블 추출 ===
        sheet1_name = wb_partpin.sheetnames[0]
        sheet1 = wb_partpin[sheet1_name]
        
        # Pin location 테이블 찾기
        pin_location_row = None
        pin_location_col = None
        address_image_row = None
        address_image_col = None
        
        for row_idx in range(1, sheet1.max_row + 1):
            for col_idx in range(1, sheet1.max_column + 1):
                cell_val = sheet1.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    cell_str = str(cell_val).lower().strip()
                    if "pin location" in cell_str:
                        pin_location_row = row_idx
                        pin_location_col = col_idx
                    elif "address image" in cell_str:
                        address_image_row = row_idx
                        address_image_col = col_idx
        
        debug_info.append(f"PinLoc: R{pin_location_row}C{pin_location_col}, AddrImg: R{address_image_row}C{address_image_col}")
        
        # Pin location 데이터 추출 (part, pin, X, Y)
        pin_location_data = []
        if pin_location_row:
            # 헤더는 pin_location_row + 1 행에 있음 (part, pin, X, Y)
            for row_idx in range(pin_location_row + 2, sheet1.max_row + 1):
                part = sheet1.cell(row=row_idx, column=pin_location_col).value
                if not part:  # 빈 행이면 종료
                    break
                pin = sheet1.cell(row=row_idx, column=pin_location_col + 1).value or ""
                x_val = sheet1.cell(row=row_idx, column=pin_location_col + 2).value or ""
                y_val = sheet1.cell(row=row_idx, column=pin_location_col + 3).value or ""
                pin_location_data.append([part, pin, x_val, y_val])
        
        # Address image 데이터 추출 (part, pin, address, address)
        # 중간에 빈 줄이 있어도 계속 읽음 (연속 10개 빈 줄이면 종료)
        address_image_data = []
        if address_image_row:
            # 헤더는 address_image_row + 1 행에 있음
            empty_row_count = 0
            for row_idx in range(address_image_row + 2, sheet1.max_row + 1):
                part = sheet1.cell(row=row_idx, column=address_image_col).value
                if not part:
                    empty_row_count += 1
                    if empty_row_count >= 10:  # 연속 10개 빈 줄이면 종료
                        break
                    continue
                empty_row_count = 0  # 데이터가 있으면 카운트 리셋
                pin = sheet1.cell(row=row_idx, column=address_image_col + 1).value or ""
                addr1 = sheet1.cell(row=row_idx, column=address_image_col + 2).value or ""
                addr2 = sheet1.cell(row=row_idx, column=address_image_col + 3).value or ""
                address_image_data.append([part, pin, addr1, addr2])
        
        debug_info.append(f"PinLoc data: {len(pin_location_data)}, AddrImg data: {len(address_image_data)}")
        
        # Address image 데이터를 딕셔너리로 변환 (part.pin -> [addr1, addr2])
        address_lookup = {}
        for row_data in address_image_data:
            part = row_data[0] if row_data[0] else ""
            pin = row_data[1] if row_data[1] else ""
            addr1 = row_data[2] if row_data[2] else ""
            addr2 = row_data[3] if row_data[3] else ""
            if part and pin:
                key = f"{part}.{pin}"
                address_lookup[key] = [addr1, addr2]
        
        debug_info.append(f"Address lookup: {len(address_lookup)} entries")
        
        wb_partpin.close()
        
        # 출력 파일 열기
        if os.path.exists(outfile_path):
            wb_out = load_workbook(outfile_path)
        else:
            wb_out = Workbook()
            if "Sheet" in wb_out.sheetnames:
                del wb_out["Sheet"]
        
        # 기존 DE requirement 시트가 있으면 삭제
        if "DE requirement" in wb_out.sheetnames:
            del wb_out["DE requirement"]
        
        # 새 DE requirement 시트 생성
        ws_de = wb_out.create_sheet("DE requirement")
        
        # A1에 "4 wire pair" 작성
        ws_de.cell(row=1, column=1, value="4 wire pair")
        
        # 첫 번째 데이터 행에서 part1, part2 이름 가져오기 (헤더용)
        part1_name = data_rows[0][2] if data_rows and len(data_rows[0]) > 2 else "part1"
        part2_name = data_rows[0][4] if data_rows and len(data_rows[0]) > 4 else "part2"
        
        # 헤더 작성 (2행) - 기존 8개 + address 4개 컬럼
        headers = ["", "NET", "part", "pin", "part", "pin", "Part &Pin", "Part &Pin",
                   f"{part1_name}_add1", f"{part1_name}_add2", f"{part2_name}_add1", f"{part2_name}_add2"]
        for col_idx, header in enumerate(headers, 1):
            ws_de.cell(row=2, column=col_idx, value=header)
        
        # 데이터 작성 (3행부터)
        for row_idx, row_data in enumerate(data_rows, 3):
            # 기존 8개 컬럼 데이터
            for col_idx, value in enumerate(row_data, 1):
                ws_de.cell(row=row_idx, column=col_idx, value=value)
            
            # Part&Pin에 해당하는 address 찾아서 추가 (I-L 컬럼, 9-12)
            part_pin1 = row_data[6] if len(row_data) > 6 else ""  # G 컬럼 (Part &Pin 1)
            part_pin2 = row_data[7] if len(row_data) > 7 else ""  # H 컬럼 (Part &Pin 2)
            
            # part_pin1의 address (I, J 컬럼)
            if part_pin1 and part_pin1 in address_lookup:
                addr1, addr2 = address_lookup[part_pin1]
                ws_de.cell(row=row_idx, column=9, value=addr1)
                ws_de.cell(row=row_idx, column=10, value=addr2)
            
            # part_pin2의 address (K, L 컬럼)
            if part_pin2 and part_pin2 in address_lookup:
                addr1, addr2 = address_lookup[part_pin2]
                ws_de.cell(row=row_idx, column=11, value=addr1)
                ws_de.cell(row=row_idx, column=12, value=addr2)
        
        # === I-L 컬럼(9-12)은 address 데이터로 이미 작성됨 ===
        # M-N 컬럼(13-14)은 빈칸으로 유지
        
        # === O 컬럼(15)에 part.pin 컬럼 추가 ===
        part_pin_col = 15  # O 컬럼
        
        # part.pin 헤더
        ws_de.cell(row=2, column=part_pin_col, value="Part.Pin")
        
        # part.pin 데이터 (3행부터) - Address image의 part와 pin을 결합
        for row_offset, row_data in enumerate(address_image_data):
            part = row_data[0] if row_data[0] else ""
            pin = row_data[1] if row_data[1] else ""
            part_pin = f"{part}.{pin}" if part and pin else ""
            ws_de.cell(row=3 + row_offset, column=part_pin_col, value=part_pin)
        
        # === P 컬럼(16)부터 Address image 테이블 추가 ===
        addr_img_start_col = 16  # P 컬럼
        
        # Address image 헤더
        ws_de.cell(row=1, column=addr_img_start_col, value="Address image")
        addr_img_headers = ["part", "pin", "address", "address"]
        for col_offset, header in enumerate(addr_img_headers):
            ws_de.cell(row=2, column=addr_img_start_col + col_offset, value=header)
        
        # Address image 데이터 (3행부터)
        for row_offset, row_data in enumerate(address_image_data):
            for col_offset, value in enumerate(row_data):
                ws_de.cell(row=3 + row_offset, column=addr_img_start_col + col_offset, value=value)
        
        # 컬럼 너비 자동 조절
        auto_adjust_column_width(ws_de)
        
        # 헤더 스타일 적용 (2행에 파란색)
        apply_header_style(ws_de, 2)
        
        # 파일 저장
        wb_out.save(outfile_path)
        wb_out.close()
        
        result_msg = f"Success: Created 'DE requirement' sheet with {len(data_rows)} rows\n"
        result_msg += "Debug: " + " | ".join(debug_info)
        return result_msg
        
    except Exception as e:
        import traceback
        return f"Error: {str(e)}\n{traceback.format_exc()}"

