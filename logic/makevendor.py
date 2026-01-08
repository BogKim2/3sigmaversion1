"""
Vendor 시트 생성 모듈
vendorspec 파일에서 cover page가 아닌 sheet를 읽어서 outfile의 vendor sheet로 저장
"""

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
import os


def auto_adjust_column_width(worksheet):
    """
    각 컬럼의 너비를 해당 컬럼에서 가장 긴 셀 내용에 맞게 자동 조절
    C 컬럼은 +10 추가
    
    Args:
        worksheet: openpyxl worksheet 객체
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        
        for cell in column_cells:
            try:
                if cell.value:
                    # 셀 값의 길이 계산 (문자열로 변환)
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # C 컬럼은 +10, 나머지는 +2
        if column_letter == 'C':
            extra_width = 10
        else:
            extra_width = 2
        
        # 최소 너비 8, 최대 너비 100
        adjusted_width = min(max(max_length + extra_width, 8), 100)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def apply_row_style(worksheet, row_numbers, fill_color="FF0000", skip_column_a=True):
    """
    특정 row에 배경색 적용
    
    Args:
        worksheet: openpyxl worksheet 객체
        row_numbers: 색상을 적용할 row 번호 리스트
        fill_color: 배경색 (기본값: 붉은색)
        skip_column_a: A열 스킵 여부 (기본값: True)
    """
    red_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)  # 흰색 글자, 굵게
    
    for row_num in row_numbers:
        for cell in worksheet[row_num]:
            # A열(column 1) 스킵 옵션
            if skip_column_a and cell.column == 1:
                continue
            cell.fill = red_fill
            cell.font = white_font


def make_vendor_sheet(vendorspec_path: str, outfile_path: str) -> str:
    """
    vendorspec 파일에서 cover page가 아닌 sheet를 찾아서
    outfile의 'vendor' sheet로 복사
    
    Args:
        vendorspec_path: vendorspec 파일 경로
        outfile_path: 출력 파일 경로
        
    Returns:
        결과 메시지
    """
    try:
        # vendorspec 파일 열기
        wb_vendor = load_workbook(vendorspec_path, data_only=True)
        
        # cover page가 아닌 sheet 찾기
        target_sheet = None
        target_sheet_name = None
        
        for sheet_name in wb_vendor.sheetnames:
            # cover page가 아닌 sheet 찾기 (대소문자 무시)
            if "cover" not in sheet_name.lower():
                target_sheet = wb_vendor[sheet_name]
                target_sheet_name = sheet_name
                break
        
        if target_sheet is None:
            wb_vendor.close()
            return "Error: No sheet found other than cover page"
        
        # 출력 파일이 존재하면 열고, 없으면 새로 생성
        if os.path.exists(outfile_path):
            wb_out = load_workbook(outfile_path)
        else:
            wb_out = Workbook()
            # 기본 시트 삭제 (나중에 vendor 시트를 추가할 것이므로)
            if "Sheet" in wb_out.sheetnames:
                del wb_out["Sheet"]
        
        # 기존 vendor 시트가 있으면 삭제
        if "vendor" in wb_out.sheetnames:
            del wb_out["vendor"]
        
        # 새 vendor 시트 생성
        ws_vendor = wb_out.create_sheet("vendor")
        
        # 데이터 복사 (B1부터 시작 - A열을 비우고 B열부터 데이터 복사)
        for row_idx, row in enumerate(target_sheet.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                # col_idx + 1로 B열부터 시작
                ws_vendor.cell(row=row_idx, column=col_idx + 1, value=cell.value)
        
        # "Design" 문자가 나오는 row 찾기
        design_row = None
        for row_idx in range(1, ws_vendor.max_row + 1):
            for col_idx in range(1, ws_vendor.max_column + 1):
                cell_val = ws_vendor.cell(row=row_idx, column=col_idx).value
                if cell_val and "Design" in str(cell_val):
                    design_row = row_idx
                    break
            if design_row:
                break
        
        # Design row 다음 row부터 끝까지 A column에 수식 추가
        # =G{row}&H{row}
        # 단, A1-A10은 빈칸으로 유지 (row 11부터 수식 추가)
        if design_row:
            start_row = max(design_row + 1, 11)  # 최소 11행부터 시작
            for row_idx in range(start_row, ws_vendor.max_row + 1):
                # A열에 수식 추가: =G{row}&H{row}
                ws_vendor.cell(row=row_idx, column=1, value=f"=G{row_idx}&H{row_idx}")
        
        # 컬럼 너비 자동 조절
        auto_adjust_column_width(ws_vendor)
        
        # A열 너비를 35로 고정
        ws_vendor.column_dimensions['A'].width = 35
        
        # 1번, 9번 row 붉은색 배경 적용 (A열 제외)
        apply_row_style(ws_vendor, [1, 9], skip_column_a=True)
        
        # 파일 저장
        wb_out.save(outfile_path)
        
        # 리소스 정리
        wb_vendor.close()
        wb_out.close()
        
        return f"Success: Copied sheet '{target_sheet_name}' to 'vendor' sheet in {outfile_path}"
        
    except Exception as e:
        return f"Error: {str(e)}"

