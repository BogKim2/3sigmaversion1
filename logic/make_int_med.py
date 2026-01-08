"""
int_med.xlsx 파일 생성 모듈
NET 파일에서 #4W 섹션을 파싱하여 그룹별 데이터를 엑셀 파일로 저장합니다.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from copy import copy
import os

from logic.file_reader import parse_4w_section


def apply_header_style(worksheet, row_number, fill_color="ED7D31", start_col=1, end_col=None):
    """
    지정된 행에 헤더 스타일 적용 (주황색 배경)
    """
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True)
    
    if end_col is None:
        end_col = worksheet.max_column
    
    for col in range(start_col, end_col + 1):
        cell = worksheet.cell(row=row_number, column=col)
        cell.fill = fill
        cell.font = font


def make_int_med_file(net_file_path: str, output_path: str = "int_med.xlsx") -> str:
    """
    NET 파일에서 #4W 섹션을 파싱하여 int_med.xlsx 파일을 생성합니다.
    
    Args:
        net_file_path: NET 파일 경로
        output_path: 출력 파일 경로 (기본값: int_med.xlsx)
    
    Returns:
        결과 메시지
    """
    try:
        debug_info = []
        
        # NET 파일 존재 확인
        if not net_file_path or not os.path.exists(net_file_path):
            return f"Error: NET file not found: {net_file_path}"
        
        # #4W 섹션 파싱
        groups, parse_debug = parse_4w_section(net_file_path)
        debug_info.extend(parse_debug)
        
        if "Error" in groups:
            return f"Error parsing NET file: {groups['Error']}\nDebug: " + "\n".join(debug_info)
        
        if not groups:
            return "Error: No #4W section found in NET file\nDebug: " + "\n".join(debug_info)
        
        debug_info.append(f"Found {len(groups)} groups")
        
        # 그룹 이름 정렬 (Group 1, Group 2, ...)
        sorted_groups = sorted(groups.keys(), key=lambda x: int(''.join(filter(str.isdigit, x)) or 0))
        
        # 최대 행 수 계산
        max_rows = max(len(groups[g]) for g in sorted_groups) if sorted_groups else 0
        debug_info.append(f"Max rows: {max_rows}")
        
        # 새 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "4W Data"
        
        # === Row 1: 헤더 (NO, Group 1, Group 2, ...) ===
        ws.cell(row=1, column=1, value="NO")
        
        col = 2  # B열부터 시작
        for group_name in sorted_groups:
            # 4개 컬럼 병합
            start_col_letter = get_column_letter(col)
            end_col_letter = get_column_letter(col + 3)
            ws.merge_cells(f'{start_col_letter}1:{end_col_letter}1')
            ws.cell(row=1, column=col, value=group_name)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')
            col += 4
        
        # Row 1 스타일 적용
        apply_header_style(ws, 1, fill_color="ED7D31", end_col=1 + len(sorted_groups) * 4)
        
        # === Row 2부터: 데이터 ===
        for row_idx in range(max_rows):
            row_num = 2 + row_idx
            
            # NO 컬럼
            ws.cell(row=row_num, column=1, value=row_idx + 1)
            
            # 각 그룹의 데이터
            col = 2
            for group_name in sorted_groups:
                group_data = groups[group_name]
                if row_idx < len(group_data):
                    for col_offset, value in enumerate(group_data[row_idx]):
                        ws.cell(row=row_num, column=col + col_offset, value=value)
                col += 4
        
        # 컬럼 너비 설정 (10으로 고정)
        max_col = 1 + len(sorted_groups) * 4
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 10
        
        # 파일 저장
        wb.save(output_path)
        wb.close()
        
        result_msg = f"Success: Created '{output_path}' with {len(sorted_groups)} groups, {max_rows} rows\n"
        result_msg += "Debug: " + " | ".join(debug_info)
        return result_msg
        
    except Exception as e:
        import traceback
        return f"Error: {str(e)}\n{traceback.format_exc()}"


def make_input_check_pin_final(outfile_path: str, int_med_path: str = "int_med.xlsx") -> str:
    """
    input check pin 시트를 outfile에 생성합니다.
    - input check pin interm 시트의 row 1-10을 복사
    - int_med.xlsx의 내용을 row 11부터 복사
    
    Args:
        outfile_path: 출력 파일 경로 (input check pin interm 시트가 있는 파일)
        int_med_path: int_med.xlsx 파일 경로
    
    Returns:
        결과 메시지
    """
    try:
        debug_info = []
        
        # outfile 존재 확인
        if not os.path.exists(outfile_path):
            return f"Error: Output file not found: {outfile_path}"
        
        # int_med.xlsx 존재 확인
        if not os.path.exists(int_med_path):
            return f"Error: int_med.xlsx not found: {int_med_path}"
        
        # outfile 열기
        wb_out = load_workbook(outfile_path)
        
        # input check pin interm 시트 확인
        if "input check pin interm" not in wb_out.sheetnames:
            wb_out.close()
            return "Error: 'input check pin interm' sheet not found. Please run Execute first."
        
        ws_source = wb_out["input check pin interm"]
        
        # 기존 input check pin 시트가 있으면 삭제
        if "input check pin" in wb_out.sheetnames:
            del wb_out["input check pin"]
        
        # 새 input check pin 시트 생성
        ws_final = wb_out.create_sheet("input check pin")
        
        # === input check pin interm 시트의 row 1-10 복사 ===
        # 병합된 셀 정보 먼저 복사
        for merged_range in ws_source.merged_cells.ranges:
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            if max_row <= 10:  # row 1-10만
                ws_final.merge_cells(str(merged_range))
        
        # 셀 데이터 및 스타일 복사 (row 1-10)
        for row in range(1, 11):
            for col in range(1, ws_source.max_column + 1):
                source_cell = ws_source.cell(row=row, column=col)
                target_cell = ws_final.cell(row=row, column=col)
                
                # 병합된 셀인지 확인 (MergedCell은 값을 쓸 수 없음)
                try:
                    # 값 복사
                    target_cell.value = source_cell.value
                    
                    # 스타일 복사
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.border = copy(source_cell.border)
                        target_cell.alignment = copy(source_cell.alignment)
                        target_cell.number_format = source_cell.number_format
                except AttributeError:
                    # MergedCell인 경우 스킵
                    pass
        
        debug_info.append("Copied rows 1-10 from 'input check pin interm'")
        
        # === int_med.xlsx의 내용을 row 11부터 복사 ===
        wb_int_med = load_workbook(int_med_path)
        ws_int_med = wb_int_med.active
        
        # 병합된 셀 정보 복사 (int_med의 row 1 -> final의 row 11로 오프셋)
        for merged_range in ws_int_med.merged_cells.ranges:
            # 새 범위 계산 (row + 10)
            new_min_row = merged_range.min_row + 10
            new_max_row = merged_range.max_row + 10
            min_col = merged_range.min_col
            max_col = merged_range.max_col
            
            new_range = f"{get_column_letter(min_col)}{new_min_row}:{get_column_letter(max_col)}{new_max_row}"
            ws_final.merge_cells(new_range)
        
        # int_med 데이터 복사 (row 11부터)
        int_med_rows = 0
        for row in range(1, ws_int_med.max_row + 1):
            target_row = row + 10  # row 11부터 시작
            int_med_rows += 1
            
            for col in range(1, ws_int_med.max_column + 1):
                source_cell = ws_int_med.cell(row=row, column=col)
                target_cell = ws_final.cell(row=target_row, column=col)
                
                # 병합된 셀인지 확인 (MergedCell은 값을 쓸 수 없음)
                try:
                    # 값 복사
                    target_cell.value = source_cell.value
                    
                    # 스타일 복사
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.border = copy(source_cell.border)
                        target_cell.alignment = copy(source_cell.alignment)
                        target_cell.number_format = source_cell.number_format
                except AttributeError:
                    # MergedCell인 경우 스킵
                    pass
        
        debug_info.append(f"Copied {int_med_rows} rows from 'int_med.xlsx' starting at row 11")
        
        wb_int_med.close()
        
        # 컬럼 너비 복사 (input check pin interm과 동일하게)
        for col in range(1, ws_source.max_column + 1):
            col_letter = get_column_letter(col)
            if ws_source.column_dimensions[col_letter].width:
                ws_final.column_dimensions[col_letter].width = ws_source.column_dimensions[col_letter].width
            else:
                ws_final.column_dimensions[col_letter].width = 10
        
        # 파일 저장
        wb_out.save(outfile_path)
        wb_out.close()
        
        result_msg = f"Success: Created 'input check pin' sheet\n"
        result_msg += "Debug: " + " | ".join(debug_info)
        return result_msg
        
    except Exception as e:
        import traceback
        return f"Error: {str(e)}\n{traceback.format_exc()}"

