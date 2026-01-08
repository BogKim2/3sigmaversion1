"""
Cover Page 생성 모듈
Excel 파일에 Cover Page 시트를 추가
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


def add_cover_page(output_file: str, operator_name: str, 
                   input_files: dict, output_file_path: str = None) -> str:
    """
    Excel 파일에 Cover Page 시트 추가
    
    Args:
        output_file: 출력 Excel 파일 경로
        operator_name: 작업자 이름
        input_files: 입력 파일 딕셔너리 {"파일유형": "파일경로", ...}
        output_file_path: 실제 출력 파일 경로 (없으면 output_file 사용)
    
    Returns:
        결과 메시지
    """
    try:
        # 파일이 존재하는지 확인
        if not os.path.exists(output_file):
            return f"Error: Output file not found: {output_file}"
        
        # 워크북 열기
        wb = openpyxl.load_workbook(output_file)
        
        # Cover Page 시트가 이미 있으면 삭제
        if "Cover Page" in wb.sheetnames:
            del wb["Cover Page"]
        
        # 새 Cover Page 시트 생성 (맨 앞에 삽입)
        ws = wb.create_sheet("Cover Page", 0)
        
        # 스타일 정의
        title_font = Font(name='Segoe UI', size=18, bold=True, color='1976D2')
        header_font = Font(name='Segoe UI', size=12, bold=True, color='424242')
        normal_font = Font(name='Segoe UI', size=12, color='424242')
        small_font = Font(name='Segoe UI', size=10, color='757575')
        
        # 헤더 배경색
        header_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        
        # 테두리
        thin_border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )
        
        # 열 너비 설정
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 60
        
        row = 2
        
        # === 프로그램 타이틀 ===
        ws.cell(row=row, column=2, value="DCR Format Converter")
        ws.cell(row=row, column=2).font = title_font
        ws.merge_cells(f'B{row}:C{row}')
        row += 2
        
        # === 프로그램 정보 섹션 ===
        ws.cell(row=row, column=2, value="Program Information")
        ws.cell(row=row, column=2).font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=3).fill = header_fill
        ws.merge_cells(f'B{row}:C{row}')
        row += 1
        
        # Version
        ws.cell(row=row, column=2, value="Version:")
        ws.cell(row=row, column=2).font = normal_font
        ws.cell(row=row, column=3, value="1.0")
        ws.cell(row=row, column=3).font = normal_font
        row += 1
        
        # Programmer
        ws.cell(row=row, column=2, value="Programmer:")
        ws.cell(row=row, column=2).font = normal_font
        ws.cell(row=row, column=3, value="Sangwoo Kim")
        ws.cell(row=row, column=3).font = normal_font
        row += 1
        
        # Acknowledgments
        ws.cell(row=row, column=2, value="Acknowledgments:")
        ws.cell(row=row, column=2).font = normal_font
        ws.cell(row=row, column=3, value="Lots of help from Opus4.5 and Gemini")
        ws.cell(row=row, column=3).font = normal_font
        row += 2
        
        # === 문서 정보 섹션 ===
        ws.cell(row=row, column=2, value="Document Information")
        ws.cell(row=row, column=2).font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=3).fill = header_fill
        ws.merge_cells(f'B{row}:C{row}')
        row += 1
        
        # Operator
        ws.cell(row=row, column=2, value="Operator:")
        ws.cell(row=row, column=2).font = normal_font
        ws.cell(row=row, column=3, value=operator_name if operator_name else "(Not specified)")
        ws.cell(row=row, column=3).font = normal_font
        row += 1
        
        # Created Date/Time
        created_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=row, column=2, value="Created:")
        ws.cell(row=row, column=2).font = normal_font
        ws.cell(row=row, column=3, value=created_datetime)
        ws.cell(row=row, column=3).font = normal_font
        row += 2
        
        # === 입력 파일 섹션 ===
        ws.cell(row=row, column=2, value="Input Files")
        ws.cell(row=row, column=2).font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=3).fill = header_fill
        ws.merge_cells(f'B{row}:C{row}')
        row += 1
        
        for file_type, file_path in input_files.items():
            ws.cell(row=row, column=2, value=f"{file_type}:")
            ws.cell(row=row, column=2).font = normal_font
            ws.cell(row=row, column=3, value=file_path if file_path else "(Not specified)")
            ws.cell(row=row, column=3).font = small_font
            row += 1
        
        row += 1
        
        # === 출력 파일 섹션 ===
        ws.cell(row=row, column=2, value="Output File")
        ws.cell(row=row, column=2).font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=3).fill = header_fill
        ws.merge_cells(f'B{row}:C{row}')
        row += 1
        
        actual_output = output_file_path if output_file_path else output_file
        ws.cell(row=row, column=2, value="File Path:")
        ws.cell(row=row, column=2).font = normal_font
        ws.cell(row=row, column=3, value=actual_output)
        ws.cell(row=row, column=3).font = small_font
        row += 1
        
        # 파일 이름만 추출
        file_name = os.path.basename(actual_output)
        ws.cell(row=row, column=2, value="File Name:")
        ws.cell(row=row, column=2).font = normal_font
        ws.cell(row=row, column=3, value=file_name)
        ws.cell(row=row, column=3).font = normal_font
        row += 2
        
        # === 시트 목록 섹션 ===
        ws.cell(row=row, column=2, value="Sheets in This File")
        ws.cell(row=row, column=2).font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=3).fill = header_fill
        ws.merge_cells(f'B{row}:C{row}')
        row += 1
        
        # Cover Page를 제외한 시트 목록
        sheet_list = [name for name in wb.sheetnames if name != "Cover Page"]
        for i, sheet_name in enumerate(sheet_list, 1):
            ws.cell(row=row, column=2, value=f"{i}.")
            ws.cell(row=row, column=2).font = normal_font
            ws.cell(row=row, column=3, value=sheet_name)
            ws.cell(row=row, column=3).font = normal_font
            row += 1
        
        # 행 높이 조정
        for r in range(1, row + 1):
            ws.row_dimensions[r].height = 20
        
        # 타이틀 행은 더 높게
        ws.row_dimensions[2].height = 30
        
        # 파일 저장
        wb.save(output_file)
        wb.close()
        
        return f"Success: Cover Page added to {os.path.basename(output_file)}"
        
    except Exception as e:
        return f"Error adding cover page: {str(e)}"

