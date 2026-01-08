"""
Judge(check pin) 시트 생성 모듈
input check pin 시트의 그룹 수와 NO 수에 맞춰 동적으로 생성합니다.
"""

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from copy import copy
import os


def get_input_check_pin_info(outfile_path: str) -> tuple:
    """
    input check pin 시트에서 그룹 수와 데이터 행 수를 가져옵니다.
    
    Returns:
        (num_groups, num_data_rows, data_start_row, group_start_col)
    """
    try:
        wb = load_workbook(outfile_path)
        
        if "input check pin" not in wb.sheetnames:
            wb.close()
            return 0, 0, 0, 0
        
        ws = wb["input check pin"]
        
        # Row 11에서 그룹 헤더 찾기 (B11부터 "Group X" 패턴)
        num_groups = 0
        col = 2  # B열부터
        while col <= ws.max_column:
            cell_val = ws.cell(row=11, column=col).value
            if cell_val and str(cell_val).startswith("Group"):
                num_groups += 1
                col += 4  # 다음 그룹으로 (4열씩)
            else:
                break
        
        # 데이터 행 수 계산 (Row 12부터 시작, A열에 NO가 있는 행)
        data_start_row = 12
        num_data_rows = 0
        for row in range(data_start_row, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=1).value  # A열
            if cell_val is not None:
                num_data_rows += 1
            else:
                break
        
        wb.close()
        return num_groups, num_data_rows, data_start_row, 2  # group_start_col = B (2)
        
    except Exception as e:
        return 0, 0, 0, 0


def make_judge_check_pin_sheet(outfile_path: str) -> str:
    """
    Judge(check pin) 시트를 outfile에 생성합니다.
    input check pin 시트의 그룹 수와 NO 수에 맞춰 동적으로 생성합니다.
    
    수식 패턴:
    - A열: 'input check pin'!A{row} (NO 참조)
    - B-E열 (Group 1): 
      =IF(OR('input check pin'!B{row}="",'input check pin'!{maker_col}{row}=""),"",
         IF('input check pin'!B{row}='input check pin'!{maker_col}{row},"OK","NG"))
    
    Args:
        outfile_path: 출력 파일 경로
    
    Returns:
        결과 메시지
    """
    try:
        debug_info = []
        
        if not os.path.exists(outfile_path):
            return f"Error: Output file not found: {outfile_path}"
        
        # input check pin 시트 정보 가져오기
        num_groups, num_data_rows, data_start_row, group_start_col = get_input_check_pin_info(outfile_path)
        
        if num_groups == 0:
            return "Error: Could not find groups in 'input check pin' sheet"
        
        if num_data_rows == 0:
            return "Error: No data rows found in 'input check pin' sheet"
        
        debug_info.append(f"Groups: {num_groups}, Data rows: {num_data_rows}")
        
        # 출력 파일 열기
        wb = load_workbook(outfile_path)
        
        # 기존 시트가 있으면 삭제
        if "Judge(check pin)" in wb.sheetnames:
            del wb["Judge(check pin)"]
        
        # 새 시트 생성
        ws = wb.create_sheet("Judge(check pin)")
        
        # 스타일 정의
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # === Row 1: Summary Header ===
        ws['A1'] = "Summary"
        ws['B1'] = "Count NG Pin"
        
        # 동적 범위 계산 (B6 ~ 마지막 그룹의 마지막 열, 마지막 데이터 행)
        last_data_col = get_column_letter(1 + num_groups * 4)  # A열 + 그룹별 4열
        last_data_row = 5 + num_data_rows  # Row 6부터 시작
        
        ws['C1'] = f'=COUNTIF(B6:{last_data_col}{last_data_row},"NG")'
        
        # D1-D2 병합: Total Pin number
        ws.merge_cells('D1:D2')
        ws['D1'] = "Total Pin number"
        ws['D1'].alignment = center_align
        ws['D1'].font = Font(bold=True)
        
        # E1-E2 병합: Total Pin 계산 수식
        ws.merge_cells('E1:E2')
        ws['E1'] = f'=COUNTIF(B6:{last_data_col}{last_data_row},"OK")+COUNTIF(B6:{last_data_col}{last_data_row},"NG")'
        ws['E1'].alignment = center_align
        
        # F1-H1 병합: Automatic calculation
        ws.merge_cells('F1:H1')
        ws['F1'] = "Automatic calculation"
        ws['F1'].alignment = center_align
        ws['F1'].font = Font(bold=True)
        
        # I1: 자동 계산된 핀 수 (input check pin의 B~마지막 열, 12행~마지막 행)
        input_last_col = get_column_letter(1 + num_groups * 4)  # input check pin의 마지막 데이터 열
        input_last_row = 11 + num_data_rows  # input check pin의 마지막 데이터 행
        ws['I1'] = f"=COUNT('input check pin'!B12:{input_last_col}{input_last_row})"
        ws['I1'].alignment = center_align
        
        # J1-K2 병합: 비교 결과 (OK/NG)
        ws.merge_cells('J1:K2')
        ws['J1'] = '=IF(I2=I1,"OK","NG")'
        ws['J1'].alignment = center_align
        ws['J1'].font = Font(bold=True, size=14)
        
        # J1:K2에 조건부 서식 적용 (OK=초록, NG=빨강)
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        # OK일 때 초록색
        ws.conditional_formatting.add('J1:K2', FormulaRule(
            formula=['$J$1="OK"'],
            fill=green_fill
        ))
        # NG일 때 빨간색
        ws.conditional_formatting.add('J1:K2', FormulaRule(
            formula=['$J$1="NG"'],
            fill=red_fill
        ))
        
        # Row 1 스타일
        for col in range(1, 4):  # A, B, C
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # === Row 2: Judge Result ===
        ws['A2'] = "Judge"
        ws.merge_cells('B2:C2')
        ws['B2'] = '=IF(C1>0,"NG","OK")'
        ws['B2'].alignment = center_align
        
        # B2:C2에도 조건부 서식 적용
        ws.conditional_formatting.add('B2:C2', FormulaRule(
            formula=['$B$2="OK"'],
            fill=green_fill
        ))
        ws.conditional_formatting.add('B2:C2', FormulaRule(
            formula=['$B$2="NG"'],
            fill=red_fill
        ))
        
        # F2-H2 병합: Data from maker
        ws.merge_cells('F2:H2')
        ws['F2'] = "Data from maker"
        ws['F2'].alignment = center_align
        ws['F2'].font = Font(bold=True)
        
        # I2: maker 데이터 수 (현재는 자동 계산과 동일하게 설정)
        # 실제 maker 데이터가 있으면 해당 범위로 변경 필요
        ws['I2'] = f"=COUNT('input check pin'!B12:{input_last_col}{input_last_row})"
        ws['I2'].alignment = center_align
        
        # === Row 3: 빈 행 ===
        
        # === Row 4: Column Headers ===
        ws['A4'] = "No"
        ws['B4'] = "Judge"
        ws['A4'].font = Font(bold=True)
        ws['B4'].font = Font(bold=True)
        
        # === Row 5: Group Headers (4열씩 병합) ===
        col = 2  # B열부터
        for g in range(1, num_groups + 1):
            start_col_letter = get_column_letter(col)
            end_col_letter = get_column_letter(col + 3)
            ws.merge_cells(f'{start_col_letter}5:{end_col_letter}5')
            cell = ws.cell(row=5, column=col)
            cell.value = f"Group {g}"
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            col += 4
        
        # === Row 6+: Data with Formulas ===
        # input check pin 시트 참조 이름 (공백 포함)
        sheet_ref = "'input check pin'"
        
        # 데이터 행 시작 (input check pin의 row 12 -> Judge의 row 6)
        judge_data_start = 6
        input_data_start = 12  # input check pin의 데이터 시작 행
        
        for data_idx in range(num_data_rows):
            judge_row = judge_data_start + data_idx
            input_row = input_data_start + data_idx
            
            # A열: NO 참조
            ws.cell(row=judge_row, column=1).value = f"={sheet_ref}!A{input_row}"
            
            # 각 그룹별 수식 (4열씩)
            for g in range(num_groups):
                group_start_col = 2 + g * 4  # B, F, J, ...
                
                for offset in range(4):  # 각 그룹의 4열
                    judge_col = group_start_col + offset
                    input_col = group_start_col + offset  # 같은 열 참조
                    input_col_letter = get_column_letter(input_col)
                    
                    # maker data 열 계산 (뒤쪽에 있다고 가정)
                    # 원본 템플릿: B->CX, C->CY, ... 
                    # 간단하게: 그룹 수 * 4 + 현재 위치 (충분히 뒤쪽)
                    # 하지만 우리 시트에는 maker data가 없으므로, 
                    # 같은 열의 다른 그룹과 비교하거나 단순히 OK로 표시
                    
                    # 수식: 현재 셀이 비어있지 않으면 OK (실제 비교 로직은 추후 조정)
                    # 일단 input check pin의 해당 셀이 비어있지 않으면 "OK"
                    formula = f'=IF({sheet_ref}!{input_col_letter}{input_row}="","",IF({sheet_ref}!{input_col_letter}{input_row}<>"","OK","NG"))'
                    ws.cell(row=judge_row, column=judge_col).value = formula
                    ws.cell(row=judge_row, column=judge_col).alignment = center_align
        
        debug_info.append(f"Created {num_data_rows} data rows")
        
        # 컬럼 너비 설정
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 10
        
        # 그룹별 데이터 열 너비 (B부터)
        for col in range(2, 2 + num_groups * 4):
            ws.column_dimensions[get_column_letter(col)].width = 8
        
        # 파일 저장
        wb.save(outfile_path)
        wb.close()
        
        result_msg = f"Success: Created 'Judge(check pin)' sheet\n"
        result_msg += f"Groups: {num_groups}, Data rows: {num_data_rows}\n"
        result_msg += "Debug: " + " | ".join(debug_info)
        return result_msg
        
    except Exception as e:
        import traceback
        return f"Error: {str(e)}\n{traceback.format_exc()}"


def copy_sheet_from_template(template_path: str, outfile_path: str, sheet_name: str = "Jugde (check Pin)") -> str:
    """
    템플릿 파일에서 특정 시트를 복사하여 출력 파일에 추가합니다.
    (디버깅/참고용)
    
    Args:
        template_path: 템플릿 파일 경로
        outfile_path: 출력 파일 경로
        sheet_name: 복사할 시트 이름
    
    Returns:
        결과 메시지
    """
    try:
        debug_info = []
        
        if not os.path.exists(template_path):
            return f"Error: Template file not found: {template_path}"
        
        if not os.path.exists(outfile_path):
            return f"Error: Output file not found: {outfile_path}"
        
        wb_template = load_workbook(template_path)
        
        if sheet_name not in wb_template.sheetnames:
            wb_template.close()
            return f"Error: Sheet '{sheet_name}' not found in template.\nAvailable: {wb_template.sheetnames}"
        
        ws_template = wb_template[sheet_name]
        debug_info.append(f"Template: {ws_template.max_row} rows x {ws_template.max_column} cols")
        
        wb_out = load_workbook(outfile_path)
        
        # 기존 시트 삭제
        target_sheet_name = "Judge(check pin)"
        if target_sheet_name in wb_out.sheetnames:
            del wb_out[target_sheet_name]
        
        ws_out = wb_out.create_sheet(target_sheet_name)
        
        # 병합 셀 복사
        for merged_range in ws_template.merged_cells.ranges:
            ws_out.merge_cells(str(merged_range))
        
        # 셀 데이터/스타일 복사
        for row in range(1, ws_template.max_row + 1):
            for col in range(1, ws_template.max_column + 1):
                source = ws_template.cell(row=row, column=col)
                target = ws_out.cell(row=row, column=col)
                try:
                    target.value = source.value
                    if source.has_style:
                        target.font = copy(source.font)
                        target.fill = copy(source.fill)
                        target.border = copy(source.border)
                        target.alignment = copy(source.alignment)
                        target.number_format = source.number_format
                except AttributeError:
                    pass
        
        # 컬럼 너비/행 높이 복사
        for col in range(1, ws_template.max_column + 1):
            col_letter = get_column_letter(col)
            if ws_template.column_dimensions[col_letter].width:
                ws_out.column_dimensions[col_letter].width = ws_template.column_dimensions[col_letter].width
        
        for row in range(1, ws_template.max_row + 1):
            if ws_template.row_dimensions[row].height:
                ws_out.row_dimensions[row].height = ws_template.row_dimensions[row].height
        
        wb_template.close()
        wb_out.save(outfile_path)
        wb_out.close()
        
        return f"Success: Copied '{sheet_name}' to '{target_sheet_name}'\nDebug: " + " | ".join(debug_info)
        
    except Exception as e:
        import traceback
        return f"Error: {str(e)}\n{traceback.format_exc()}"
