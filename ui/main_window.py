"""
메인 윈도우 UI 모듈
Material Design 스타일 탭 기반 인터페이스 제공
"""

from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QPushButton, QLineEdit, QTextEdit,
    QLabel, QFileDialog, QGroupBox, QInputDialog, QMessageBox,
    QFrame, QSizePolicy, QApplication, QComboBox,
    QRadioButton, QButtonGroup, QListWidget, QAbstractItemView
)
from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QFont, QColor, QPalette, QPixmap
import pandas as pd
from datetime import datetime
import os

from logic.file_reader import read_net_file, read_xlsx_file
from logic.makevendor import make_vendor_sheet
from logic.make_de_requirement import make_de_requirement_sheet
from logic.make_input_check_pin import make_input_check_pin_sheet
from logic.make_int_med import make_int_med_file, make_input_check_pin_final
from logic.make_judge_check_pin import make_judge_check_pin_sheet
from logic.make_dcr import make_dcr_sheet
from logic.make_form_measurement import create_form_measurement_file, fill_impedance_data, fill_impedance_data_from_files, fill_dimension_data, fill_lslusl_data
from logic.visualizer import save_dcr_plots_from_file, save_form_plots_from_workbook, save_lslusl_plots_from_data
from logic.calculate_lsl_usl import calculate_lsl_usl_full
from logic.config_manager import save_file_paths, load_file_paths, get_app_dir


# Material Design 스타일 시트
MATERIAL_STYLE = """
QMainWindow {
    background-color: #FAFAFA;
}

QWidget {
    font-family: 'Segoe UI', 'Roboto', 'Arial', sans-serif;
    font-size: 10pt;
}

/* 헤더 스타일 */
QFrame#header_frame {
    background-color: #1976D2;
    border: none;
    border-radius: 0px;
}

QLabel#header_title {
    color: white;
    font-size: 18pt;
    font-weight: bold;
}

QLabel#header_version {
    color: rgba(255, 255, 255, 0.87);
    font-size: 9pt;
}

QLabel#header_programmer {
    color: rgba(255, 255, 255, 0.7);
    font-size: 9pt;
}

/* 그룹박스 스타일 */
QGroupBox {
    font-weight: bold;
    color: #424242;
    border: 1px solid #E0E0E0;
    border-radius: 8px;
    margin-top: 12px;
    padding-top: 8px;
    background-color: white;
}

QGroupBox::title {
    subcontrol-origin: margin;
    left: 16px;
    padding: 0 8px;
    color: #1976D2;
}

/* 입력 필드 스타일 */
QLineEdit {
    border: 1px solid #BDBDBD;
    border-radius: 4px;
    padding: 8px 12px;
    background-color: white;
    color: #212121;
    selection-background-color: #1976D2;
}

QLineEdit:focus {
    border: 2px solid #1976D2;
    padding: 7px 11px;
}

QLineEdit:read-only {
    background-color: #F5F5F5;
}

/* 버튼 스타일 */
QPushButton {
    background-color: #1976D2;
    color: white;
    border: none;
    border-radius: 4px;
    padding: 8px 16px;
    font-weight: bold;
    min-width: 80px;
}

QPushButton:hover {
    background-color: #1565C0;
}

QPushButton:pressed {
    background-color: #0D47A1;
}

QPushButton:disabled {
    background-color: #BDBDBD;
}

QPushButton#browse_btn {
    background-color: #757575;
    min-width: 60px;
    padding: 8px 12px;
}

QPushButton#browse_btn:hover {
    background-color: #616161;
}

QPushButton#auto_execute_btn {
    background-color: #388E3C;
    font-size: 11pt;
    padding: 12px 32px;
    min-width: 150px;
}

QPushButton#auto_execute_btn:hover {
    background-color: #2E7D32;
}

QPushButton#auto_execute_btn:pressed {
    background-color: #1B5E20;
}

QPushButton#execute_btn {
    background-color: #1976D2;
    font-size: 10pt;
    padding: 10px 24px;
}

QPushButton#print_btn {
    background-color: #757575;
    font-size: 10pt;
    padding: 10px 24px;
}

/* 탭 위젯 스타일 */
QTabWidget::pane {
    border: 1px solid #E0E0E0;
    border-radius: 8px;
    background-color: white;
    top: -1px;
}

QTabBar::tab {
    background-color: #EEEEEE;
    color: #616161;
    border: 1px solid #E0E0E0;
    border-bottom: none;
    border-top-left-radius: 4px;
    border-top-right-radius: 4px;
    padding: 10px 20px;
    margin-right: 2px;
    font-weight: bold;
}

QTabBar::tab:selected {
    background-color: white;
    border-bottom: 2px solid #1976D2;
    color: #1976D2;
}

QTabBar::tab:hover:!selected {
    background-color: #E3F2FD;
}

/* 텍스트 출력 영역 */
QTextEdit {
    border: 1px solid #E0E0E0;
    border-radius: 4px;
    padding: 8px;
    background-color: #FAFAFA;
    font-family: 'Consolas', 'Courier New', monospace;
    font-size: 9pt;
}

/* 콤보박스 스타일 */
QComboBox {
    color: #212121;
    background-color: white;
    border: 1px solid #BDBDBD;
    border-radius: 4px;
    padding: 6px 12px;
}

QComboBox QAbstractItemView {
    color: #212121;
    background-color: white;
}

/* 라디오버튼 스타일 */
QRadioButton {
    color: #424242;
}

/* 리스트 위젯 스타일 */
QListWidget {
    color: #212121;
    background-color: white;
    border: 1px solid #BDBDBD;
    border-radius: 4px;
}

/* 레이블 스타일 */
QLabel {
    color: #424242;
}

QLabel#section_label {
    font-weight: bold;
    color: #1976D2;
    font-size: 10pt;
}

/* 스크롤바 스타일 */
QScrollBar:vertical {
    border: none;
    background-color: #F5F5F5;
    width: 10px;
    border-radius: 5px;
}

QScrollBar::handle:vertical {
    background-color: #BDBDBD;
    border-radius: 5px;
    min-height: 20px;
}

QScrollBar::handle:vertical:hover {
    background-color: #9E9E9E;
}
"""


class MainWindow(QMainWindow):
    """메인 윈도우 클래스"""
    
    # 프로그램 정보
    PROGRAM_NAME = "DCR Format Converter"
    VERSION = "1.1"
    PROGRAMMER = "Sangwoo Kim"
    ACKNOWLEDGMENTS = "Lots of help from Opus4.5 and Gemini"
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"{self.PROGRAM_NAME} v{self.VERSION}")
        self.setMinimumSize(1000, 1000)
        self.resize(1100, 1100)  # 초기 크기 설정
        
        # Material Design 스타일 적용
        self.setStyleSheet(MATERIAL_STYLE)
        
        # JSON에서 파일 경로 로드
        config = load_file_paths()
        self.net_file_path = config["net_file"]
        self.xlsx_file_path = config["vendorspec_file"]
        self.partpin_file_path = config["partpin_file"]
        self.outfile_path = config["outfile"]
        self.etching_dir_path = config.get("etching_dir", "")
        self.form_outfile_path = config.get("form_outfile", "Form_measurement_result.xlsx")
        self.dimension_file_path = config.get("dimension_file", "")
        self.dimension_sheet_name = config.get("dimension_sheet", "")
        self.lslusl_file_path = config.get("lslusl_file", "")
        self.merged_file_path = config.get("merged_file", "")
        self.operator_name = config.get("operator_name", "")
        self.item_name = config.get("item_name", "")
        self.item_code = config.get("item_code", "")
        self.output_base_dir = config.get("output_base_dir", "")
        
        # 진행 상황 로그
        self.progress_logs = []
        self.current_progress_timer = None
        
        self._setup_ui()
        self._load_saved_paths()
    
    def _log_progress(self, message: str, tab_index: int = None):
        """
        실시간 진행 상황 로그 추가 및 UI 업데이트
        
        Args:
            message: 로그 메시지
            tab_index: 업데이트할 탭 인덱스 (None이면 현재 선택된 탭)
        """
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}"
        self.progress_logs.append(log_entry)
        
        if tab_index is None:
            tab_index = self.tab_widget.currentIndex()
            
        # 모든 로그를 합쳐서 현재 탭의 텍스트 에디터 업데이트
        full_log = "\n".join(self.progress_logs)
        
        if tab_index == 0 and hasattr(self, 'output_text'):
            self.output_text.setText(full_log)
            self.output_text.verticalScrollBar().setValue(self.output_text.verticalScrollBar().maximum())
        elif tab_index == 1 and hasattr(self, 'form_output_text'):
            self.form_output_text.setText(full_log)
            self.form_output_text.verticalScrollBar().setValue(self.form_output_text.verticalScrollBar().maximum())
        elif tab_index == 2 and hasattr(self, 'lsl_output_text'):
            self.lsl_output_text.setText(full_log)
            self.lsl_output_text.verticalScrollBar().setValue(self.lsl_output_text.verticalScrollBar().maximum())
            
        QApplication.processEvents()
    
    def _clear_progress(self):
        """진행 로그 초기화"""
        self.progress_logs = []
    
    def _save_log_file(self, log_content: str, tab_name: str = ""):
        """로그를 파일로 저장 (output 폴더 내 plain ASCII .dat 파일)"""
        try:
            output_dir = self._get_output_dir()
                
            operator = self.operator_input.text().strip() if hasattr(self, 'operator_input') else "Unknown"
            if not operator:
                operator = "Unknown"
            
            date_str = datetime.now().strftime("%Y%m%d")
            
            # 로그 파일명 생성: log_{Operator}_{date}.dat
            log_filename = f"log_{operator}_{date_str}.dat"
            log_path = os.path.join(output_dir, log_filename)
            
            # 로그 파일 저장
            with open(log_path, 'w', encoding='ascii', errors='replace') as f:
                f.write(f"DCR Format Converter - Log File\n")
                f.write(f"{'=' * 60}\n")
                f.write(f"Operator: {operator}\n")
                f.write(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"{'=' * 60}\n\n")
                f.write(log_content)
            
            return f"Log saved: output/{log_filename}"
        except Exception as e:
            return f"Failed to save log: {str(e)}"
    
    def _get_output_dir(self) -> str:
        """
        출력 디렉토리 경로를 반환합니다.
        {output_base_dir}/{ItemName}_{ItemCode}/ 구조를 사용합니다.
        """
        # 사용자 지정 base dir, 없으면 기본 output 디렉토리
        base_dir = self.output_dir_edit.text().strip() if hasattr(self, 'output_dir_edit') and self.output_dir_edit.text().strip() else os.path.join(get_app_dir(), "output")
        
        # Item Name + Item Code 폴더명
        folder_name = self._get_output_folder_name()
        if folder_name and not folder_name.startswith("("):
            output_dir = os.path.join(base_dir, folder_name)
        else:
            output_dir = base_dir
        
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        return output_dir
    
    def _get_output_filename(self, base_name: str, extension: str = ".xlsx", 
                              suffix_type: str = "operator_date") -> str:
        """
        출력 파일명 생성 (output 폴더 경로 포함)
        
        Args:
            base_name: 기본 파일명
            extension: 확장자
            suffix_type: "operator_date" (Operator_날짜), "final" (_final만), "none" (접미사 없음)
        """
        output_dir = self._get_output_dir()
            
        operator = self.operator_input.text().strip() if hasattr(self, 'operator_input') else ""
        date_str = datetime.now().strftime("%Y%m%d")
        
        # 파일명만 추출 (경로가 포함되어 있을 수 있음)
        pure_base_name = os.path.basename(base_name)
        
        # 확장자 제거 후 베이스 이름 추출
        if pure_base_name.endswith(extension):
            name_without_ext = pure_base_name[:-len(extension)]
        else:
            name_without_ext = pure_base_name
            
        # suffix_type에 따라 파일명 생성
        if suffix_type == "final":
            if name_without_ext.endswith("_final"):
                final_name = f"{name_without_ext}{extension}"
            else:
                final_name = f"{name_without_ext}_final{extension}"
            
        elif suffix_type == "none":
            final_name = f"{name_without_ext}{extension}"
            
        else:  # operator_date
            # 이미 날짜가 붙어있는지 확인
            import re
            date_pattern = r"_\d{8}$"
            if re.search(date_pattern, name_without_ext):
                final_name = f"{name_without_ext}{extension}"
            else:
                if operator:
                    final_name = f"{name_without_ext}_{operator}_{date_str}{extension}"
                else:
                    final_name = f"{name_without_ext}_{date_str}{extension}"
                    
        return os.path.join(output_dir, final_name)
    
    def _setup_ui(self):
        """UI 구성"""
        # 중앙 위젯
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 메인 레이아웃
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # 헤더 프레임
        header_frame = QFrame()
        header_frame.setObjectName("header_frame")
        header_frame.setFixedHeight(100)  # 높이 약간 확대
        header_layout = QHBoxLayout(header_frame)  # 수평 레이아웃으로 변경
        header_layout.setContentsMargins(30, 15, 30, 15)
        header_layout.setSpacing(20)
        
        # 로고 섹션 (왼쪽)
        logo_label = QLabel()
        logo_label.setFixedSize(70, 70)
        logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path).scaled(70, 70, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            logo_label.setPixmap(pixmap)
        else:
            # 로고가 없을 경우 텍스트로 대체 (Sumitomo Group Emblem 스타일)
            logo_label.setText("SEI")
            logo_label.setStyleSheet("color: white; font-weight: bold; font-size: 24pt; border: 3px solid white; padding: 5px;")
            logo_label.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(logo_label)
        
        # 텍스트 정보 섹션 (중앙)
        text_info_layout = QVBoxLayout()
        header_title = QLabel(self.PROGRAM_NAME)
        header_title.setObjectName("header_title")
        text_info_layout.addWidget(header_title)
        
        header_version = QLabel(f"Sumitomo Electric Industries | Version {self.VERSION}")
        header_version.setObjectName("header_version")
        text_info_layout.addWidget(header_version)
        header_layout.addLayout(text_info_layout)
        
        header_layout.addStretch()
        
        # 개발자 정보 섹션 (오른쪽)
        header_programmer = QLabel(f"Programmed by {self.PROGRAMMER}")
        header_programmer.setObjectName("header_programmer")
        header_programmer.setAlignment(Qt.AlignBottom | Qt.AlignRight)
        header_layout.addWidget(header_programmer)
        
        main_layout.addWidget(header_frame)
        
        # 컨텐츠 영역
        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(16, 16, 16, 16)
        content_layout.setSpacing(12)
        
        # === 공통 설정 영역 (탭 위에 배치) ===
        common_group = QGroupBox("Common Settings")
        common_layout = QVBoxLayout(common_group)
        
        # Row 1: Operator Name + Auto Execute
        operator_layout = QHBoxLayout()
        operator_label = QLabel("Operator Name:")
        operator_label.setFixedWidth(120)
        operator_label.setStyleSheet("font-weight: bold; color: #1976D2;")
        self.operator_input = QLineEdit()
        self.operator_input.setPlaceholderText("Enter your name...")
        self.operator_input.setText(self.operator_name)
        self.operator_input.textChanged.connect(self._on_operator_changed)
        operator_layout.addWidget(operator_label)
        operator_layout.addWidget(self.operator_input)
        operator_layout.addSpacing(20)
        
        # Auto Execute 버튼
        self.auto_execute_btn = QPushButton("Auto Execute All")
        self.auto_execute_btn.setObjectName("auto_execute_btn")
        self.auto_execute_btn.setToolTip("Executes all tabs sequentially (Tab1 → Tab2 → Tab3).\nThis may take several minutes. Please wait patiently.")
        self.auto_execute_btn.clicked.connect(self._auto_execute_all)
        operator_layout.addWidget(self.auto_execute_btn)
        common_layout.addLayout(operator_layout)
        
        # Row 2: Item Name + Item Code
        item_layout = QHBoxLayout()
        item_name_label = QLabel("Item Name:")
        item_name_label.setFixedWidth(120)
        item_name_label.setStyleSheet("font-weight: bold; color: #1976D2;")
        self.item_name_input = QLineEdit()
        self.item_name_input.setPlaceholderText("e.g., FPC")
        self.item_name_input.setText(self.item_name)
        self.item_name_input.textChanged.connect(self._on_item_info_changed)
        item_layout.addWidget(item_name_label)
        item_layout.addWidget(self.item_name_input)
        item_layout.addSpacing(20)
        
        item_code_label = QLabel("Item Code:")
        item_code_label.setFixedWidth(80)
        item_code_label.setStyleSheet("font-weight: bold; color: #1976D2;")
        self.item_code_input = QLineEdit()
        self.item_code_input.setPlaceholderText("e.g., 7S3493")
        self.item_code_input.setText(self.item_code)
        self.item_code_input.textChanged.connect(self._on_item_info_changed)
        item_layout.addWidget(item_code_label)
        item_layout.addWidget(self.item_code_input)
        common_layout.addLayout(item_layout)
        
        # Row 3: Output Directory
        output_dir_layout = QHBoxLayout()
        output_dir_label = QLabel("Output Directory:")
        output_dir_label.setFixedWidth(120)
        output_dir_label.setStyleSheet("font-weight: bold; color: #1976D2;")
        self.output_dir_edit = QLineEdit()
        self.output_dir_edit.setReadOnly(True)
        default_output_dir = self.output_base_dir if self.output_base_dir else os.path.join(get_app_dir(), "output")
        self.output_dir_edit.setText(default_output_dir)
        self.output_dir_edit.setPlaceholderText("Default: {app_dir}/output/")
        output_dir_browse_btn = QPushButton("Browse")
        output_dir_browse_btn.setObjectName("browse_btn")
        output_dir_browse_btn.clicked.connect(self._browse_output_directory)
        output_dir_layout.addWidget(output_dir_label)
        output_dir_layout.addWidget(self.output_dir_edit)
        output_dir_layout.addWidget(output_dir_browse_btn)
        common_layout.addLayout(output_dir_layout)
        
        # Row 4: 출력 폴더 미리보기
        preview_layout = QHBoxLayout()
        preview_label = QLabel("Output Folder:")
        preview_label.setFixedWidth(120)
        self.output_folder_preview = QLabel(self._get_output_folder_name())
        self.output_folder_preview.setStyleSheet("color: #388E3C; font-weight: bold; font-style: italic;")
        preview_layout.addWidget(preview_label)
        preview_layout.addWidget(self.output_folder_preview)
        preview_layout.addStretch()
        common_layout.addLayout(preview_layout)
        
        content_layout.addWidget(common_group)
        
        # 탭 위젯 생성
        self.tab_widget = QTabWidget()
        content_layout.addWidget(self.tab_widget)
        
        main_layout.addWidget(content_widget)
        
        # 첫 번째 탭: make DCR format
        self._create_dcr_tab()
        
        # 두 번째 탭: make Form Measurement Result file
        self._create_form_measurement_tab()
        
        # 세 번째 탭: calculate LSL USL
        self._create_lsl_usl_tab()

    def _create_dcr_tab(self):
        """make DCR format 탭 생성"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(12)
        
        # 파일 선택 그룹박스
        file_group = QGroupBox("Input Files")
        file_layout = QVBoxLayout(file_group)
        
        # .NET 파일 선택 (inputfile1)
        net_layout = QHBoxLayout()
        net_label = QLabel("NET File:")
        net_label.setFixedWidth(120)
        self.net_path_edit = QLineEdit()
        self.net_path_edit.setReadOnly(True)
        self.net_path_edit.setPlaceholderText("Select .NET file...")
        net_browse_btn = QPushButton("Browse")
        net_browse_btn.setObjectName("browse_btn")
        net_browse_btn.clicked.connect(self._browse_net_file)
        
        net_layout.addWidget(net_label)
        net_layout.addWidget(self.net_path_edit)
        net_layout.addWidget(net_browse_btn)
        file_layout.addLayout(net_layout)
        
        # .xlsx 파일 선택 (vendorspec)
        xlsx_layout = QHBoxLayout()
        xlsx_label = QLabel("Vendorspec File:")
        xlsx_label.setFixedWidth(120)
        self.xlsx_path_edit = QLineEdit()
        self.xlsx_path_edit.setReadOnly(True)
        self.xlsx_path_edit.setPlaceholderText("Select vendorspec .xlsx file...")
        xlsx_browse_btn = QPushButton("Browse")
        xlsx_browse_btn.setObjectName("browse_btn")
        xlsx_browse_btn.clicked.connect(self._browse_xlsx_file)
        
        xlsx_layout.addWidget(xlsx_label)
        xlsx_layout.addWidget(self.xlsx_path_edit)
        xlsx_layout.addWidget(xlsx_browse_btn)
        file_layout.addLayout(xlsx_layout)
        
        # .xlsx 파일 선택 (partpin)
        partpin_layout = QHBoxLayout()
        partpin_label = QLabel("Partpin File:")
        partpin_label.setFixedWidth(120)
        self.partpin_path_edit = QLineEdit()
        self.partpin_path_edit.setReadOnly(True)
        self.partpin_path_edit.setPlaceholderText("Select partpin .xlsx file...")
        partpin_browse_btn = QPushButton("Browse")
        partpin_browse_btn.setObjectName("browse_btn")
        partpin_browse_btn.clicked.connect(self._browse_partpin_file)
        
        partpin_layout.addWidget(partpin_label)
        partpin_layout.addWidget(self.partpin_path_edit)
        partpin_layout.addWidget(partpin_browse_btn)
        file_layout.addLayout(partpin_layout)
        
        layout.addWidget(file_group)
        
        # 출력 파일 그룹박스 (UI에는 정보를 표시하지만 사용자가 직접 수정하지 않도록 read-only로 설정 가능)
        output_file_group = QGroupBox("Output Settings")
        output_file_layout = QHBoxLayout(output_file_group)
        
        outfile_label = QLabel("Output Info:")
        outfile_label.setFixedWidth(120)
        self.outfile_info_label = QLabel("Filename will be auto-generated: DCR_format_yamaha_{Operator}_{Date}.xlsx")
        self.outfile_info_label.setStyleSheet("color: #666666; font-style: italic;")
        
        output_file_layout.addWidget(outfile_label)
        output_file_layout.addWidget(self.outfile_info_label)
        output_file_layout.addStretch()
        
        layout.addWidget(output_file_group)
        
        # 버튼 레이아웃
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        # Print 버튼
        self.print_btn = QPushButton("Print")
        self.print_btn.setObjectName("print_btn")
        self.print_btn.setToolTip("Displays the contents of selected input files in the output area.")
        self.print_btn.clicked.connect(self._print_files)
        btn_layout.addWidget(self.print_btn)
        
        # Execute 버튼
        self.execute_btn = QPushButton("Execute")
        self.execute_btn.setObjectName("execute_btn")
        self.execute_btn.setToolTip("Converts files and generates output sheets.\nThis process may take some time. Please wait patiently.")
        self.execute_btn.clicked.connect(self._execute)
        btn_layout.addWidget(self.execute_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # 출력 텍스트박스
        output_group = QGroupBox("Output Log")
        output_layout = QVBoxLayout(output_group)
        self.output_text = QTextEdit()
        self.output_text.setReadOnly(True)
        self.output_text.setPlaceholderText("Execution log will be displayed here...")
        self.output_text.setMinimumHeight(350)  # 로그 창 높이 확대
        output_layout.addWidget(self.output_text)
        layout.addWidget(output_group)
        
        # 탭에 추가
        self.tab_widget.addTab(tab, "make DCR format")
    
    def _on_operator_changed(self, text: str):
        """Operator 이름 변경 시 호출"""
        self.operator_name = text
        self._save_config()
    
    def _on_item_info_changed(self):
        """Item Name 또는 Item Code 변경 시 호출"""
        self.item_name = self.item_name_input.text().strip()
        self.item_code = self.item_code_input.text().strip()
        # 미리보기 업데이트
        if hasattr(self, 'output_folder_preview'):
            self.output_folder_preview.setText(self._get_output_folder_name())
        self._save_config()
    
    def _get_output_folder_name(self) -> str:
        """Item Name + Item Code로 출력 폴더명 생성"""
        name = self.item_name_input.text().strip() if hasattr(self, 'item_name_input') else self.item_name
        code = self.item_code_input.text().strip() if hasattr(self, 'item_code_input') else self.item_code
        if name and code:
            return f"{name}_{code}"
        elif name:
            return name
        elif code:
            return code
        else:
            return "(Enter Item Name and Item Code above)"
    
    def _browse_output_directory(self):
        """출력 기본 디렉토리 선택"""
        current_dir = self.output_dir_edit.text() if self.output_dir_edit.text() else ""
        dir_path = QFileDialog.getExistingDirectory(
            self,
            "Select Output Base Directory",
            current_dir
        )
        if dir_path:
            self.output_base_dir = dir_path
            self.output_dir_edit.setText(dir_path)
            self._save_config()
    
    def _load_saved_paths(self):
        """저장된 파일 경로를 UI에 반영"""
        if self.net_file_path:
            self.net_path_edit.setText(self.net_file_path)
        if self.xlsx_file_path:
            self.xlsx_path_edit.setText(self.xlsx_file_path)
        if self.partpin_file_path:
            self.partpin_path_edit.setText(self.partpin_file_path)
        # Tab 1의 outfile_path_edit가 제거되었으므로 관련 코드 삭제 또는 수정
        
        # Form Measurement 탭 경로 로드
        if self.etching_dir_path:
            self.etching_dir_edit.setText(self.etching_dir_path)
        # form_out_path_edit 제거됨
        if self.dimension_file_path:
            self.dimension_file_edit.setText(self.dimension_file_path)
            # 시트 목록 로드
            try:
                xl = pd.ExcelFile(self.dimension_file_path)
                sheet_names = xl.sheet_names
                xl.close()
                self.dimension_sheet_combo.blockSignals(True)
                self.dimension_sheet_combo.clear()
                self.dimension_sheet_combo.addItems(sheet_names)
                if self.dimension_sheet_name and self.dimension_sheet_name in sheet_names:
                    self.dimension_sheet_combo.setCurrentText(self.dimension_sheet_name)
                self.dimension_sheet_combo.blockSignals(False)
            except Exception:
                pass
        if self.lslusl_file_path:
            self.lslusl_file_edit.setText(self.lslusl_file_path)
    
    def _save_config(self):
        """현재 파일 경로를 JSON에 저장"""
        save_file_paths(
            self.net_file_path,
            self.xlsx_file_path,
            self.partpin_file_path,
            "DCR_format_yamaha.xlsx", # 기본 파일명 사용
            self.etching_dir_edit.text() if hasattr(self, 'etching_dir_edit') else "",
            "Form_measurement_result.xlsx", # 기본 파일명 사용
            self.dimension_file_edit.text() if hasattr(self, 'dimension_file_edit') else "",
            self.dimension_sheet_name if hasattr(self, 'dimension_sheet_name') else "",
            self.lslusl_file_edit.text() if hasattr(self, 'lslusl_file_edit') else "",
            self.merged_file_path if hasattr(self, 'merged_file_path') else "",
            self.operator_input.text() if hasattr(self, 'operator_input') else "",
            item_name=self.item_name_input.text().strip() if hasattr(self, 'item_name_input') else "",
            item_code=self.item_code_input.text().strip() if hasattr(self, 'item_code_input') else "",
            output_base_dir=self.output_dir_edit.text() if hasattr(self, 'output_dir_edit') else ""
        )
    
    def _browse_net_file(self):
        """.NET 파일 선택 다이얼로그"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select NET File",
            "",
            "NET Files (*.NET *.net);;All Files (*.*)"
        )
        if file_path:
            self.net_file_path = file_path
            self.net_path_edit.setText(file_path)
            self._save_config()
    
    def _browse_xlsx_file(self):
        """.xlsx 파일 선택 다이얼로그 (vendorspec)"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select XLSX File (vendorspec)",
            "",
            "Excel Files (*.xlsx);;All Files (*.*)"
        )
        if file_path:
            self.xlsx_file_path = file_path
            self.xlsx_path_edit.setText(file_path)
            self._save_config()
    
    def _browse_partpin_file(self):
        """.xlsx 파일 선택 다이얼로그 (partpin)"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select XLSX File (partpin)",
            "",
            "Excel Files (*.xlsx);;All Files (*.*)"
        )
        if file_path:
            self.partpin_file_path = file_path
            self.partpin_path_edit.setText(file_path)
            self._save_config()
    
    def _browse_outfile(self):
        """출력 파일 저장 위치 선택 다이얼로그 (더 이상 사용되지 않음)"""
        pass
    
    def _print_files(self):
        """선택된 파일들의 내용을 텍스트박스에 출력"""
        self._clear_progress()
        
        # .NET 파일 읽기
        if self.net_file_path:
            self._log_progress("=" * 60)
            self._log_progress("[ NET FILE (inputfile1) ]")
            self._log_progress("=" * 60)
            self._log_progress(f"Path: {self.net_file_path}")
            self._log_progress("-" * 60)
            net_content = read_net_file(self.net_file_path)
            for line in net_content.split('\n')[:50]:  # 처음 50줄만
                self._log_progress(line)
            if len(net_content.split('\n')) > 50:
                self._log_progress("... (truncated)")
            self._log_progress("")
        else:
            self._log_progress("[ NET FILE ] - No file selected")
            self._log_progress("")
        
        # .xlsx 파일 읽기 (vendorspec)
        if self.xlsx_file_path:
            self._log_progress("=" * 60)
            self._log_progress("[ XLSX FILE (vendorspec) ]")
            self._log_progress("=" * 60)
            self._log_progress(f"Path: {self.xlsx_file_path}")
            self._log_progress("-" * 60)
            xlsx_content = read_xlsx_file(self.xlsx_file_path)
            for line in xlsx_content.split('\n')[:30]:
                self._log_progress(line)
            self._log_progress("")
        else:
            self._log_progress("[ XLSX FILE (vendorspec) ] - No file selected")
            self._log_progress("")
        
        # .xlsx 파일 읽기 (partpin)
        if self.partpin_file_path:
            self._log_progress("=" * 60)
            self._log_progress("[ XLSX FILE (partpin) ]")
            self._log_progress("=" * 60)
            self._log_progress(f"Path: {self.partpin_file_path}")
            self._log_progress("-" * 60)
            partpin_content = read_xlsx_file(self.partpin_file_path)
            for line in partpin_content.split('\n')[:30]:
                self._log_progress(line)
            self._log_progress("")
        else:
            self._log_progress("[ XLSX FILE (partpin) ] - No file selected")
            self._log_progress("")
        
        # 출력 파일 정보 표시
        self._log_progress("=" * 60)
        self._log_progress("[ OUTPUT INFO ]")
        self._log_progress("=" * 60)
        self._log_progress("Output files will be automatically generated in the 'output/' directory.")
        self._log_progress("Naming rule: {OriginalName}_{Operator}_{Date}.xlsx")
    
    def _execute(self, for_auto_execute=False):
        """Execute 버튼 클릭 - 모든 시트 순차적 생성"""
        if not for_auto_execute:
            self._clear_progress()
        
        # 출력 파일 경로 자동 생성 (DCR_format_yamaha_{Operator}_{Date}.xlsx)
        current_outfile = self._get_output_filename("DCR_format_yamaha.xlsx")
        
        # Operator 이름 확인
        operator = self.operator_input.text().strip()
        if not operator:
            QMessageBox.warning(self, "Warning", "Please enter operator name before executing.")
            return
        
        self._log_progress(f"Starting DCR format conversion...")
        self._log_progress(f"Operator: {operator}")
        self._log_progress(f"Output file: {current_outfile}")
        self._log_progress("")
        
        # === Step 1: Make Vendor Sheet ===
        self._log_progress("=" * 60)
        self._log_progress("[ Step 1: Make Vendor Sheet ]")
        self._log_progress("=" * 60)
        
        if not self.xlsx_file_path:
            self._log_progress("Error: Please select vendorspec file first")
            return
        
        self._log_progress(f"Source: {self.xlsx_file_path}")
        self._log_progress(f"Output: {current_outfile}")
        
        result1 = make_vendor_sheet(self.xlsx_file_path, current_outfile)
        self._log_progress(result1)
        self._log_progress("")
        
        # === Step 2: Make DE Requirement Sheet ===
        self._log_progress("=" * 60)
        self._log_progress("[ Step 2: Make DE Requirement Sheet ]")
        self._log_progress("=" * 60)
        
        if not self.partpin_file_path:
            self._log_progress("Error: Please select partpin file first")
            return
        
        self._log_progress(f"Source: {self.partpin_file_path}")
        result2 = make_de_requirement_sheet(self.partpin_file_path, current_outfile)
        self._log_progress(result2)
        self._log_progress("")
        
        # === Step 3: Make Input Check Pin Sheet ===
        self._log_progress("=" * 60)
        self._log_progress("[ Step 3: Make Input Check Pin Sheet ]")
        self._log_progress("=" * 60)
        
        self._log_progress(f"Processing input check pin sheet...")
        result3 = make_input_check_pin_sheet(current_outfile, self.net_file_path)
        self._log_progress(result3)
        self._log_progress("")
        
        # === Step 4: Create int_med.xlsx ===
        self._log_progress("=" * 60)
        self._log_progress("[ Step 4: Create int_med.xlsx ]")
        self._log_progress("=" * 60)
        
        if not self.net_file_path:
            self._log_progress("Error: Please select NET file first")
            return
        
        self._log_progress(f"Processing NET file for int_med.xlsx...")
        result4 = make_int_med_file(self.net_file_path, "int_med.xlsx")
        self._log_progress(result4)
        self._log_progress("")
        
        # === Step 5: Create 'input check pin' sheet ===
        self._log_progress("=" * 60)
        self._log_progress("[ Step 5: Create 'input check pin' sheet ]")
        self._log_progress("=" * 60)
        
        self._log_progress(f"Merging input check pin data...")
        result5 = make_input_check_pin_final(current_outfile, "int_med.xlsx")
        self._log_progress(result5)
        self._log_progress("")
        
        # === Step 6: Create 'Judge(check pin)' sheet ===
        self._log_progress("=" * 60)
        self._log_progress("[ Step 6: Create 'Judge(check pin)' sheet ]")
        self._log_progress("=" * 60)
        
        self._log_progress(f"Creating Judge(check pin) sheet...")
        result6 = make_judge_check_pin_sheet(current_outfile)
        self._log_progress(result6)
        self._log_progress("")
        
        # === Step 7: Create 'DCR' sheet ===
        self._log_progress("=" * 60)
        self._log_progress("[ Step 7: Create 'DCR' sheet ]")
        self._log_progress("=" * 60)
        
        self._log_progress(f"Creating DCR sheet...")
        result7 = make_dcr_sheet(current_outfile)
        self._log_progress(result7)
        self._log_progress("")
        
        # === Step 8: Add Cover Page ===
        self._log_progress("=" * 60)
        self._log_progress("[ Step 8: Add Cover Page ]")
        self._log_progress("=" * 60)
        
        self._log_progress(f"Adding cover page...")
        from logic.cover_page import add_cover_page
        result8 = add_cover_page(
            current_outfile,
            operator,
            {
                "NET File": self.net_file_path,
                "Vendorspec File": self.xlsx_file_path,
                "Partpin File": self.partpin_file_path
            }
        )
        self._log_progress(result8)
        self._log_progress("")
        
        # === Step 9: Generate Plots ===
        self._log_progress("=" * 60)
        self._log_progress("[ Step 9: Generate Statistical Plots ]")
        self._log_progress("=" * 60)
        
        self._log_progress(f"Generating plots from output file...")
        try:
            plots = save_dcr_plots_from_file(current_outfile, operator, output_dir=self._get_output_dir())
            if plots:
                self._log_progress(f"Generated {len(plots)} plots:")
                for p in plots:
                    self._log_progress(f"  - {os.path.basename(p)}")
            else:
                self._log_progress("No plots generated (data not found or insufficient)")
        except Exception as e:
            self._log_progress(f"Warning: Plot generation failed - {str(e)}")
        
        self._log_progress("")
        self._log_progress("=" * 60)
        self._log_progress("All steps completed successfully!")
        self._log_progress(f"Output saved to: {current_outfile}")
        self._log_progress("=" * 60)
        
        # 개별 실행인 경우에만 로그 저장
        if not for_auto_execute:
            log_result = self._save_log_file("\n".join(self.progress_logs), "")
            self._log_progress(log_result)
        
        # 출력 파일 경로 업데이트 (DCR 파일 경로 저장 - 이후 탭에서 참조용)
        self.dcr_output_path = current_outfile
        self._save_config()
    
    # ========== Tab 2: Form Measurement Result ==========
    
    def _create_form_measurement_tab(self):
        """make Form Measurement Result file 탭 생성"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(12)
        
        # 입력 디렉토리 설정 그룹박스
        input_group = QGroupBox("Input Files")
        input_layout = QVBoxLayout(input_group)
        
        # === Etching 모드 선택 ===
        etching_mode_group = QGroupBox("Etching File Mode")
        etching_mode_layout = QVBoxLayout(etching_mode_group)
        
        # 라디오 버튼 (자동/수동)
        radio_layout = QHBoxLayout()
        self.etching_mode_group_btn = QButtonGroup(self)
        self.etching_auto_radio = QRadioButton("Auto (Directory Scan)")
        self.etching_manual_radio = QRadioButton("Manual (Select Files)")
        self.etching_auto_radio.setChecked(True)
        self.etching_mode_group_btn.addButton(self.etching_auto_radio, 0)
        self.etching_mode_group_btn.addButton(self.etching_manual_radio, 1)
        self.etching_auto_radio.toggled.connect(self._on_etching_mode_changed)
        radio_layout.addWidget(self.etching_auto_radio)
        radio_layout.addWidget(self.etching_manual_radio)
        radio_layout.addStretch()
        etching_mode_layout.addLayout(radio_layout)
        
        # --- 자동 모드 위젯 ---
        self.etching_auto_widget = QWidget()
        auto_layout = QHBoxLayout(self.etching_auto_widget)
        auto_layout.setContentsMargins(0, 0, 0, 0)
        etching_label = QLabel("Etching Directory:")
        etching_label.setFixedWidth(120)
        self.etching_dir_edit = QLineEdit()
        self.etching_dir_edit.setReadOnly(True)
        self.etching_dir_edit.setPlaceholderText("Select etching directory containing DK files...")
        etching_browse_btn = QPushButton("Browse")
        etching_browse_btn.setObjectName("browse_btn")
        etching_browse_btn.clicked.connect(self._browse_etching_directory)
        auto_layout.addWidget(etching_label)
        auto_layout.addWidget(self.etching_dir_edit)
        auto_layout.addWidget(etching_browse_btn)
        etching_mode_layout.addWidget(self.etching_auto_widget)
        
        # --- 수동 모드 위젯 ---
        self.etching_manual_widget = QWidget()
        manual_layout = QVBoxLayout(self.etching_manual_widget)
        manual_layout.setContentsMargins(0, 0, 0, 0)
        
        # 파일 리스트
        self.etching_file_list = QListWidget()
        self.etching_file_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.etching_file_list.setMinimumHeight(100)
        self.etching_file_list.setMaximumHeight(150)
        self.etching_file_list.setStyleSheet("QListWidget { font-size: 9pt; }")
        manual_layout.addWidget(self.etching_file_list)
        
        # 파일 추가/제거 버튼
        file_btn_layout = QHBoxLayout()
        add_files_btn = QPushButton("Add Files")
        add_files_btn.setObjectName("browse_btn")
        add_files_btn.clicked.connect(self._add_etching_files)
        remove_files_btn = QPushButton("Remove Selected")
        remove_files_btn.setObjectName("browse_btn")
        remove_files_btn.setStyleSheet("QPushButton { background-color: #C62828; } QPushButton:hover { background-color: #B71C1C; }")
        remove_files_btn.clicked.connect(self._remove_etching_files)
        clear_files_btn = QPushButton("Clear All")
        clear_files_btn.setObjectName("browse_btn")
        clear_files_btn.setStyleSheet("QPushButton { background-color: #E65100; } QPushButton:hover { background-color: #BF360C; }")
        clear_files_btn.clicked.connect(self._clear_etching_files)
        
        file_btn_layout.addWidget(add_files_btn)
        file_btn_layout.addWidget(remove_files_btn)
        file_btn_layout.addWidget(clear_files_btn)
        file_btn_layout.addStretch()
        manual_layout.addLayout(file_btn_layout)
        
        self.etching_manual_widget.setVisible(False)  # 초기에는 숨김
        etching_mode_layout.addWidget(self.etching_manual_widget)
        
        input_layout.addWidget(etching_mode_group)
        
        # Dimension 파일 선택
        dimension_layout = QHBoxLayout()
        dimension_label = QLabel("Dimension File:")
        dimension_label.setFixedWidth(120)
        self.dimension_file_edit = QLineEdit()
        self.dimension_file_edit.setReadOnly(True)
        self.dimension_file_edit.setPlaceholderText("Select dimension file (e.g., 7E3493-00003.xlsx)...")
        dimension_browse_btn = QPushButton("Browse")
        dimension_browse_btn.setObjectName("browse_btn")
        dimension_browse_btn.clicked.connect(self._browse_dimension_file)
        
        dimension_layout.addWidget(dimension_label)
        dimension_layout.addWidget(self.dimension_file_edit)
        dimension_layout.addWidget(dimension_browse_btn)
        input_layout.addLayout(dimension_layout)
        
        # Dimension Sheet 선택 (드롭다운)
        sheet_layout = QHBoxLayout()
        sheet_label = QLabel("Dimension Sheet:")
        sheet_label.setFixedWidth(120)
        self.dimension_sheet_combo = QComboBox()
        self.dimension_sheet_combo.setMinimumWidth(200)
        self.dimension_sheet_combo.setPlaceholderText("Select sheet after loading file...")
        self.dimension_sheet_combo.currentTextChanged.connect(self._on_dimension_sheet_changed)
        
        sheet_layout.addWidget(sheet_label)
        sheet_layout.addWidget(self.dimension_sheet_combo)
        sheet_layout.addStretch()
        input_layout.addLayout(sheet_layout)
        
        # LSLUSL 파일 선택
        lslusl_layout = QHBoxLayout()
        lslusl_label = QLabel("LSLUSL File:")
        lslusl_label.setFixedWidth(120)
        self.lslusl_file_edit = QLineEdit()
        self.lslusl_file_edit.setReadOnly(True)
        self.lslusl_file_edit.setPlaceholderText("Select LSLUSL file...")
        lslusl_browse_btn = QPushButton("Browse")
        lslusl_browse_btn.setObjectName("browse_btn")
        lslusl_browse_btn.clicked.connect(self._browse_lslusl_file)
        
        lslusl_layout.addWidget(lslusl_label)
        lslusl_layout.addWidget(self.lslusl_file_edit)
        lslusl_layout.addWidget(lslusl_browse_btn)
        input_layout.addLayout(lslusl_layout)
        
        layout.addWidget(input_group)
        
        # 출력 설정 그룹박스
        output_file_group = QGroupBox("Output Settings")
        output_file_layout = QHBoxLayout(output_file_group)
        
        outfile_label = QLabel("Output Info:")
        outfile_label.setFixedWidth(120)
        self.form_outfile_info_label = QLabel("Filename will be auto-generated: Form_measurement_result_{Operator}_{Date}.xlsx")
        self.form_outfile_info_label.setStyleSheet("color: #666666; font-style: italic;")
        
        output_file_layout.addWidget(outfile_label)
        output_file_layout.addWidget(self.form_outfile_info_label)
        output_file_layout.addStretch()
        
        layout.addWidget(output_file_group)
        
        # 버튼 그룹
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        self.form_execute_btn = QPushButton("Execute")
        self.form_execute_btn.setObjectName("execute_btn")
        self.form_execute_btn.setToolTip("Creates Form Measurement Result file from template.")
        self.form_execute_btn.clicked.connect(self._execute_form_measurement)
        btn_layout.addWidget(self.form_execute_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # 출력 텍스트 영역
        output_log_group = QGroupBox("Output Log")
        output_log_layout = QVBoxLayout(output_log_group)
        self.form_output_text = QTextEdit()
        self.form_output_text.setReadOnly(True)
        self.form_output_text.setPlaceholderText("Execution log will be displayed here...")
        self.form_output_text.setMinimumHeight(350)  # 로그 창 높이 확대
        output_log_layout.addWidget(self.form_output_text)
        layout.addWidget(output_log_group)
        
        # 탭에 추가
        self.tab_widget.addTab(tab, "make Form Measurement Result file")

    def _on_etching_mode_changed(self, checked):
        """에칭 모드 변경 시 UI 전환"""
        if not checked:
            return
        is_auto = self.etching_auto_radio.isChecked()
        self.etching_auto_widget.setVisible(is_auto)
        self.etching_manual_widget.setVisible(not is_auto)
    
    def _add_etching_files(self):
        """수동 모드: DK 파일 추가"""
        file_paths, _ = QFileDialog.getOpenFileNames(
            self,
            "Select DK Files",
            "",
            "Excel Files (*.xls *.xlsx);;All Files (*.*)"
        )
        if file_paths:
            for fp in file_paths:
                # 중복 방지
                existing = [self.etching_file_list.item(i).text() for i in range(self.etching_file_list.count())]
                if fp not in existing:
                    self.etching_file_list.addItem(fp)
    
    def _remove_etching_files(self):
        """수동 모드: 선택된 DK 파일 제거"""
        for item in self.etching_file_list.selectedItems():
            self.etching_file_list.takeItem(self.etching_file_list.row(item))
    
    def _clear_etching_files(self):
        """수동 모드: 모든 DK 파일 제거"""
        self.etching_file_list.clear()
    
    def _browse_etching_directory(self):
        """Etching 디렉토리 선택"""
        start_dir = self.etching_dir_edit.text() if self.etching_dir_edit.text() else ""
        dir_path = QFileDialog.getExistingDirectory(
            self,
            "Select Etching Directory",
            start_dir
        )
        if dir_path:
            self.etching_dir_edit.setText(dir_path)
            self._save_config()
    
    def _browse_dimension_file(self):
        """Dimension 파일 선택"""
        start_path = self.dimension_file_edit.text() if self.dimension_file_edit.text() else ""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Dimension File",
            start_path,
            "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            try:
                xl = pd.ExcelFile(file_path)
                sheet_names = xl.sheet_names
                xl.close()
                
                self.dimension_file_edit.setText(file_path)
                
                # 콤보박스에 시트 목록 추가
                self.dimension_sheet_combo.blockSignals(True)
                self.dimension_sheet_combo.clear()
                self.dimension_sheet_combo.addItems(sheet_names)
                
                # 이전에 저장된 시트가 있으면 선택
                if hasattr(self, 'dimension_sheet_name') and self.dimension_sheet_name in sheet_names:
                    self.dimension_sheet_combo.setCurrentText(self.dimension_sheet_name)
                elif sheet_names:
                    self.dimension_sheet_name = sheet_names[0]
                    self.dimension_sheet_combo.setCurrentIndex(0)
                
                self.dimension_sheet_combo.blockSignals(False)
                self._save_config()
                    
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to read Excel file:\n{str(e)}")
                return
    
    def _on_dimension_sheet_changed(self, sheet_name):
        """Dimension sheet 콤보박스 변경 이벤트"""
        if sheet_name:
            self.dimension_sheet_name = sheet_name
            self._save_config()
    
    def _browse_lslusl_file(self):
        """LSLUSL 파일 선택"""
        start_path = self.lslusl_file_edit.text() if self.lslusl_file_edit.text() else ""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select LSLUSL File",
            start_path,
            "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.lslusl_file_edit.setText(file_path)
            self._save_config()

    def _execute_form_measurement(self, for_auto_execute=False):
        """Form Measurement Result 파일 생성 실행"""
        if not for_auto_execute:
            self._clear_progress()
            
        # 출력 파일 경로 자동 생성 (Form_measurement_result_{Operator}_{Date}.xlsx)
        output_path = self._get_output_filename("Form_measurement_result.xlsx")
        
        etching_dir = self.etching_dir_edit.text()
        dimension_file = self.dimension_file_edit.text()
        lslusl_file = self.lslusl_file_edit.text()
        
        # === Step 1: 템플릿 복사하여 출력 파일 생성 ===
        self._log_progress("=" * 60, tab_index=1)
        self._log_progress("[ Step 1: Create Form Measurement Result File ]", tab_index=1)
        self._log_progress("=" * 60, tab_index=1)
        self._log_progress(f"Output: {output_path}", tab_index=1)
        
        result1 = create_form_measurement_file(output_path)
        self._log_progress(result1 if isinstance(result1, str) else result1.get("message", ""), tab_index=1)
        self._log_progress("", tab_index=1)
        
        tdr_map = {}
        dim_map = {}

        # === Step 2: DK 파일에서 Impedance 데이터 ===
        is_auto_mode = self.etching_auto_radio.isChecked()
        
        if is_auto_mode:
            # 자동 모드: 디렉토리 스캔
            if etching_dir:
                self._log_progress("=" * 60, tab_index=1)
                self._log_progress("[ Step 2: Fill Impedance Data from DK Files (Auto Mode) ]", tab_index=1)
                self._log_progress("=" * 60, tab_index=1)
                self._log_progress(f"Etching Directory: {etching_dir}", tab_index=1)
                
                result2 = fill_impedance_data(output_path, etching_dir)
                if isinstance(result2, dict):
                    tdr_map = result2.get("tdr_map", {})
                    self._log_progress(result2.get("message", ""), tab_index=1)
                else:
                    self._log_progress(result2, tab_index=1)
            else:
                self._log_progress("Note: No etching directory selected. Skipping DK file processing.", tab_index=1)
        else:
            # 수동 모드: 사용자가 선택한 파일 리스트
            file_count = self.etching_file_list.count()
            if file_count > 0:
                self._log_progress("=" * 60, tab_index=1)
                self._log_progress("[ Step 2: Fill Impedance Data from DK Files (Manual Mode) ]", tab_index=1)
                self._log_progress("=" * 60, tab_index=1)
                
                file_list = [self.etching_file_list.item(i).text() for i in range(file_count)]
                self._log_progress(f"Selected {len(file_list)} files:", tab_index=1)
                for fp in file_list:
                    self._log_progress(f"  - {os.path.basename(fp)}", tab_index=1)
                
                result2 = fill_impedance_data_from_files(output_path, file_list)
                if isinstance(result2, dict):
                    tdr_map = result2.get("tdr_map", {})
                    self._log_progress(result2.get("message", ""), tab_index=1)
                else:
                    self._log_progress(result2, tab_index=1)
            else:
                self._log_progress("Note: No DK files selected in manual mode. Skipping DK file processing.", tab_index=1)
        
        self._log_progress("", tab_index=1)
        
        # === Step 3: Dimension 데이터 ===
        if dimension_file:
            self._log_progress("=" * 60, tab_index=1)
            self._log_progress("[ Step 3: Fill Dimension Data (Circuit Width/Thickness) ]", tab_index=1)
            self._log_progress("=" * 60, tab_index=1)
            self._log_progress(f"Dimension File: {dimension_file}", tab_index=1)
            
            sheet_name = self.dimension_sheet_name if hasattr(self, 'dimension_sheet_name') else ""
            if not sheet_name:
                try:
                    xl = pd.ExcelFile(dimension_file)
                    sheet_names = xl.sheet_names
                    xl.close()
                    
                    if len(sheet_names) > 1:
                        selected_sheet, ok = QInputDialog.getItem(
                            self,
                            "Select Sheet",
                            f"Multiple sheets found ({len(sheet_names)}).\nSelect sheet:",
                            sheet_names,
                            0,
                            False
                        )
                        if ok and selected_sheet:
                            sheet_name = selected_sheet
                            self.dimension_sheet_name = sheet_name
                            self._save_config()
                        else:
                            self._log_progress("Cancelled: User did not select a sheet.", tab_index=1)
                            return
                    elif len(sheet_names) == 1:
                        sheet_name = sheet_names[0]
                        self.dimension_sheet_name = sheet_name
                        self._save_config()
                except Exception as e:
                    self._log_progress(f"Error reading dimension file sheets: {e}", tab_index=1)
                    return
            
            self._log_progress(f"Sheet: {sheet_name}", tab_index=1)
            result3 = fill_dimension_data(output_path, dimension_file, sheet_name)
            if isinstance(result3, dict):
                dim_map = result3.get("dim_map", {})
                self._log_progress(result3.get("message", ""), tab_index=1)
            else:
                self._log_progress(result3, tab_index=1)
        else:
            self._log_progress("Note: No dimension file selected. Skipping dimension processing.", tab_index=1)
        
        self._log_progress("", tab_index=1)
        
        # === Step 4: LSLUSL 데이터 ===
        if lslusl_file:
            self._log_progress("=" * 60, tab_index=1)
            self._log_progress("[ Step 4: Fill LSL/Center/USL Data ]", tab_index=1)
            self._log_progress("=" * 60, tab_index=1)
            self._log_progress(f"LSLUSL File: {lslusl_file}", tab_index=1)
            
            result4 = fill_lslusl_data(output_path, lslusl_file)
            self._log_progress(result4, tab_index=1)
        else:
            self._log_progress("Note: No LSLUSL file selected. Skipping LSL/USL processing.", tab_index=1)
        
        # === Step 5: Add Cover Page ===
        self._log_progress("", tab_index=1)
        self._log_progress("=" * 60, tab_index=1)
        self._log_progress("[ Step 5: Add Cover Page ]", tab_index=1)
        self._log_progress("=" * 60, tab_index=1)
        self._log_progress(f"Adding cover page...", tab_index=1)
        
        operator = self.operator_input.text().strip() if hasattr(self, 'operator_input') else ""
        from logic.cover_page import add_cover_page
        result5 = add_cover_page(
            output_path,
            operator,
            {
                "Etching Directory": etching_dir,
                "Dimension File": dimension_file,
                "LSLUSL File": lslusl_file
            }
        )
        self._log_progress(result5, tab_index=1)
        
        self._log_progress("", tab_index=1)
        self._log_progress("=" * 60, tab_index=1)
        self._log_progress("Form Measurement Result file created successfully!", tab_index=1)
        self._log_progress(f"Output saved to: {output_path}", tab_index=1)
        self._log_progress("=" * 60, tab_index=1)

        # === Visualization (PNG 저장) ===
        try:
            plots = save_form_plots_from_workbook(tdr_map, dim_map, operator, output_dir=self._get_output_dir())
            if plots:
                for p in plots:
                    self._log_progress(f"Plot saved: {p}", tab_index=1)
            else:
                self._log_progress("Note: No plot generated (insufficient data).", tab_index=1)
        except Exception as e:
            self._log_progress(f"Warning: Plot generation failed - {e}", tab_index=1)
        
        # 개별 실행인 경우에만 로그 저장
        if not for_auto_execute:
            log_result = self._save_log_file("\n".join(self.progress_logs), "")
            self._log_progress(log_result, tab_index=1)
        
        # 출력 경로 업데이트
        # self.form_out_path_edit.setText(output_path) # 에디트 박스 제거됨
        self._save_config()

    def _create_lsl_usl_tab(self):
        """calculate LSL USL 탭 생성"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(12)
        
        # 입력 파일 그룹
        input_group = QGroupBox("Input Files")
        input_layout = QVBoxLayout(input_group)
        
        # Merged File 선택
        merged_layout = QHBoxLayout()
        merged_label = QLabel("Merged File:")
        merged_label.setFixedWidth(120)
        self.merged_file_edit = QLineEdit()
        self.merged_file_edit.setReadOnly(True)
        self.merged_file_edit.setPlaceholderText("Select merged_file.xlsx...")
        if self.merged_file_path:
            self.merged_file_edit.setText(self.merged_file_path)
        merged_browse_btn = QPushButton("Browse")
        merged_browse_btn.setObjectName("browse_btn")
        merged_browse_btn.clicked.connect(self._browse_merged_file)
        merged_layout.addWidget(merged_label)
        merged_layout.addWidget(self.merged_file_edit)
        merged_layout.addWidget(merged_browse_btn)
        input_layout.addLayout(merged_layout)
        
        # DCR File (자동 입력 정보)
        dcr_layout = QHBoxLayout()
        dcr_label = QLabel("DCR File (Auto):")
        dcr_label.setFixedWidth(120)
        self.lsl_dcr_file_info = QLabel("Auto-selected from Tab 1 output")
        self.lsl_dcr_file_info.setStyleSheet("color: #1976D2; font-weight: bold;")
        dcr_layout.addWidget(dcr_label)
        dcr_layout.addWidget(self.lsl_dcr_file_info)
        dcr_layout.addStretch()
        input_layout.addLayout(dcr_layout)
        
        layout.addWidget(input_group)
        
        # 출력 설정 그룹박스
        output_file_group = QGroupBox("Output Settings")
        output_file_layout = QHBoxLayout(output_file_group)
        
        outfile_label = QLabel("Output Info:")
        outfile_label.setFixedWidth(120)
        self.lsl_outfile_info_label = QLabel("Filename will be auto-generated: Calculate_3Sigma_LSLUSL_final.xlsx")
        self.lsl_outfile_info_label.setStyleSheet("color: #666666; font-style: italic;")
        
        output_file_layout.addWidget(outfile_label)
        output_file_layout.addWidget(self.lsl_outfile_info_label)
        output_file_layout.addStretch()
        
        layout.addWidget(output_file_group)
        
        # 실행 버튼
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        execute_btn = QPushButton("Execute")
        execute_btn.setObjectName("execute_btn")
        execute_btn.setToolTip("Calculate LSL/USL statistics. This may take a while for large files.")
        execute_btn.clicked.connect(self._execute_lsl_usl)
        btn_layout.addWidget(execute_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # 출력 텍스트
        output_log_group = QGroupBox("Output Log")
        output_log_layout = QVBoxLayout(output_log_group)
        self.lsl_output_text = QTextEdit()
        self.lsl_output_text.setReadOnly(True)
        self.lsl_output_text.setPlaceholderText("Execution log will be displayed here...")
        self.lsl_output_text.setMinimumHeight(350)  # 로그 창 높이 확대
        output_log_layout.addWidget(self.lsl_output_text)
        layout.addWidget(output_log_group)
        
        self.tab_widget.addTab(tab, "calculate LSL USL")
    
    def _browse_merged_file(self):
        """Merged 파일 선택"""
        start_path = self.merged_file_edit.text() if self.merged_file_edit.text() else ""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Merged File",
            start_path,
            "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.merged_file_edit.setText(file_path)
            self.merged_file_path = file_path
            self._save_config()
    
    def _browse_lsl_dcr_file(self):
        """LSL 계산용 DCR 파일 선택"""
        start_path = self.lsl_dcr_file_edit.text() if self.lsl_dcr_file_edit.text() else ""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select DCR File",
            start_path,
            "Excel Files (*.xlsx)"
        )
        if file_path:
            self.lsl_dcr_file_edit.setText(file_path)
    
    def _browse_lsl_output_file(self):
        """LSL 출력 파일 선택 (더 이상 사용되지 않음)"""
        pass
    
    def _execute_lsl_usl(self, for_auto_execute=False):
        """LSL/USL 계산 실행"""
        if not for_auto_execute:
            self._clear_progress()
            
        merged_file = self.merged_file_edit.text()
        # DCR 파일은 Tab 1의 출력 파일을 자동으로 사용 (설정된 게 없으면 기본 파일명 사용)
        dcr_file = getattr(self, 'dcr_output_path', "output/DCR_format_yamaha.xlsx")
        # operator는 항상 미리 확보 (dcr_file 존재 여부와 무관)
        operator = self.operator_input.text().strip()
        
        # 파일이 실제로 존재하지 않으면 기본 경로 시도
        if not os.path.exists(dcr_file):
             # 현재 날짜와 Operator 이름이 포함된 최신 파일을 찾거나 기본 이름 사용
             date_str = datetime.now().strftime("%Y%m%d")
             possible_name = f"DCR_format_yamaha_{operator}_{date_str}.xlsx"
             possible_path = os.path.join("output", possible_name)
             if os.path.exists(possible_path):
                 dcr_file = possible_path
        
        if not merged_file:
            self._log_progress("Error: Please select a merged file.", tab_index=2)
            return
        
        if not os.path.exists(dcr_file):
            self._log_progress(f"Error: DCR file not found at {dcr_file}. Please execute Tab 1 first.", tab_index=2)
            return
        
        # Tab3 출력 파일은 Calculate_3Sigma_LSLUSL_final.xlsx 로 고정
        output_file = self._get_output_filename("Calculate_3Sigma_LSLUSL.xlsx", suffix_type="final")
        
        self._log_progress("=" * 60, tab_index=2)
        self._log_progress("[ Calculate LSL/USL Statistics ]", tab_index=2)
        self._log_progress("=" * 60, tab_index=2)
        self._log_progress(f"Merged File: {merged_file}", tab_index=2)
        self._log_progress(f"DCR File: {dcr_file}", tab_index=2)
        self._log_progress(f"Output File: {output_file}", tab_index=2)
        self._log_progress("-" * 60, tab_index=2)
        self._log_progress("Processing... This may take a while for large files.", tab_index=2)
        
        result = calculate_lsl_usl_full(merged_file, dcr_file, output_file, operator=operator)
        self._log_progress("", tab_index=2)
        self._log_progress(result, tab_index=2)
        
        # Add Cover Page
        self._log_progress("", tab_index=2)
        self._log_progress("=" * 60, tab_index=2)
        self._log_progress("[ Add Cover Page ]", tab_index=2)
        self._log_progress("=" * 60, tab_index=2)
        self._log_progress(f"Adding cover page...", tab_index=2)
        
        operator = self.operator_input.text().strip() if hasattr(self, 'operator_input') else ""
        from logic.cover_page import add_cover_page
        result_cover = add_cover_page(
            output_file,
            operator,
            {
                "Merged File": merged_file,
                "DCR File": dcr_file
            }
        )
        self._log_progress(result_cover, tab_index=2)
        
        # === Generate Statistical Plots ===
        self._log_progress("", tab_index=2)
        self._log_progress("=" * 60, tab_index=2)
        self._log_progress("[ Generate Statistical Plots ]", tab_index=2)
        self._log_progress("=" * 60, tab_index=2)
        
        try:
            # 출력 파일에서 데이터를 읽어 플롯 생성
            from openpyxl import load_workbook
            import pandas as pd
            import numpy as np
            
            wb = load_workbook(output_file, data_only=True)
            
            # 'tinh LCLUCL' 시트에서 데이터 추출 (행=NET, 열=측정값)
            if "tinh LCLUCL" in wb.sheetnames:
                ws = wb["tinh LCLUCL"]
                data_rows = []
                lsl_list = []
                usl_list = []
                
                # 데이터 읽기 (행 2부터, A열=NET, B~끝=측정값)
                for row in range(2, ws.max_row + 1):
                    row_data = []
                    for col in range(2, ws.max_column + 1):
                        val = ws.cell(row=row, column=col).value
                        if val is not None:
                            try:
                                row_data.append(float(val))
                            except (ValueError, TypeError):
                                pass
                    if row_data:
                        data_rows.append(row_data)
                        # LSL/USL 계산 (평균 ± 3σ)
                        arr = np.array(row_data)
                        avg = np.mean(arr)
                        std = np.std(arr)
                        lsl_list.append(max(0, avg - 3 * std))
                        usl_list.append(avg + 3 * std)
                
                if data_rows:
                    data_df = pd.DataFrame(data_rows)
                    plots = save_lslusl_plots_from_data(data_df, lsl_list, usl_list, operator, output_dir=self._get_output_dir())
                    if plots:
                        self._log_progress(f"Generated {len(plots)} plots:", tab_index=2)
                        for p in plots:
                            self._log_progress(f"  - {os.path.basename(p)}", tab_index=2)
                    else:
                        self._log_progress("No plots generated (insufficient data)", tab_index=2)
                else:
                    self._log_progress("No data found in 'tinh LCLUCL' sheet for plotting", tab_index=2)
            else:
                self._log_progress("Sheet 'tinh LCLUCL' not found - skipping plot generation", tab_index=2)
            
            wb.close()
        except Exception as e:
            self._log_progress(f"Warning: Plot generation failed - {str(e)}", tab_index=2)
        
        self._log_progress("", tab_index=2)
        self._log_progress("=" * 60, tab_index=2)
        self._log_progress("LSL/USL calculation completed!", tab_index=2)
        self._log_progress(f"Output saved to: {output_file}", tab_index=2)
        self._log_progress("=" * 60, tab_index=2)
        
        # 만약 개별 실행인 경우에만 여기서 로그 저장 (Auto Execute가 아니면)
        if not for_auto_execute:
            all_logs = "\n".join(self.progress_logs)
            log_result = self._save_log_file(all_logs, "")
            self._log_progress(log_result, tab_index=2)
        
        # 출력 경로 업데이트
        # self.lsl_out_path_edit.setText(output_file) # 에디트 박스 제거됨

    
    def _auto_execute_all(self):
        """모든 탭 자동 실행"""
        # Operator 이름 확인
        operator = self.operator_input.text().strip()
        if not operator:
            QMessageBox.warning(self, "Warning", "Please enter operator name before executing.")
            return
        
        # 확인 다이얼로그
        reply = QMessageBox.question(
            self,
            "Confirm Auto Execute",
            "This will execute all tabs sequentially:\n\n"
            "1. make DCR format\n"
            "2. make Form Measurement Result file\n"
            "3. calculate LSL USL\n\n"
            "This may take several minutes. Continue?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        
        if reply != QMessageBox.Yes:
            return
        
        # 버튼 비활성화
        self.auto_execute_btn.setEnabled(False)
        self.auto_execute_btn.setText("Executing...")
        QApplication.processEvents()
        
        try:
            # Tab 1 실행
            self._clear_progress()
            self._log_progress("=" * 60)
            self._log_progress("AUTO EXECUTE: Starting Tab 1 (make DCR format)")
            self._log_progress("=" * 60)
            self.tab_widget.setCurrentIndex(0)
            QApplication.processEvents()
            
            self._execute(for_auto_execute=True)
            
            # Tab 2 실행
            self._log_progress("")
            self._log_progress("=" * 60)
            self._log_progress("AUTO EXECUTE: Starting Tab 2 (Form Measurement Result)")
            self._log_progress("=" * 60)
            self.tab_widget.setCurrentIndex(1)
            QApplication.processEvents()
            
            self._execute_form_measurement(for_auto_execute=True)
            
            # Tab 3 실행
            self._log_progress("")
            self._log_progress("=" * 60)
            self._log_progress("AUTO EXECUTE: Starting Tab 3 (calculate LSL USL)")
            self._log_progress("=" * 60)
            self.tab_widget.setCurrentIndex(2)
            QApplication.processEvents()
            
            # DCR 파일을 Tab 1의 출력 파일로 설정
            if hasattr(self, 'dcr_output_path') and self.dcr_output_path:
                dcr_file_for_lsl = self.dcr_output_path
            else:
                # Tab 1이 실행되지 않았을 경우를 대비해 추정 경로 사용
                operator = self.operator_input.text().strip()
                date_str = datetime.now().strftime("%Y%m%d")
                dcr_file_for_lsl = os.path.join("output", f"DCR_format_yamaha_{operator}_{date_str}.xlsx")
            
            self._execute_lsl_usl(for_auto_execute=True)
            
            # 전체 로그 저장 (한 번만)
            all_logs = "\n".join(self.progress_logs)
            log_result = self._save_log_file(all_logs, "")
            self._log_progress(log_result)
            
            # 완료 메시지
            QMessageBox.information(
                self,
                "Auto Execute Complete",
                "All tabs have been executed successfully!"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"An error occurred during auto execution:\n{str(e)}"
            )
        finally:
            # 버튼 다시 활성화
            self.auto_execute_btn.setEnabled(True)
            self.auto_execute_btn.setText("Auto Execute All")
